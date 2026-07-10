import warnings

warnings.simplefilter("ignore", DeprecationWarning)

import hashlib
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import threading
import unicodedata
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from rag_core import IMAGE_FILE_TYPES, INDEX_PATH, KNOWLEDGE_DIR, UserFacingError, answer_question_with_sources, build_index


ROOT = Path(__file__).resolve().parents[1]
ALLOWED_EXTENSIONS = {".txt", ".md", ".docx", ".xlsx", ".pptx", ".pdf", ".json"} | IMAGE_FILE_TYPES
DB_PATH = ROOT / "data" / "usage_log.db"
INDEX_META_PATH = ROOT / "data" / "index_meta.json"
TXT_REPORT_DIR = ROOT / "informes_txt"
ASSETS_DIR = ROOT / "assets"
INDEX_BUILD_STATE = {"running": False}
ADMIN_USERS = {
    "USUARIO": {"CONTRASENA", "contrasena"},
}
ADMIN_SESSIONS: dict[str, dict] = {}
USER_ROLES = {"invitado", "super_usuario"}
ROLE_LABELS = {
    "invitado": "Invitado",
    "super_usuario": "Super usuario",
}
ACTION_LABELS = {
    "user_login": "Ingreso al chat",
    "admin_login": "Ingreso super usuario",
    "admin_logout": "Cierre super usuario",
    "user_save": "Usuario creado o actualizado",
}
ROLE_PERMISSIONS = [
    {
        "role": "invitado",
        "role_label": "Invitado",
        "description": "Puede ingresar al chat y hacer consultas. No puede ver reportes ni administrar usuarios.",
    },
    {
        "role": "super_usuario",
        "role_label": "Super usuario",
        "description": "Puede usar el chat, ver reportes, exportar Excel/PDF, revisar auditoria y administrar usuarios.",
    },
]


PAGE = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Asistente MADEVAL</title>
  <script src="https://unpkg.com/lucide@latest/dist/umd/lucide.min.js"></script>
  <style>
    :root {
      --bg: #f7faf8;
      --panel: #ffffff;
      --ink: #17211c;
      --muted: #68736d;
      --line: #dce5df;
      --teal: #007f73;
      --teal-dark: #00665c;
      --coral: #ef6f61;
      --mint: #dff4ea;
      --amber: #f6c453;
      --shadow: 0 18px 42px rgba(28, 50, 40, .10);
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      color: var(--ink);
      background:
        linear-gradient(180deg, rgba(223,244,234,.85), rgba(247,250,248,.96) 44%),
        var(--bg);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }

    .shell {
      width: min(920px, calc(100% - 32px));
      margin: 0 auto;
      padding: 24px 0 36px;
    }

    .topbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 10px 0 24px;
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 12px;
      min-width: 0;
    }

    .top-actions {
      display: flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }

    .local-clock {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0;
      white-space: nowrap;
    }

    .mark {
      width: 54px;
      height: 54px;
      display: block;
      border-radius: 8px;
      box-shadow: var(--shadow);
      flex: 0 0 auto;
      object-fit: cover;
      background: white;
      border: 1px solid var(--line);
    }

    h1 {
      margin: 0;
      font-size: clamp(24px, 4vw, 42px);
      line-height: 1.05;
      letter-spacing: 0;
    }

    .subtitle {
      display: none;
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 15px;
      max-width: 740px;
    }

    .status {
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 9px 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(255,255,255,.82);
      color: var(--muted);
      white-space: nowrap;
    }

    .grid {
      display: grid;
      grid-template-columns: 1fr;
      gap: 18px;
      align-items: stretch;
    }

    aside.panel { display: none; }

    .panel {
      background: rgba(255,255,255,.94);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      overflow: hidden;
    }

    .panel-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 18px 18px 14px;
      border-bottom: 1px solid var(--line);
    }

    h2 {
      margin: 0;
      font-size: 18px;
      letter-spacing: 0;
    }

    .panel-body { padding: 18px; }

    .control-box {
      border: 1px solid #cfe4dc;
      border-radius: 8px;
      padding: 18px;
      background: #f4fbf7;
      display: grid;
      gap: 12px;
    }

    .text-input {
      width: 100%;
      min-height: 42px;
      padding: 10px 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      color: var(--ink);
      font: inherit;
      outline-color: var(--teal);
    }

    .btn-row {
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
    }

    button {
      border: 0;
      border-radius: 8px;
      min-height: 42px;
      padding: 0 14px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      font-weight: 700;
      cursor: pointer;
      transition: transform .15s ease, background .15s ease, opacity .15s ease;
    }

    button:hover { transform: translateY(-1px); }
    button:disabled { opacity: .58; cursor: not-allowed; transform: none; }

    .primary { background: var(--teal); color: white; }
    .primary:hover { background: var(--teal-dark); }
    .secondary { background: #eef4f0; color: var(--ink); border: 1px solid var(--line); }
    .accent { background: var(--coral); color: white; }

    .files {
      display: grid;
      gap: 10px;
      margin-top: 16px;
      max-height: 320px;
      overflow-y: auto;
    }

    .file {
      display: grid;
      grid-template-columns: 34px 1fr;
      gap: 10px;
      align-items: center;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
    }

    .file-icon {
      width: 34px;
      height: 34px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      color: var(--teal-dark);
      background: var(--mint);
    }

    .file-name {
      font-weight: 750;
      overflow-wrap: anywhere;
    }

    .file-meta {
      margin-top: 2px;
      color: var(--muted);
      font-size: 13px;
    }

    .chat {
      min-height: calc(100vh - 150px);
      display: grid;
      grid-template-rows: auto 1fr auto;
    }

    .messages {
      padding: 18px;
      display: flex;
      flex-direction: column;
      gap: 12px;
      overflow-y: auto;
      min-height: 460px;
      background: linear-gradient(180deg, #ffffff, #fbfdfb);
    }

    .message {
      max-width: 84%;
      padding: 12px 14px;
      border-radius: 8px;
      line-height: 1.45;
      overflow-wrap: anywhere;
    }

    .message p {
      margin: 0 0 10px;
      white-space: pre-wrap;
    }

    .message p:last-child { margin-bottom: 0; }

    .message ul {
      margin: 0 0 10px 20px;
      padding: 0;
    }

    .message-table-wrap {
      max-width: 100%;
      overflow-x: auto;
      margin: 12px 0;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: 0 8px 22px rgba(10, 74, 66, 0.08);
    }

    .message table {
      width: 100%;
      border-collapse: collapse;
      min-width: 560px;
      font-size: 14px;
      line-height: 1.35;
    }

    .message th,
    .message td {
      padding: 10px 12px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      vertical-align: top;
      overflow-wrap: normal;
      word-break: normal;
    }

    .message th {
      background: #edf8f3;
      color: var(--teal-dark);
      font-weight: 800;
      position: sticky;
      top: 0;
      z-index: 1;
      white-space: nowrap;
    }

    .message tbody tr:nth-child(even) td { background: #fbfefd; }

    .message td:first-child,
    .message th:first-child {
      font-weight: 750;
      color: #063f39;
    }

    .message tr:last-child td { border-bottom: 0; }

    .assistant {
      align-self: flex-start;
      background: #eef8f4;
      border: 1px solid #cfece0;
    }

    .rating-box {
      display: flex;
      align-items: center;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 12px;
      padding-top: 10px;
      border-top: 1px solid #cfece0;
      color: var(--muted);
      font-size: 13px;
    }

    .rating-btn {
      min-height: 32px;
      min-width: 34px;
      padding: 0 8px;
      border: 1px solid var(--line);
      background: white;
      color: var(--teal-dark);
      font-weight: 800;
    }

    .rating-btn.selected {
      background: var(--teal);
      color: white;
      border-color: var(--teal);
    }

    .image-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 10px;
      margin-top: 12px;
    }

    .image-result {
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 12px;
      text-decoration: none;
    }

    .image-result img {
      width: 100%;
      aspect-ratio: 4 / 3;
      object-fit: cover;
      border-radius: 8px;
      border: 1px solid var(--line);
      background: white;
    }

    .user {
      align-self: flex-end;
      color: white;
      background: #1f7f77;
    }

    .composer {
      padding: 14px;
      border-top: 1px solid var(--line);
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 10px;
      background: white;
    }

    .button-bot {
      width: 20px;
      height: 20px;
      border-radius: 999px;
      object-fit: cover;
      background: white;
    }

    textarea {
      width: 100%;
      min-height: 48px;
      max-height: 150px;
      resize: vertical;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 13px 14px;
      font: inherit;
      color: var(--ink);
      outline-color: var(--teal);
    }

    .hint {
      margin-top: 12px;
      padding: 12px;
      border-left: 4px solid var(--amber);
      background: #fff9e7;
      border-radius: 8px;
      color: #64512a;
      font-size: 14px;
      line-height: 1.4;
    }

    .report-list {
      display: grid;
      gap: 8px;
      margin-top: 12px;
      max-height: 210px;
      overflow-y: auto;
    }

    .report-item {
      padding: 10px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      font-size: 13px;
      line-height: 1.35;
    }

    .report-item strong {
      display: block;
      margin-bottom: 3px;
    }

    .icon-button {
      width: 44px;
      min-width: 44px;
      padding: 0;
      border: 1px solid var(--line);
      background: rgba(255,255,255,.88);
      color: var(--teal-dark);
    }

    .modal-backdrop {
      position: fixed;
      inset: 0;
      display: none;
      align-items: center;
      justify-content: center;
      padding: 18px;
      background: rgba(11, 23, 18, .42);
      z-index: 20;
    }

    .modal-backdrop.active { display: flex; }

    .modal {
      width: min(440px, 100%);
      background: white;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 22px;
      display: grid;
      gap: 14px;
    }

    .modal h2 { font-size: 22px; }

    .admin-panel {
      position: fixed;
      top: 0;
      right: 0;
      height: 100vh;
      width: min(440px, 100%);
      background: white;
      border-left: 1px solid var(--line);
      box-shadow: var(--shadow);
      transform: translateX(105%);
      transition: transform .18s ease;
      z-index: 18;
      display: grid;
      grid-template-rows: auto 1fr;
    }

    .admin-panel.active { transform: translateX(0); }

    .admin-body {
      padding: 18px;
      overflow-y: auto;
    }

    @media (max-width: 860px) {
      .topbar { align-items: flex-start; flex-direction: column; }
      .grid { grid-template-columns: 1fr; }
      .composer { grid-template-columns: 1fr; }
      .message { max-width: 100%; }
      .status { white-space: normal; }
    }
  </style>
</head>
<body>
  <main class="shell">
    <section class="topbar">
      <div>
        <div class="brand">
          <img class="mark" src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAGAAAABgCAYAAADimHc4AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFP4SURBVHhefb0HWFRZnj7s/ndn4+zM7OzMzk6e6e7paGh7bHPOgjlgwgQoICCCGADFBIpIMqMi5qyYI4LknHPOOUezvt/z/m6douyZ/c7z1FO3qu6555xfTudUn6SkJOOKqiqvDx8+eLW3t3vFxcV5vXz50uvt27deCQlxXo2NjfJbSkqKV1lZmVzn5+d7ZWamy3VNTY1cv3//3utc0E2vmJgYrw8f3nu9ffvaKyEhQd8/NTX1B/0z9f35bMPxe3p6vN6/5/gf9y8pKZHrwqJCr8xsrX9VVZVXkmH/BM6/x+v169decQb909PT9f3zCwu9srOz9eMnJSXJdWdnp6yZ43P9SUm9/bnGxsZar0Xz1nt9+tuJXunpWv+KigqZZ2//BH1/w/kbjl9YWKhff5/09PQjL16+gGpFxUX665KSYrx7906uKyoq0NnZKdfNzU2oqamR6zdv3qCmphq73I7B1mofHGw9P+r//v17ua6qqtL3b2pqQm2t1v/FixcoLy/X9yksLNRfl5aW4v373vHb29vluqWlBXV1dXLNuZcZ9C8q6p1/cXHv+JWVlfr+zc0tqK+vl+uXL1/KOIZ9VCstLdGP39XdjoN+Z7HNJQCHDgXDzfmQfN/d043Ssr/fv6SkRA+/6upqdHR0yHVLS7N+/n1IuR8+fEBGRgbKyspkwgUFBcjOzpZrAi4lJUUexIUnJiagq6sLr169QlJSEnpedOLWjRCYrdiOVx8AZ+fDcN28H7W1VdK/sqoSqala/7a2NunT3d0tgE9ISBBA8D7D8fPy8mR8zouITktLw9u3b2V89u/q7pbxk1OS0dDQIH0yMzNRVkaEvUd+fj5ycnLkmgvn+CSU5uZmJCYmyvjSPzkZ9fV1fzM+kaj6c/2ZWZloamzG/Nn2yCmsQWvPO5gudcbF8zfQ2dku96VnpOsRTiLKzc3V909NTdWvn/AjIb5+/VqIqg/ZllREbFVXV8lNnAg/85oAKCouFuppqK9Hfl6ecAB/Ky4uQldXO2ytduP85adoan+Ju0/iYTLXHjW11dKHACgqKkRbexsaGxtRWFgggOQrPz9POIH3aeNX68cn9bW2tqG2tlYWRuohsAlcNT4BRQSyP+/hglpbW4Wj+FLjk6D4Pfur8dvaWlFQkC/r47M4nhqf3MI5qPVXV1fiwb0wmJtvR11LN8pqWrHJ+SB2bvPHq9c9MjfOhWNyHM6jvLxMxhf4FRXJs7h+zoXz53383Icyl4OxsUNERIRQB7EXGRkhN7GRchR7Z2dnIT09TWOtmhrMMl6LkMg01DR04N6TBCyc54jmphYkJMSjsUnrT2rv7Z8tVM1WUVGOmJgYueZCIiMjdeO/lfEJNDZSvmJvbfwMua6qqkZMrNafiwqPCBdq5/yjoiL1oiY+Pl4vanJzc4Qr2CorKxAXF6cbX1s/xRK5Lzo6Wr9+X69AODr5oq65G/nFtdi9Jwh7dp1Ae3sLYmNjtf6dnTJncjfH1/o3yW/a+jXxmpubJ1zB1oeKgwtWFEQkaBRYivaOXgoicCi78wtIgc1CRaSmqsoqrDR1xfOYTFTVdeBZeBrmTF+HhPgUEVWUdWRJPpfAIGuyb3t7m1Azv+Oz+SyOz2u+y/jt7cLC7K8fX8cBBDbnpfpz/ryX16REchT7E8AfzT8/D01N2vx5TRFEkUB9xXE1DiwVjmAfUjDF6Ymjl7HZ+SBqGjtRXt2C40F3YbNmJ9rbW2UcElfv+GV6riUH8zeRIA0NIl619WvwEB2gZCgXTVklMvQHMoxUxUWnpCQLYHt6eoSK6uvqsHrVDjx5noLaxm4EnbmPmdNskJycJs/iAtLT0/HmzVu9DOaCX754KbqFQBEZmp6mV1qcPOfAawJFyXCKMfZhf46vZPi7d2+RlZUl97IPAU4ZrvpnZKTLNefP8bn4Fy+6kZKaLBRORZuZmSHjav3zkZenrV8Tofm4fP4O1q/fj8aOl4hLzIOV1R5scvTCq1c90p8wItLU/ElovKZY5/qVDiUny/pfvtR0AE0wZYVwYmSbnhc9smCyVpNOhBDwlPkaC+WK0mJra2/Bovn2uHwjFJW1bbC22gNL8x0yqfj4OD0LE3AEMBuBo0QQEUygsLW2tog4InA5YYoNJUIyMtNRX69ZTvUNNSgr18RRaxu5MkeuP+AtsrLT8f7DO7x58/ojEUIAKRHG8QlwNgJYP34Lx9fWr80/Xi8Cb16/jyWLNqG5+zWC70Rg3twNuHHlCbq7O0SxstG64ZgUQSSK2NgYvbWjIUiJwFxBClsf2tGKHXgzFVBdXa0oR2KI33ESnCixKUqpplp+Z5+u7g4cPXQeTht8UVBci9kzHbBn11F0dWmihCzI+9hPe26tPIuAr6mp1Sm5ajQ2NshYJAbqlYaGejS3NOHlq260tDYhLiEBN27cw5nTN7HV2Rd21u7YssEPdlZ7YGWxA3vdj8PfNwjHjp1BbFwiaqjcO1pkfiSi3jFrxGzW5lUn6+Ca+ZkvXnO9HJ/z53M4L8512aKNCI/KRFp2BaZOtER8bLKIQzV/9uf8uU4+g8/SRFgNqqqrUFtbJ8/SYKmZpX3oDBATVBKKQiin0vVKskIogY3yMTo6Cj06JUMOIYW9e/ceK5Y4Y8VSN6xa5irYzsnJlj6cUK+S6xAKJ/vRrKTCVRROcaI4pKq6AoVF+cgvKIGXZyDmGjtg1jgnzBm5DYtGe8JiWgDsZp+D3azzsJ97GevmXIDFlACYjveBydjdmDJkPWZNXg9P95O4c+cxUnQKl0o2OiZaxieFk0KJBDaKBjU+15+apilJrl/NPy4mDbOM1mHR/C3Y7noQeXnZegqnSawonEYGxR4bAR0XpylpEnpUtBr/vYzXhx5hTW2tWB6Ux5S3HR2dwkZkEw5Am5Vsy8m8fv1GlArZiN+TWmvrqnD+7B306fMl3HcekcFoa1PM1NXVy328n/0p2pSM5/ekSooziox3718jNy8Hp4OuwXTBZswYsRlmUw7BdelteJg/hadFKPatDoOXZTh81kbBxyYGvjax8LGJxn6rSPlt7+oQ7F3zFNuX34XV9ADMHe2GudMc4L7rkJjTbGXlZQJkw/G5fpqmXDMBRV9DdFx9vcw5Ny9XLJ7F85zwoz4DcPXSfbx7/06oXPWnjuGapX9XlzyLnMD+2TnZAnCOaajj+jAswEEIMFIzlRllMamVQCMC+DB24GR5HzFNy6Sru0tj6doqnDpxHT/556HYsztABtT65wiFqf6cLK/LSjU/g8/iM0vLSvD27RvcDn6GmRPXYc5wN2wyuQJ3i0dwWngZFtOOwXzKEVhND8SG+Zexbekd7F71GHssnslrt9kTuJnehdP8y7CeEYhVUw5i5eRDWDf3DHaZPcC25cEwGe2O2ZOdhFBqaymCqkUZUqxSaXMuNBJIvXwnt/NawYZWWkNDHVYsdsZ//dtwnAm8gZevXgqFE5liZTU2CszYl0gg0RJ+Wv9ClFeUa+sv06wkIkMvgqiIFAsSCcpOJYAUC9J0pNhQdjLFiVJSPl6B+PE/DkbQyZuor6/WsyCBruz89rZ2hIeHC3exP5/79u1LlBRXYfoUS4z8eg08Vj+Sl4XRMcwZ6Y5Zw3Zi/qg9WDbRD2uMjsNu9lk4LbgC58XB2Lr0jrxcFgdj44IrsJtzFmuMA7Bskh8WjN6LmcN2YtawXdJ364qbcF/zEPNG78KsKbZ4HhqH/PwcUZyy/qREfRiE4jNNJ4K4fhoTbC9e9GDWtLX45X+OweUL95GTmynIkv6JiYJINiJEGSlcv/ITiPCoqCj9+omEPgwQKbuewOSX5ASKI4oaYpCsRazxYfye71VVlfI9lVRDYx1OBlzDj/9pMDx2HcPLVy/keeQUUpC65mLYn3Y2HbBXr17g/JnbMDXZgiOHr8Fk3A5YzwzErOG7MHvEbswfvVf/Wjx+P1ZMPog1RgFYOzMI6+ach8O8i/Kyn3seNrOCBEG8Z/G4/dJn3qg98pozwh0zh+3A8skHYGV8HItnOYvvsmPrIXR0tqG6hvNqFCVKJPCacyZB8jteU5fRKFi+aAt+/u8jEBR4Q3QKFSxhQ1iQ27X+mp/CNRN+hC/vIbfwWcrTJmeIGUrZzEbWIVUSQ5T1hmYY5ZmhkmJ8hI0WRWlZEc6dvoX//KfB2OK0T76nkiVVcDJsomRLtf6FRYUoKS3Cnp0BWLvaHU3NHSivacLY79bA+Hs3zB/t+RHw54/eg/lj9mLR+P1YNtEfZlOOYLXRcVganxSxtMb4BMymHsGySf5yj+qjIcBD3rXvPDH4Dxbw2nseb95/wHbXACxbtBnJKZqSJnC4Zq6f849P6DWDqQ/q6mqwZtV2/OxfhwnBsdHRowJno58SFx8n+oNimIZNrxmcguISTQeRUzKzsuT6o2CcFn1kMKlA5DeviS0VjNOCSYno6qIj8UIGVgg6cuAC/uMfB8l7Y2ONPphWXV0j4owLam5pFufn3bs32Oy4H3Zr96CotBY5RZVIK6zAolnOmNDXCSZjvX6AAO21YLQnTMbuw+Lx3lg26QAWjNiHGYN2Yck4X5hO9JXfDO9XHDBvlPbZZIwXhvzJHIFnHqK4tgndr9/hoN9lmMx2QGoKDYYavP/wQSLChsFIzp9W04uXPVgwex3++z9G4tL5+0JE9Q0N+PBBC8aRS7hmCWZmZemCibW6YNxbvSNIPWAQjGv1IsaViOBNBDpFRLsuGEWx1KILHlEsaSKrWa7Jot3dXThx9Cp+8qOh8PM+jbb2Znken0WbmJwj1/V1aGyqw3aXA9jk6Iv65m4EnrmHyMRc5FXUYd/eC/j+92ZY+AMELBizD/NG7MW079wwb6QnFo3bj8nfbcaSuW5w2XQMRqM2YObQnTAZsw/zR3lixvc7BVkfPWPsPhj/dRsmD7JDfHoxknNKEZ9RhNbuVwg8fgvzptvLXDs7tVCCmj/Xx2sSH8XVskWb8PN/G4Wzp24JwpSIZijDMJinAnPsrwKbWjBQE1HkNn0wTtmv1ORUEioYR69OiRDDYJz4CTpPjoMVFOTgwpk7+Nm/DofrZh/5/u3bd/IsxYIUQc3NDbhy4QHMlruhres14tMLkZBZjISMIqTlliEqPhdj+1nC6DtXPRcsGOOJWcN3Y9oIR1iYemDqUCcYfbcNSxZsRUFFLbpef8CdhzGYNNgec4d7YMpQRyye7YapgzdhwaheJPB5f/3NCjhvOIbi2mYkZZUgLr0QMan5aO7sga/3RaxZuU08buWJ03hQ609KSkRtXTXWr/XET/95BI4cvKhbfwVidXa+gt/fM1KSDYKJhJ+KBPTp6upi9kgwY2hm8mayCjmAJiSvORmyJsPE/MwHkQI4IJMVVMKH/S9IgKqwkKHqLnkmWZImW0ZaLmYZ2SAsPA3nLj4SCoxNK0B8ehESM4pRWNWI7S6B6P/LRXrKXzB6H8YMsMO5y4/R0vMaBw9dQ//fmmK3RxCqmzpQXtWCzOIqLJrtgmF/scbBI9fR0NmDra4BmNB3I0xG78PCcfsxpZ8Thn2+AuGxOcgurkFSdokgnuNHp+ShtqUDlma7cdD3HNramlBcVCSw4PoYtKMF093diZVLnPHf/zZKzO7KyjK9mUnipEgxNDP5PTkgL0+DH71mmqbkDD6PffswdUdsE7BkGyKDD1OOEtmKZunrV6+EbcTRaGsXKmEAr7GhQeS7v/dp/Mf/G4TDBy5K6IBAZ39SCO/jta2lOy5dfIo9e07j3tM4ocC4tAIkpBcKAtLzK5CSXY4Zox0w7A9rhPpNxuzHmAFrcf9pLHo+AEHnH+C7P62Cqcl2FFU1oPv1B9x7EiscMOab9Th49BrqOjrh4nwUE/o5YdE4b8z4fiv6/twE/r43UFLbgpTsMuEAIiAurVAQwHnkF1dj/gxHPH0SjspKDRZcP7mdRPbq9QusMnUREXTy2DWUUgfoHLW83Dw9/Ah8tX4SMNdPqaKHX7sGPxJnH+YtDfMBDDVwMCLlhyJEsZBhPJ9mWHFJAc6euo3//Kch2OIkboUoJlpU9Q2aFXH4wBlssPfByZO3ce12uIie6GRt4fFphUjKLEZydinyK+rxODQNw/6yAsP+aI75ozwwe/guLJy5FTu2B2Lm+E2YPcwdkwduxrIFO7DN+QSMRm/AjCE7MH/UXkwZugEmM1xgNGQLFo/zxvRBzvjqp3Nga+6DkpoWZBRUITWnvBcBOjEUkZCNzKJKXL4SAmvzXTJn5jP0oZKkZDQ11WOzozf+80fD4Od9Rr7X/AQtVEM/iTkIlQ8wFEFaMFOJoNxeEVRfX89ksi5opAXhVIBMBah4zQdxMhKg0gXjKP+1eHo7Thy7Jo4Yg2KdXVomi/ezHye5Yslm+PpcxOGA60jKKZEFK8qjLkjMKkFKThlSc8pQWtuCGzejMPiTZfj+NysxZ9gOzB62GxP6bcKc4R5i7dCiMR60HeP7btZ/R45ZMGofZg7ZjfkjPTC5nyO+/tl8WCz1QEFZA3JL65CeV6EhILtERGBceoEgICopF2FxmSiorION5V7cuvEUzc2NMnct2NYogT0G/n76z8Nw9NBF8QsoRrhOdZ+CnxZ8ZLCvVn77ONip9aEo0nvCNDUNw6WGniztWTYqmdi4OGGf129eSzCLUUM2KqUf/78hOHrwioSLlSfY0tKEK5duwcpiN3z8L+JhWCIik3IQmZgjCIhN1RBAiiQC0nLLkZ5bgbL6NjwNTYPxSAd8/bN5GPO5FWYN2SZ2PXUDlSpflO8m43hNfbEXc4bvxNQBTvjrr03xzS/nY9vmk5JCLK5oQkZ+FdLzK5GaqyFA44ACxKTkCQJIFHEZ+bh1NxIrlm6R+ROo5H42iiDTRRvxX/86AoHHr6O4ON/ASOnN+ImflKn5SexPTmKjhKFXrJQ8k/l9WO5BDFGOMwZkmLTmNbFFdqJrTlFF7c6BlJ2smVzFYhf/7F+Gw3WLLz58eIdXr1+hoDBPBqaF5LYtAE+jUhGemI3whGxEJuSICBIEUAQRAdllSMspR0ZeJTILqlBe14bcwlps23wCQ/+yAn3/ywRDfrsK476wxZT+GzBt4CYYfbdZ3if1dcDoT60w8H+Wou8vTDB7wkZcvxmF+tYelFQ0IbeoVp5JPfMRAtIMEJCYjbC4DCTnlGDpws24fesROjraZM0SWGusxUYHb/znPw3HQd8L8j1hl5ycJPKeFK2KDnqT/r1FB8rPYmqSypj9JRRBLKqUmZY0bpaH0TJiDFtdE9i85qAURXJNlqqvwZlTwWKGMhTB1EhLawsO+AfCdbM/poyxwPHA24jPLERYfKaGAD0H5CM+rehjBORXIquwWqyVksomNLT2ICGpCJ47L2LuhM0Y8ukK9P/VIvT9bxP0/e8F6PffJhj42yUY298SlqaeuHQxDBW17WhofYGi0gYUFNcjr6gOmQXVGgI+0gFEgCaCyAHP4zORlFMMx/U+uHjulqylrb1VS6y3N8PJfj9++qOROOx/UfwfwoSOq0q69/pJWsqWIkhL+vf6WYQdiZmc8FEw7u8pWXbSB9N0SXsqaXz4gNiYWL2S8fcOEiV8NvA2KqvKYTxpDRZO8MDONQ8wdZgTjp++hZjUPDyPy0R4fJZQf1J2MVLzy0XsUPykZJcKAigmiICc4loUlNahuKwBNfXtaO16jaq6diQmF+H27VicOx2C0ycf48rFMDx7loG8wlq0dL5Ga+drVNS0oaS8CcWlDSgsqUNescYBaXkVguCc0lrkltcho6gKaQUVSC+oQGJWMcITshCfUYjduwOxaPYGHD1yVmL4bB8+vMXqFVvx838djWOHrkgwTp8PSO7NBzCKYAg/FcxUwTjlJ0gwTh+O7uoWeUaTUyWNiQgCmJiiWUqs0Z5lapKd+b0omeoKnDpxEz/90RD4ep2BnfVuWM0KxC3vOgR7V2PJVE8cCbyB6NQ8objU3DI8DEtG4Ln7OBZ4GxeuPsPz2CzklTWgqKpJKF/jgFoUlGgIoBgpq2oWBDS2vkB7zzt0vQJ6XgPdr4H27ndoaOmR38urW1Fa2SIIKCqrFwQUlTegtLZVnh8amYkDh69h40Z/2Nnuwzo7L2xw9IOX7wU8i05HZkkl3PcGYsZQVyydtB/Txq3FrVuPAbyDy8YD+Mk/jcB+z1MSjqGMJ2wICxZ8EWYKfhQzCn6ULDRTDf0sCUdnSkKm5m/sVAKa13yAllDJkQfwPg7Iz69evRStX11TgfOn7+J/fjwKUydaYNkMD1z3KoWHxTN4mD/DjGHOOHXhLtKLKhEQdBtmq7ZjxRIXrLf1xBYnH9iv3SufLVbtwn6fSwiPzRaTsaqpE5V1bSivbhbgE7CkbAK5ur4TNQ1dqGvqQV1TtxQE1DTwuw7tt8ZO1Lf0CLKYq46MyYaf/1VYrfaAxYrt2O12DFcvP0Z4eDIiI1MQfDMMO92OYuXSbQgIuoO9vkEwm34Aj4+04eDmWIzob46YmHh47QnCj/9xKLz2BoreJAES0L0JrVQROQQ05X5DQ6PAj/cQ8KoChQhhf70ISkruTcnxZpUPoIxTdq5iISatmUChzausoA32nvjz/07BV59Oh9vqYHjbhGGraTB2rXgM42GbcPjENVhZekgs5eaNEBSVVaOirhllNU2orG9BeU0jIiJTsG9PIFavcoOlxW647w7ClauhiEvIFzleVd+JxrZXaOt6g46ed+h88QFdL0n979HU9koQUlzaiJS0Ejx6koSTgXexzfUoVpvthLXFTuHOmJgM1Ld0oL69C3nldUjJLUVCVjFS8kpRVtuE7NxSuGw+hP5fzcOSiXtx2DEBF3YXYLf1fSxfvBkOdnvxs38dgUP+F1BYxCI1LR9AS4cyn82waIHBSJUPIGEzn6L8BBFBLC6lYqACoSgilTMeziQzRQ7tWV4TEcQ271O+AeUffYCAIxcxbqANHNcdwvLFbnBZEozNJtfhuuQWdq14hCnfb8B3/ecg8PgNVNQ1ITG7CE8iU/AsJl0UM98vXA/FxeuhiIjNlghpckoeThy9js2OPrC22IXVq3bCysIdttaecLDbD0d7b1GUfF9vux+2Vnthvdodlma75P4N9vvhviMA587cRUZmMbpevkVpVRPuPU3AuashuPs0Acm5ZUjMKcHz+CyZw6OIZDyLTUdhVT22bPTDpIGO8FwTih0rHiDILQMzx23CkO8WSULm2OHLEh0lHAgn+kOEB+Gn8ikM6knxgs4XItwoUQhTfiZB92FlMrXy+w/vhXVUPJzutWFZiFZWoikZlRNmCwuLxoivzbB95T2st/HDSpOdcDK5Aoe5F7HB5Boc51/GiH5muH7jCXLLqnE/LB4PnyciOiUXj8NTsGGDPxbN3whL8+3Y7OAt4QqTuY7YsMFPFOcHAI3NnSiraEB2dikSEnIQFZWGiIhURESkISY2Eykp+cjNYTFWHWrqWtHS/gKdL96g+9V7vPwA3L4fi+XLtsJ0sTPWrd0jFRWrlm3F0sUu8DlwBUk5pYhIykZIdBqeRCbjYXgScksrsGj2RlhMOSr55X2WoTCdtg+Hjl2FycwtWDDDUeL+TG3SgKHnTz+JPhNlPnPchJ9S0hTnSknToqSuZdOLIMOEC2U8lQkbMUzbVhU2MT1H/fD2zRsUFRfAbs0uWM4IwMIxXliywBkLp2+F/ZzzsGPGasEVjO5ri4MHziOntAp3n8XhUXiSWEOnLz7C5PFrxJxraWqTsVR70fMCl8/fx5zp9jgZdB+1zT0oqWxGdV0H6pt60NT2Eq2db9DW/U7ET2vnWzS2vhQ9UFnXgbKqVjFfKbJcnY9JHvfZU00MGLbkxCzYrN6NVSt3IDwxR0xQcubj50mISslB4OmbGP+tHVxNb8NlyS0sGrsP3n5n4edxHUvHe2PVss3o7ukUi0YV97Jp8NMArBw5ihyKIOVnUS/w/j7Nzc1ePd3dImJUyIEcodKPxCA/a9daoawSRWSxZXO2wsL4qMRiFs/bglkTnGA9I0gQwETJjMlrkZxVgIfhiXgUkSTsfv5qCKaMW4O0FI2L3nR1IPRRFM6de4B7waFoq9O4rr6uCdMnWeHS5WeiTEsrm1FR3YrKWk0J1zZ0o67pBeoae1BT3ymlkVTSZVUtgiRHB1842vWWyxdk5IrivXzpEVJjNTORzc/rNObP3SBWWkhUKh6HJ8lcw2JTMWOCPSyNA7HB5IoAffu2Y3DddBjWM07DaPhGkQBd3Z0ieggzwoXvfFE/KLFE+FGakAuUiOLvfRKSPg7GKSVBitfyAVowjlygL67NyUZ2diZa29phPNoei8f5YO4IDyyZ74rp4xxgPjUAtnPOYfwAR+z3Po2YtDxBwOPIZITGZcJo6lpER2ru/YO74Ri0xBv9dz3HgEOZGLAvBqMtj+PcqWD5vaK8BnOM1yO3sE6smYqaVlTpEFDXqCFALCADBJAbbt6MxLKFm+QZL9o7sNHlGAY5XkF/32T080vDAKdbWON4GA2VWrUdxd+27ceFO8mlfEWn5cJipRvmDfeAw/zLWDx2P+xtvbB7R6Akd1ZMO4BtW/xRUKgREiUE4afqpmj/9+YTkv6uku6j2w0j8QtSOp0EiiKyEa+pMCiOlJnKjpR1lHGhoREY1n8VFo7dj7kj92Lp/G2YPnY9lk86CLNpxzBhiC3uPIpESEyqRv2JmdjnewEb7LWIaWxEEhasOwz7yFYM84nHIMsgjPB4hpH3OvHJ6rO4duG+3OezLwh7954VJ0shoKa+S0NAc4/eBOX3lTXtqG/ugdny7cjKyBdPds16f3zllYSJN+oxzPEqBtmcw5iLFfjmeAmmm/vgVVcXWprbMGeGPR6FJ+NxVLLMNyo1F5s2+2Pqd1uk+MtklCdWm+3G/n0XMO07F8nc2VjsxZu3NDOzJYvIvIeCH2FGaUGCJSyZRaQuoK5VPpQkZKhwWfVMtuEPZBmVUKBYInJ4TQSouhfKsJCQcAzra4bF48kBe7B43lYYjVuHpRN8JREyfeI6PItOEcrngiJTcrBipRtCRR5/wNw1+1Hcpu3OGf7dEnzyX2Pwyc9G4U8TXfHX252YYnsSLzs6kZtThBWmLqhveaH5AbWarU8EENj0AzQEdAgyYuPzYbbcVZ4b/jQafZ3v4/tTRfj9n+bii1+Mx5e/HI/PP1+IWRfzMWp/OAJPatzmuM4T/kevIiwhA48ikxCRnA0PzyBM6O8A65lBmDvcHavNduHQoRsw+utWzBuxB/ONHNHR2aGlJGuqBX4MslHREmYk4Lz8PPme8FMJGVZUSEKG+QC1xYaUTbFD4JKFaPMrK8iwdI8KhoG2vNxiDP3aHEsm+GoImOOKqWNtJQkya+guzDbegGdx6XgcqcnU0LgMLJzvhMryWjRWVmOw4yX4572A/74gfPPnmfjs9zMwbOBiHN5/Bocz2zBk52PEhCdJQtxs2VYUljYKlRPQFDl/I4JqO9DY+grXbzyHs5O3zHX7vosYfLYUa3zuYPaYVfjst9Px1Z/n4Ms/GKOquBTNb97BxvWE3MuIrvO2IwJ4zpeWkd+hSxj9jTXMjQIwY+gOWK/Zg+PH72DKQGcJfU8abon4BK2496OUJFO6BvAzLL38KB/Q2sqkfI9oZNqmxJIKFvGaGFRJZSoPYlarduvEw4ehGPb1aiyZ4Id5I/dg0WwiwA4mY/djxpCdmG3khGdxaYIAvp5Fp2Hxgs1orG9BYUYuxro/xsTrtfj1Lybjqz/OxB9/NRUhT3qtlR0XI3Dl+jO5Xr1yO/IK60TOKwQQ8OIJCwK65Hs6ZJcuh2DH1sNaP9dT8Imp0j+Txby//+Vk/OU303D23AM8aXqPpVtOyW/Mcm12OYSIlGw8jU7BvZA4OG3yxeh+Vlgx9QiMvt+GdWu9cCLgDiYN2AKTMd6YNtoOVdXVshVLaoeam0WSEIYKfgzdKF9LgnVUwC3NWlJelaXI/oCyMql3lP0BOVq4mQ8gtrRwtVZWoTbbMc4/5AszHQI8sWj2VkwdYyt53DkjPIQbHkck6WUqKcp0iQvSU3LR2dSMyRvPYnEy8P2ULfjsp2Pwx5+PR1KUppzZ/E/dQWxMBlqb27B6xQ4BcqVOBCkE0CzVrCANAfz8LDQNDjZ75Rk7Pc8iPE1TfmxXz97BL/9lCD792RiMdr2FcffbsH57kPzG8MTe/edEVz2JSJbQtNuOAIz4yhLLJx8Sql9v54Njx4IxaYAzFo71xaThNlJ6kpObLWJbCzcXidjmNWtn6UOp/RWyx023x07KUj4qzDJIGEhhUnz8301JkgsqKsqQGJeOQZ8tx9KJfpg3yhOLZ2/DtDG2khqkZTR8gDkuBYcgNDZNrIrIpGxsdjksZh+b844TmHqjAsvD2zBs7k58/tVSrHQOQv4r4FlZO6wcD4iNffn8PWzZeBBtXW/FDCUXkPKbWl+iue0VGlteoq6hWxBAE7W8qhWmC7egtbUDGWn52OJ+Gj1MDr15j+V2B/BF3+WYuPkSzPOB4dvu4tnDSJnPonkbcONuJEKiUvA8LgOnLtyHybxNmDjQEcsmH8TEAZuwYb0vjh0NxuRvnbF4vB9Gf7caISHPpT+tSFo+PT3dYkXSKVPRYoZ2lBVJAtc7YtwfQPGiUmZEBq/56k2f1Yl4IqbVfa1tLXjy6Dm+/2KFIGD+SE8sme0Go/F2Io444SFfrobbrgBEp+SIc/M0KgV3nsRhlpEdmpta0VLXgLnWvph6vQwWeYB56gcsSgHGBVdhyCo/ZKbmCgLmzViHsPAMqSOqa2IArlvi/UUVTSgsa0RZdRuqGzrR0PICtQ1daO16Cw+PM9jtplVqHz8RjIkuF7A8ugvLE9/DKhMwywaGejzBth3H5Z5L5+5h2bKt4gs8fp6I8PhMuO8LwqRxVjD+3hWmkw5gQn8nbHLyx7GjtzB5oIaAkd+aISkpVZJZKv3K8ANrhlQqUoUfVNqXKV1eswpd7wkbKgnDHRzcWGAYjFM7WNiuXbuNoX01JTxv5F6YztuOGZPXSS3m0om+mPa9C4ynrhNWJkuLaZeSA3fP01LewdbV3AKXHccxfctZzNj3ALN2XIOtcwBK87S50D533xWE9h7GcloQnZiPE6fvY9OWw7BY4wFrGy/Y2Xtjn89FBN+PQWZ+NarqO8QqWmKyBQ/vRchzbl5/igUORzHb4y5m7b2PRZsDcfrkTfmNDuG0ida4FxKPkJg0hMWlI/hBBOwd9mPkX82k4GvpRH+M7+8E5y2HcOToTT0CRg00R2VlDfLycvSesAa/3qIFfUq3U9sfofwswrsPd2wTG7R8qDTIKtTmZCNeE4NSH6+r7+d93ENVXl6KouIyjBhgIVYPK9dWmLhjjrGjVDQvnuAtiPnrX1bA++AlRCXnCAJo3sWk5WPj5oMwXbgZRQXa9qh3L7pRX16Nt12afmmob4KlmZvE6bkvKzopH65uxzFtii3Gj7GA6dJtWLt2P4yn2sPc3B3GRuswbsxqLDRxhv/hm8guqpE05MJ5m3DI74I8k629vhFNVbX6z9cuPcLEMeY4fy1EnDAq3ztPY7DTIxBLFrtgbP91MJ3kjyUT/TCunyNcXI/g4OHrgoAlRMC3a/DsWYTAhSKIuXRKFFpC1KsKtqpuiLJf6QjZH8Bt82QRWjX02tQuRj5EFV5RacgDKisFOVqUlNtsajFy4ArMGe6O+aP2wXyxJ+bPdsL0wW5YPMFHEGA8eBvGDl+N20/jEBavmaRPolIQm1GAA0dviIzduN4b58/cwYO7Ybh49i62bj6AZYtd4H/oGhIyy+DueQZTJ66F0eR1cNt5Co9CU5BdVCt55k0Oh3Do8HWk5VXi0vVQ2Njux/gxVpgx3QFHTtxGZEI+Nm06hNUrd8J//1ncCQ7B3eBnOOB9FubL3LBqxQ7ceRyL2PR8hMakIzw+A5eCw7B8hRsGfWUq61g0wVfE7Lh+DnDdegR+flcw5dstOgRYITycGUPmhyvFYCGQCScqX1pFpHi114C/0ewn9fO6Nx9gYOd/JIIMUmrKT2BHtsSkJIwdshzj+zpg0TgfWCzeh4XzN2LqIBcs4cTH7xfqoS6YOWM9bj2Jk3Avo45cLFN/USl5Umy1a3egZKhcth7F8cB7eBKWigNHbmDKJGtMGL8GO91PIyqhAGXVrRLnKSyug72ND65eSYSL02H0vAIaml6guLJFELRh02GMGmEuiDgccBvXb0fB58BVbHI+hE2bD2KP5xncehgjqcjY1DyExpI4UnHu6lNYWLhjxBBTjOhnjqVSce0jImhsv/VwdT2M/fsvYspAFywZ74sh36xCcbFmxNAHoOPKRvhR2bLRaFGbAqW8patLLCQRQYoDKN81DsgTc4nsRA4g1viblpTXSu/ISiXF2g6XNSu3o///LsOM77djpYmHhHwnf7dFqJ8IIOVMGuSE4YOXY4aRPfb7X8HdkAQNCbEZEhuKSs1HXEYRnoSnCtVaWu3F1ElrMXP6emzfFYhnkVmSYqTpWV3XjtrGLqRnlsLK3BOHDj6Fk70/Wjpe6wJ0XWKeFlc248nzNLhsPS5iavIEa1it3YejJ+/i/rNEhMZl4VF4CoIfxuDijVB4+1/BKrNdGDPCHBYWHpIVG/NXKywax3VoCBjTdx22bjuCPe5nMX3QdswZthuDPl+CyMh4kfWq3JAEzCItTYrky2ZtXjMLJkZNQz2qaqo0DtDrgDcsq6AOSNPvw+VGPSJH9jjpShd5TXGkzlI4deIGPvv5XIz4xArzpm2RdOPEb50E8AoBY/utw8YtB7HLIwiTx1tj3hwnWFrtwXpHP9ja+WDZsu2YNWM9jKbYCtDNzHfB5+A1hEZnIr+kXkxPRjjLGA1lLKi+XQC91sIT82duhZ/3ZcmMUfmKGapLXfK6sKwBz2OyBPErV+7EDKP1MJpkh5nG6zF3jpOMN2WiNYymroO5hTsOHQ9GWHwWbtyPwKiBa/QIICcP/9Ia27Yfw/Ztp2A0cCvGf+2AwV8uRVhYFMorysR0J7wohujckljJCSRoiiHG0aRgq75OtiuJDvj7hVk5Bik1FhZpWtzQCiLw09NTkZmZgyH9TDHot2aYN9UFllbuGPaXNb0ImKTJTnv7/ZL6uxwcBrddJ2Gxxh2mptuwYsUOWK/dC1e3AAScvifJ+oTMEmQV1aCwvFES8kWSlG/UEKALRze3v0ZERBbcXI6jrKIFDS0vBQEMVZRXt0nomkn50oomlFY2Ib+0AUlZZXgYmoyTZx/Ac/957HQPhMe+MzgaeFs4ISwuC1FJOSKSLgc/EwQsnuAnCpgi9a9/XAHnrUex1soLg/+4BiM/s8GwfqZoaGRa8oP4UEoEUQGrpBVFjYIndQL1LYEv+QDlB6iUpKEfwGu151WVKvJ7YpGb23jd3dMBf58g/O4/psFozAah9H6/WYjp32/D4oneWDbZH5MHbsEaCw8kZhVJ2YdWAJWFkOgMeY9OzkdCRjFS88qQWVSlVUUUVSOnSKuK0COgigjQKJtOV0vHG/S8gQC/mp6w4oAqckyz9GFfPoPnO7DMhc9laUpyTpmUobA2SEoUWaCVnCvzY7bu0s2nogNY4j53hDum9N+E/r9diJ17TsF85W4M/oM5BvzvcpiZbpN9A8pnInVTbxJmhvBsaNT5CNyy1NIsL+5F+MgTVp4cT09h0p37W5UnZ7jTnVs203UYleLc4gJcOn8Pk8etwYihyzDgjwsx5gtbTOy3QRAx+qv1WLpoK+LS8/E8PkMQEKMrzCUAkjJLkCyliaVIyS3T6oIKqpBbXKuVlOgQwEoH6gFJyNTR1tfkPXUCryVKWtcpjhrFkFaWoqsLKqpFRkEl0vLKpf6UGzSIACnO1RVmRSawai8L0Wl5OHf1MQZ+ughjPrfD6M9t5cV1HTh+A5bme9H3F0vxp18YIT42HfUNtXoK5+Ek8fEJIo4odghPwlXt/Kcera3jTn9N2nwUC1Klh9pZC71nRRD4fCA5hbsJKYpYesjvG+o1BFVUlCIvPx/Xrz5Gvz8vwJgvtImP+twW3//BAjOnOiIyOUc8zN7SxFytMo7FubrqaNYMpefpEFBUi8JSIqBeZH9MWhGyShvQ8RJoan8lQTgVkKNI6nwFlDZ0IjK5UJy20opmPQfwWQoBrIRQlXEsjWSFnipNDE/IFEIJOHUL/X5vgrFf2GPMl/YY/cU6fPfFYmzdeQJjhppj+IDlOHv6hjhghMfrN29kQzjhRRgylkZ/ideEqyrl5B4BiipKEUGAnBf0d7coaaWHFD2MAfGa3GC4C5Df87vWNm2LTn1DHd5/eIM509ZjwK+XY+xX2uRHfGaDsUPWSL6Vi4yIz9IQkKRDQMbfIoBVbFIZV1KH6oYOPHyShHn/OxVW/ZbAc3cgYjJKUNP+Ck0971Db8QrxmeU44HMZtmMtsegnY+DnHoS61pdaaWJJHXKKapBRqENATi8CflgbSgQkZBZhn/dZ9P21CcZ96YjxXzlg+CeWGPD5XKxa5iKV0aWlZXKMAuFB8d3SQhiUCGAJGwJdRUQJV7VlieKIfZR/9YPzgrR9wDSPtHxApD4Yx7oXFUwyLL3jgKySZiMiM7PScfdOCL7+/Tx8//vVGP+1I8Z+sR5//XwJrt55LtTG0kSpDWVxbooBB2QZIKBQE0EEXn3rCwQcuwWHPkNw+F/GweEfhsH8t8awm7gWW0xcsG6CDVb/bgYc/t8weP9oFPb3GYr18zajRhDQiIJickBNLwcQAZn/BwJYG5pdAls7L/zpxzMw8Ldm+OZ/FuPz38zChfO3ZZ08y0g75kCL9efn5evrqEic9JVI+YQjzw+i5KCSpgGjYE3CpaT5aIsSKZ9YISYJ7O6uLn1G7OMtSszodPzNTnA+nNc8QeRM0HWMHmSGL/5nAfr9aik+/cUMHAq4hYSMQpGzgoAkVZyr7ZBJzi7REJCvIYBym/Kb9v29p0mw+OkEHP3PyTj8s6nw//cJ8OgzEm59hsm737+Mx5GfTMLRn03Gpn8Yht0uAWhof63nAEMd8BEHpH5cnBuRmCX7BGYa2WPx3E1YuWQr1phtQ2ICZfwHca5IzWpPBWFAAPMzrR7CizAgDAkP5TepDJmCIQ0a8QO4Rentu3f6nLBCBhUtbyD7yBYbXfk1lQ2VipTepfMIGe0sCQ6uSu+0qGqphC5OnbwmpYeTRq6GnZ0vYlILEJmoq45WOuD/QACplggoLm8U2W45ZR32/sNIeP/7RBz75VTc/HIuHn67AHf7z8e5P87A/h9PwJF/G4/lv5qK0LhcCchRf4gOKP4/EKDjADWf+IwC+B+5iXmz1gvAuSmjuqYSb9+9lviY8oe4TgKbkoDESLufsGHwUouXFejrP3k/4clrJmPYh8gj3D/aoqRCDeo0k6hoblHqPTLs74UqOAkVLWW5Nt3xmmpNwTA2kp2ts5aqG8ThuXInSiKi3KRBqlMI4JYlQUCeDgHUAUU1ArzC0no5KuxxeAbm/nwSHn8xC61LbdC+ch06zO3RbmaP9hV2KJy4Ehb/+D083c+gueuNgRlab6CEyz5GAMvTZZ9CPu4+icWRE8FYuWonPHYdQV6+VmjLxuimOuaM5iPFM6mYjcBWRxuQwqlkKcIJR4odEiw/81r5WkqfyhYl2c1YrcWotXi1diYQKfmHW2w030DbYqPKFXt9A81XoO9AT6+2rhY8iaWO23HaWrBs8SY4OR8VZ4vOzkdW0A85QKwgIqBegEguaO58g4s3IrBj7AqkGS9D+1IrdC2zRqepNcrmmCNg5CK4ux1HdctLyZyxolpZQXmFH+sAbonSdknmIy41XxB/4VoItu8OwryZDigsKEJNje58pGrtjCH6Pr3xfg1Omp2v2f9MS6ozgXhdI+Wb2jYv8YB18ORpLXxnYv4jT7h3E1nvFhsOqDxhYpInSkk8+/07bROeQekizS42xjrIPZryAUpKS0U5P7wXCqPJdjh14QluPYrVFp/OHTKaP6BHQF6FDgFUwhoCCEx6ty2dbxGRUoI9LkewZ/Y6HJxmiX1Glti5ejeu3Y1Fc/c78RHEEzZAQC4RkP+xCCLiOSY50Nf/Es5fD4OxkT2uXLgnp0FSClA6cP2kZmXXKzjV1tSIxCDnSx1QfJxwPRslBE15Nj5D+QkU44ZBuz6dne0aAgyO1KI9r0oTKyoq9WchKEeN8p+OGvPDfBDjQ7QCVOERkUJ242DyvLw8+Y3nATnY7oHJAhecvRqKy8HPEZGYI06Y7JLMLEEqd8nkVshuRmWGUpHSqWIogvuCufOluecDcirbEJ9dg7TCJtS1v5NSRS0Uwf0BzSgub0Kh7JDRIYAcIJtBNGsrp6QGT8JS4LjBD57eF7Fi5U7YWu5CaWmxPu9NcauVFr4TGW54+iG5gs6qWicVsvKZqCMolihmVFkiZT85hvAikfJaRBC/YOfKigrBMjHMjoz5UMzwN35PeUjO4LuygshaKnLKODf78CA/2UPc2CifqTv4HAb7CvLzpWJtydJtOHPlGU5feorbj+LEHOXmPFIpX5kFlbI/jMDrRUCLxIIqatv0daLNbW/Q1PpKyhS1iol2uUdDQCOKSqiEa5FfXIPs4mpkFVYJEkIi0uHlfREODr44fOI2lphug8UKVxEv3JhN7uWa1Z5pEh3XnJuj7ZMg1VMkZWRo6+S9XCcJlyKdn4k82aT+8qUuWZMhcCBylAjvk5SS5EV5zSa7IGNj5QYihXuGlR9gGKwz3D9ABKhD7zgwWVKxF+/Xjj54qe0g1/UvLSmG+TJnTJu6Dn5HgnH2yjMEXXyEG3ejJAKamFmC7KJqFJTWawE1HeVLQl4HfFo4jIjWNb7QVUZ0y2ctZK1DRE2rcE1RaT0y8ypE2T4KTcalm2Fwdj2GnR6n4X3gKubP2YiN6z2FQlnnyUaCIrcTDmzkApVmJHfQX6J4ZqOIieFhh7pULcUP4cVGbqAPRaSyEbFiqChPWAXjqL2pmYldvqtrYopIoGxT9e0cmC9+r5S0Kj5lHy2AV6+dpaOeJ4pa7Teuld30gQFXMG3CWqxcuQsHAoJx5koIAi88QtClJ7h86zluP47Hs8gMxCUVIC2rDNkF1aITSiq4W0YF5Rgb6hCxQ5HDzXhp2ZWITynE8+hMPAhJxNXb4Th96QlOnnuE05ef4Pz1UKH6Zct3YOJYM1w6f1tOwCVFc+5cB40Qxm1UIS3XyN/5mWsg0LV7ew/r45r5u1LWzU1a7Q8pvXftLNbVjjggwX+0SU95uiyZMDyMw/DEJ3WyLhupXSlhYpkleGw0XZ8/f65XwlJJp9t9GRERLvY0W01tJeJiE2TnyvRptlhgshluu0/h+NkHOH0lBKcv8/VE/wq6/ARnr4QIAC/eCMOlm8/ldfFmmCRVzl17hjOXnyLo0mPt3qtPcZbfXQ3BiXMP4OV/BXbrfTFjuj3mzXLE6ZPBcvBsWnoKmpq1Ilqe/RkRqduIKMdlZujNbDZKBaWECVDDddLWV6f4srGfyoTxeXyuOjOJiKFo68Mj5SnTDeMUEutpaxOZxkE4AMUJkaFiGypWRFZSsSICmH1IEbymKOLZO7zWuIiVd83yzvv4PWVpW1sz8nLzcfXSA6yz9sD0qWsxZ5ajJEiYEfP0uQAv/8vCJUdP3UXA6fs4fuY+Tpx9gJPnHsg7cwlHTt6WPPNe7wvY4XEa9o7+koQxWbAZC+Y6YdG8jdjksA8hj6Nl7xaPHiCxcB6cMwmQn7nFlmvluvgb10+qVfEchpYJJ3K4WufH95YLh/Berpew4z2EHZWv1pfH/HRrO2Q4GBsVBaldHLF3mpmpHBHDwixDM9VQB9AbZB8l3zgJpVPIPZwYGydA6qBVwUbZq0zY9+/fICw0AneCn0itzmwjO4wdZYH58zZi9gwHyWAxtTljmj1mGPOzI2YaO8B4mj2mT7PD3NmOWLp0q6RAVy7djIAjF5CSlI2qSi1k3NyiOZY8ITE6qldfVeq24xIGbKTOj+eYqjcleY+hHCeAuU4qXjbep3Qkq8gJH66ZTczQaAMdoE5LMSw7UZqf15WVVaLZec3OahcgFU6m7ALUdlFSMZOqec3veB8pQoU2OElyltqFqb5Xpe98Hi0K3keT7+3bVygrK8bs6Tawc/CB35FrOHXuIc5eeoozl57i7OUQnLvyDOcuP8PZSyE4c/GJZLr8jl7HgWM3sXChM44cvKQDEK0wplO10hHOQbNGtLnyMymfcyIM1NGVao4MLdOQUOEGEhZ/o8RgX7VOEigJk4jgmggTdWpKe0eHPJfjajDSmaH8pwcigDfzC2pt2rmkcloApFoOymsCnoDl5DgIB+cC2Id6g4jjNSdGjiH7Sel7fr4ghwhQ/fk9J8d+CgE8EZ338Rk8eTc6KlZOqN264wRcth2Dj98l+B24gmMnbiHozAOcP/8E5y88xZmzj+TzyaB7ktrcvPUYTE3dsH9PkC51qlUo8LlUfJwDAcE5cP6cK0UkfycMSO30Ut+8fYsC3X4uAo2Sgc8iAtiXipSnoFBKCNwKi+R3IkojSuYEtO2oalzlBygj528O7aOjwc4EinZ2tKac2Flt0iNVKCXNSRARXCgHERGkiwUZsqaqJ2LjpJRoYmN/JdIoF/kMUi2bvc1eqYDzPnAZN+9F4WlEKsKZu43PQ2xiASKjs3HvQRzOXXyCQ0euY72DD8zMduPbr01w5MB5VFRoz+GaDA0NAo1GhDIl+c45KRFEAvyhmFT1nJopaWCGlpfrNzeyETa9IkhzWPPzNSdVjVut66udFaFTHqQCUjuxw4cR0MrRUnvDeK2qvPhwll5zQCKAwCNweR+fxwmS8jioUsL8ntTHZzMvyoJgIorUwGcT4XwGzV7ea2flgckTbLBi+XZYW+/BWuu9sFrjgbWWe2Br5SlbVDfY+2Kj0wHs3BUIZ9ej2L47UCoujCauloOUqOg5B85bzYFr5DjcpE7RSoXKOfF7fub6eD/7cV6Gc6S4IoIIG95LPSLraW6W30loPJqTn7ke7Wh8bfuq5uj1nrLyUWEWAcMJyj4nnVNBylRmmDoHjYOzEeA0y7gFR/3OxkXwjGa9eZZL/aBZCTRDyaLaszP0Ji4bKV89++XLHjwPC8eEUebYsOkAtrsHIvDsA5y9+BjXbkbg7oM4hISmIjwyE6Hh6XjwNBHXb0XC0/s8NrkcwibXIzCeshb1dQ1IzzDgsJ5uMaWV80mg8DBZzo2NIpB/AqHM0Kzs7I/MUK5fcRFF7fPn4XquoRnK0ARFOhsJkdtYee5QQyPPFKLlqBk13NbKJjlhXpAaCEh2IiJU9FNF8tR7b8RTi/oRYQxK8Zq2NAcnAghMfkfO4POosBkd5TUdGP7GZ/HFseigqIhhW1sLel504dnTSPT/fDZWrtqFQ8dv4sbdCNx+GCOvO49i8eBJAh6FJOHB03g5uO/KjTDs8DgFh43+MFnogn6fz8bz0Fi8eNklx85oANEcR/17Xe9nOlDqICq1dlYxqH5qvUSWErm894cwomSQeI8OhoQLk/VcN4mSwGdIm9yj5wB6anQUDM0wUqeianKIiodTqZASVOkdP1Ouqc98sKF81WRmbwSQFhd/V5TC83aydAfBAu9x7MhZLJ63HoMHLMGqVbvR7/N5GNJ/OVYs2gUbq33YtPEQduw4id27g7Bzxyls3ngEq1d5YK7xRkwbtQFDvjTHZ78zgsMGf8lsWZntwJOHUYiLi0dzi+Y4UvzRiVJzosh5/jxU0rJs5MQfzlHJda1pcGGjyNH2BWiczaaqIBTnvejpQUREpF63cfe8OGI8qkBzkbUjhrmjg/KeWOt1KrRzMwlYsijlGa95T2+grVW2tLIvqYDPIvZp33ORytETZ0cnZzUK5GGpb0UUnA68gdUr3OBguw+PHkRj+9YjYv977LsAo4nrMfgTC8wYvB3T/uqKSd9u1l4DtmDKQFfZp7xkgjeMB7li8kgbXLrxDIeO38L0qTZyuJLLRl8sXeCE3dsPIyoyHp2dbXjz9pVUKfCERM5VOVXKYlO6gABW6+c19QHXqQDO+2kNKbFFmCmnjYhWBQ+ECWFNGJDzxBHT/UGaTERhk2EF/sjGMAIxT04gVkkJNFnZOAFqeFUBxoAbrSjl2IkuiI4SL5KNQOb9POZFm3gxwsIi5UyI5Ys2w8neCzHRGWju6EFxdSMa27tx9dJjzDa2xxbXozAz24WhX1vItljz6QFYMeWgnKC7fPJBmBkdxZhv7DB9sg0eR6Rgq1sAFsxaj6TkPDR1vJCzQfPyK3D00BWYmjhJTWtEmHaiFTNfPPNCAS8hvnf9XCstILX+1JRUATYlAgFMUUSA0hAhApTZq0IQtKJo5iq9wT4flaWo/5Ch/FMP5UDKuiGw+JkOEwHNawKekyIV8DMnq5mp2n+o0P7ns0hNb9+RuhuQmJCgH7SwoBjuOw/BdOFGfPmnGVhn5YmCggrUtXRIzD6WZ8hla2fI1bV1IT2jSBBkbeMJ30NXYTTeHiO/tMHC0fuwcsohOax7yOerYWO9ByFR6Vi6xAUb1+9HbUMbSmqa5TlyKGBeGcrruYu+C9evhuDPv5qKiaPN4b7jKKKjklBSWoyXuv+EoU7kmrg2+gMUSap0n4DmOskdvIe6gJ/J7QQ8gc5jPnnN5/A3RgKICCJY9GSjzg/gRm0qDAKLJiiBxwHINsQ+J0AvkOFUPkyFFYgQcgvv5QTIKeoz2ZcbFQhwfiaV5eRkiXw9E3gb82etxyH/i3JGz5/+dzKu3ghDVkmVJGdYKMWKNQKNmTG+Smta0NLeDa89gViy0Fk2U2zfeRxjBq3GsM8tMbTvMgScCEZIeBpmT7dH0MlgdHS/QWF5g9QCpRdUSsUdnxudwixcIeJS8vH1JzPh5OArJ6osX+QMkznrcP9umBAJrReuUVE1jQj6AYQHv+OLCKCMp3giXEi0RBKtPBIp4Uci5W/UrUQIpQP7UjpQLEksiDepmA+pmCakcip4SrrKiFF7U8QoEURM0+xS2TMOSvZSeQNSPpUvayfra5uE9dfb7pWTT94CePg0EZ/8ZgpOnX8ohbssD+FBqqwRYtYqPb9cqiNyimvkxKue1+/x8H4UjKdYIyDwLjLyK3Dm/D0kZxTh6PFg+fus+PgcdL54J4n8vJJaOVWXOWaFAAKfhWDcEf/5H6dj08aD6HrzAS3tPXj0MAaWZtuxcK4DsrM1x6mnhxZU+0cZP8KLCtrQYCFcaIyw8V5VN6XyAarvDw0SyQeoWL4Wy9aUEM0mIoUYI6sop4JUrhwQ3q+ipOp3Ffnjb2Ka1VUjOioRc41tcfP6M7x5D+1wpcZuXLn+HJ/91hhXg8ORkFX0NwhIUwgoUSWKDbJTsrS8HpZmO7He3kdOQFy/zkcAV1XTKiWLzISxqloQUFQthV56BKQVIC6jEE8jU/HNp3NgY+WJstpWyTF0vniPV2+B+/ejMMfYDvduP8fr1zyWTKts47q4RoFRaalQseQKqqsELrwmzMgZLMnhNQ0VGjF8hsC5oV7v1EplnDJDmXygWUZxomE1T6q/lBlK5amyPPI5KVFvllGmcVe4MkNJGXwWHZCKslrMn2mP+IQcNLS/QEFpg2S4Khs6cSLwLj77nTFuP4hBfGaRJh4+4oAKPQeQorVy82Y5noDccNDvIqaTGw5fw8s3kDpRw2S8hoAqEWOCgEwdAtIKEBaTgYFfmcB08TYUVjZp2beKJtmPUN3MrFoLVi11wZlTV5GTq4UgKEpY6abyxSQ2+UeRV7rcQUaGfjcRG6VDbz7gBcIjIvTSQXn8ffh3rqRmveNh4IjxpZwvXhPjdFR4jzJdq6u0chZxzHQnyKpnsXB38lgLzJhig5t3IqXsPLe0Vs4F5eF5/v5X8PkfjXE/JEEqFGKS/y8E1EpaUfLCkppkNqwDnS8/oL3ztQCfG7X5PX/XI4AnJRYaIEA4oFAO7GZtKv0MhrXJJeo0xfyKBukTfDcKq0zdMGOyNVpbtXwBS07E2aquFipnKYpaa69DqcGML8JIk/W9pxDzfvalg0dCleNqVLaKjaYhsaxsWlpChq44ZZ06SZaNWFZxcuoNUgQLV9m4RfRXPxmLT35tjD/9aqr8/dPTiDRtD0BJNfZ6nhEr6GFokr5Cje96BLA8RVcjqhBACmf5OcVFFQ92LahCXkE1GlteoK37LcprtLJ0HlcpCDDkAIWA1HypSx05eAXGjbCQkhjuXcgqrca5y08wc5o9Pv2NMX71kwkYP8JCMz9TUz/a80UON9QBdEw/dto0acHvxGnVhViUdKBIYuvT2t7uRc1O7KhgFGUXKZnswmvlkPGlHBJeV1Rq/71I5FGuUYxpzka9yH6eqP7bn09G389M8PnvZ+KLP83AuJEWuHw9FKmFZXDddgRffzJL8rbqAFXhALGCdAhQ+wRK61Fc3oDm9lfIzK2Ay+Yj8q9Ik4fZY+JQGzm+/tSpe6hv6ZZzhUQJy4aMv4MA3ThTJqzFoH6L8DAsQUoSXV2PYvTwlej72Rx88utp+PQ3MzFrmp38x4wWhONZD6oaukWonva9odNG4uM1X4Qfv9OcVO0/ediHfZVEMdABlaKdFVaptXkeMv96g000u64qQD6nper/R4ZmLGWfYSgiPSMVt24+wl9+NxN/+d0sjBqySk7K8vK/hHGjLHD6ymNscT2Mbz6djTuP4xCvO0T7bxCgE0GsD+IJuk+eJWLM96sx6FMzDPncAv1/b4oBf1yOUX3tMOQvllizyl12x5RVt2gIKNZOy/0hB8RnFGPhAhd8+aeZuHA9RP6CcRZ3VJ64DWeXw5g4eg3+5ydj4ecThELdEcxa8DFarwPEYjSohqAprrKDVJ0JiQkSzGPjSZO8t9dC1A7ukINbldVCEaJsW9rsFEP8wx06IRyEWpuJd4ostQmNJinzvuxLbqCtTO5hXx75u3TBRnz+hxkSr9+3/6wcWeNz8AomjbfC8mVuGPDlfNy8Hy2moXaauoEO0COgBuU1LUjOKMbo7y0wpp+dHJNjNMwRgYduw3fPJYzsb4Vx363D4E/XYL2ttxxr8HfNUEFAgZijlpZ7RQcZT7XBzOn2uHIrHDt3nZAN214+FzBptJn+z6NJXOVlZbJORlIJI1I1P9MCIhz4d7b8TCDzfsKAyOJvvIe/aT6StlmPhMu/NJfiXBWfZmfKeQJXw6p2jDFlGTuo5AmdE/oBtIxydSV47MPzR4lEWk8vX3bLLpI//HKyFEI9j83AbvdARCblYr2jL/7yOyMM/tYUV25H6BCg4wDWieaUiGLUNmrUyCHc/P+BwX+xwIh+/LelHXh4o1cX7Vp/GUO/sMTEQQ4Y/s1qPHwcj/K61l4doNsTxudTz3AOG5wO4A+/mIgh3y3FyXMPsWPXcTyLTJX9C9zNudZihzyba6EuVOEEUj4PLySns1EspSQn65M3/F2F1enEaqEIlRNuEWmh9wPUFiVSLR0uyictFFGsS69pf+lEyiYmFSL42TA0of6miZ+pEyiycrJzYW2+A25bA7B4wRbkldfiyo1Q7PM+j3uhSZg8fi36f2GC89ee6fdq6TmA50jnlYsnm1dWj9zCahiNscfwr9fAaOgWzBvuifhIDfFsux0vYM6wPRg5wAojv7aBm2sAapu7eq0gtScsrVBqQUOi0yV9+elvpsLd8yy27z6JyzdDcfN+JKaMs8Kp0/flfzEP+p2W3LRKKxKY5HSuk4TLzyoUQQNGpVQJI5qg5AT+xnsIX30oQldbpE/KEyMEGhHAzuQEApnAJWYJbJUnfv36FfJyc+Xh/My/NCRbsQ/vJQIKCvJwLugGFs52RGR8FrbvOIk15u7IKqmB576zOH/1iWxXHfj1Qpy68FjEgmYFGThiREBBJYoqGxEdl4MJ39tiVH8rzBuzC7OHeCDQ76EA/+WLVzCf6YfF430xdqAdRve1g7X5XlQ1tAsHaAW5JXpHjCbvo+epUhM6cdxq+Xds/0OXEfwgGmOGr8D5y09wJTgMrlsOCRfwv2Iohgh4SgBSPtfJqDCr5ag3CQeKZsKEv5FwqRNI+fyN97AvpYOWlNdOpzSIhmpmKIFIs1MpFiKBYkYfDUxN1eeDGaRKSkzUhyaopNhXKZpH9yNga70HJ0/fxvFTd7BmtQccHX3lmALGcg6fvIUxw81w9OQd3X8J/MATziUCKlBY0Yj4xDxMGroOo/pbY/ZIN5hy7+7EfYgKSUfA/juYNdQdy6f4Y/QAK4zpZw87Sy9UN3YY+AGKA/ifNUV4EJqEtWv3SOCO4jD4fjSmTrCUhP/hgBs4euImbt+Lhsf2AFkLYaBKSygdWIClMn78nspXiSASozLNVS7asCyFZqleBBUVFXnRBKUSpsKgKCKl80YClAkEKmLKd/5OOcgJ8DPZilgnwPmZ3EMKoIJpa2vHJof9GP7Xpbh5L1Ioyn3PaRhPscNer7Pw8DojacZJ49bA9/A1RCTk6LzUXiuIlCtFuiW1stlivtFGjPzGWqh88Vhf7Ha4hPKSWiRE5mCl0QHMHb0bIwesxogvLbF/3znUNnf+HQQUCgfceRoPS0sPjBq2EkcCb2HBPCdsczsO34OXcersPZQ1tMF0qSscbT2lnIXrVkc18J1ER72ptiAxEkrDhflnBuVI1IQfKZ8w4T28l33YlzB/aXhYB81Q5kqVGUrA8q/8GJplo4xXxanaZ+6S18xQyjyGLVQBKourHGx3YavzQTx5noypE63h5OSPi9eewdn1CAYPXCKh5TWW7uKN+hy8KudGSJRSt2HPEAE5RdVi3/v4XMT3n1lg3Hd2GPmVLby33UBP5ysU5dZg8SQPDPpiBYyGumDkQHMkJOfLWdGihA3N0HQtFHHzYTRWr3HHoH4LsXihsyBgn/cFnDx9RwyAVSu2Y63VHjk39Njhc7IuSgUWVXGnOxuJk6alClxSWqh/12Yj5WdnaWYorUbmRgxDEYU0Qzs7O2WbKlOSuXl5wgkq10lMks1I+cQoB+Q1f+eDqER4TW4hwigTGTns6GyXqrRtWw/j6q3nAtSdu05i/pwNWL7cTSrbBnwxDyMGmWL0kOXYu/8CnkSmiiWkRBA3TiglnFNcjeKKRpTXtmKpyVYM/9IGE753wOAvVmHSkHUY0W81hvddDeOhLhj45+U4evSG7CHIK67RI4D7AVQ0lBtDrt4Jx/LlWxF8/THmTrfF8uU7sHP3STx+noz5sx1hZb1X9i/s3HFCchFdneTwZqFmZv5IqAzCcd3kfMKFO18oHfgbX7ym0uVvSjqwHKXD4B/MpS6I8qlFpwOKS0q0AJNBjYvaIcMWbxCKoF7gn/oY5j0ZxmhtbYKLkx+OnQjG8VO34eN/AXcex+De01i4bj2GebM34g+/moQj/pewf+8pbNh8EA/CkmR/LhHAd0EAdYAOAQyWVde3y6a71at2Y/AXZhjx9VqM+sYOY/ray0EaI781x7GjN+Q4G3rNvcE4tS2pWMQPjyI4f/0ZlptuRXJSGpYtcsLSJVtxKOAmpvBUFUtPEZMHj13H07BkrLXYqZPjWkiZ8DKUFipw2RuKSNYHLvkdTdZ83V/dUkyxCoPxITbJCatwtPrvdQlH64pWyQmkbAk3NzejnKHWioqPwtMqHE3uoZMSFhqNGVMsMXOaLXx8L4p5dyjgOrz9LspxZTvcT2HqJFucOn4NPvtOwcHJX8rRuWVU4jIKAUoEMR8gmzQaZV8AAXz3bhRcNh3BmhUeWGuxD157zyElrUj+W4CHcxBR+T9EQHYxknNL8fB5shT0rli6VUxlty3+mDdnI2Ya22H+HCf4Hb6KkMhUBN+LwmqLXRJKb2rSApH6cLSucFn7zHC0VnqpwtHUEwpGhqFs9vk/w9Hh4b3haB4uqoWj5aOuKqK3KoCyzjAczTAG2ZHN3todN4LDEBmfIzF7loLbrN0LR0c/LFvmJon2CaMsEBEeg+0u/rB38Jd9AXTI6Izxf8YoMmiCcqNGHnMB5Q3642qq6zvkDxxevAW6Xr5H9yvINU9Q5Oa88motIlpQVifIIxK5O5KynX9REnie1dR3sXyxC7Izc6SeZ9XSLRjw1Vy47TgBD8/TMJnnhIXzt8Bjz2msMd+By+cfyNpY2aDyxUSCtgFFq/BjFlD9+TObFo7WpAO9YtZEqaoIfTiaW5QIdGKEmFOySQux8tQPdeqfFnZWIVb1WQvBqj8o00K2Fsu3Ij23DIVVTWjqfisH6V24EgIX12MigqaMt8JB3zOyZ8zL4wSsbfbh+v0onLnyFJeCn+PW4zjcfhyHO0/i5Y8W7oUk4v6zJDwITcbDsFQ84ut5Gh6HG7zC0vH4eZr89jAsRe699ywJ954misVz+0k8bjyIxtlrITh3LUTGWjTfCfm5BejoaEXQiesY/O1SLFywGRbmu3DxegjyyutR3/lK/kLFzNQF5eVa4dr/X+hebcTgi1TfC6N6VFZUSjWE2klJbumTlpZ2RLnUbESGCqqx8QEUM6oR60SUakVFxfIw1chJzk7+WGfDo95PYveOIOzacRq7dp6G85YAuLicgMmczYgM14JWEaEpmDxuLRzWH8A6Wz/Y2fjCxnq/vGzXesNmrTds1/rCdq2PvGz4uw0/+8F2rb/uN91nG8OXP+xs/LXrtf6wsfaDzVpfrLP1h/26A7C29Mbc6ev1Yffga2H44g9zYGXpDY+9F3H46G0cP/EI586E4srlCCye64z8PM2CIeBUJQibinSqRqSomig2VbaoGqWL4pr/D3B0yklYzExBAAAAAElFTkSuQmCC" alt="Asistente MADEVAL" />
          <h1>Asistente MADEVAL</h1>
        </div>
        <p class="subtitle">Consulta la informacion de la empresa con una experiencia clara y rapida.</p>
      </div>
      <div class="top-actions">
        <div class="local-clock" id="localClock"></div>
        <button class="secondary" type="button" id="logoutUserBtn"><i data-lucide="log-out"></i>Cerrar usuario</button>
        <button class="icon-button" type="button" id="adminOpen" title="Super usuario"><i data-lucide="settings"></i></button>
      </div>
    </section>

    <section class="grid">
      <section class="panel chat">
        <div class="messages" id="messages">
        </div>
        <form class="composer" id="askForm">
          <textarea id="question" placeholder="Escribe tu requerimiento." required></textarea>
          <button class="accent" type="submit"><img class="button-bot" src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAGAAAABgCAYAAADimHc4AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFP4SURBVHhefb0HWFRZnj7s/ndn4+zM7OzMzk6e6e7paGh7bHPOgjlgwgQoICCCGADFBIpIMqMi5qyYI4LknHPOOUezvt/z/m6douyZ/c7z1FO3qu6555xfTudUn6SkJOOKqiqvDx8+eLW3t3vFxcV5vXz50uvt27deCQlxXo2NjfJbSkqKV1lZmVzn5+d7ZWamy3VNTY1cv3//3utc0E2vmJgYrw8f3nu9ffvaKyEhQd8/NTX1B/0z9f35bMPxe3p6vN6/5/gf9y8pKZHrwqJCr8xsrX9VVZVXkmH/BM6/x+v169decQb909PT9f3zCwu9srOz9eMnJSXJdWdnp6yZ43P9SUm9/bnGxsZar0Xz1nt9+tuJXunpWv+KigqZZ2//BH1/w/kbjl9YWKhff5/09PQjL16+gGpFxUX665KSYrx7906uKyoq0NnZKdfNzU2oqamR6zdv3qCmphq73I7B1mofHGw9P+r//v17ua6qqtL3b2pqQm2t1v/FixcoLy/X9yksLNRfl5aW4v373vHb29vluqWlBXV1dXLNuZcZ9C8q6p1/cXHv+JWVlfr+zc0tqK+vl+uXL1/KOIZ9VCstLdGP39XdjoN+Z7HNJQCHDgXDzfmQfN/d043Ssr/fv6SkRA+/6upqdHR0yHVLS7N+/n1IuR8+fEBGRgbKyspkwgUFBcjOzpZrAi4lJUUexIUnJiagq6sLr169QlJSEnpedOLWjRCYrdiOVx8AZ+fDcN28H7W1VdK/sqoSqala/7a2NunT3d0tgE9ISBBA8D7D8fPy8mR8zouITktLw9u3b2V89u/q7pbxk1OS0dDQIH0yMzNRVkaEvUd+fj5ycnLkmgvn+CSU5uZmJCYmyvjSPzkZ9fV1fzM+kaj6c/2ZWZloamzG/Nn2yCmsQWvPO5gudcbF8zfQ2dku96VnpOsRTiLKzc3V909NTdWvn/AjIb5+/VqIqg/ZllREbFVXV8lNnAg/85oAKCouFuppqK9Hfl6ecAB/Ky4uQldXO2ytduP85adoan+Ju0/iYTLXHjW11dKHACgqKkRbexsaGxtRWFgggOQrPz9POIH3aeNX68cn9bW2tqG2tlYWRuohsAlcNT4BRQSyP+/hglpbW4Wj+FLjk6D4Pfur8dvaWlFQkC/r47M4nhqf3MI5qPVXV1fiwb0wmJtvR11LN8pqWrHJ+SB2bvPHq9c9MjfOhWNyHM6jvLxMxhf4FRXJs7h+zoXz53383Icyl4OxsUNERIRQB7EXGRkhN7GRchR7Z2dnIT09TWOtmhrMMl6LkMg01DR04N6TBCyc54jmphYkJMSjsUnrT2rv7Z8tVM1WUVGOmJgYueZCIiMjdeO/lfEJNDZSvmJvbfwMua6qqkZMrNafiwqPCBdq5/yjoiL1oiY+Pl4vanJzc4Qr2CorKxAXF6cbX1s/xRK5Lzo6Wr9+X69AODr5oq65G/nFtdi9Jwh7dp1Ae3sLYmNjtf6dnTJncjfH1/o3yW/a+jXxmpubJ1zB1oeKgwtWFEQkaBRYivaOXgoicCi78wtIgc1CRaSmqsoqrDR1xfOYTFTVdeBZeBrmTF+HhPgUEVWUdWRJPpfAIGuyb3t7m1Azv+Oz+SyOz2u+y/jt7cLC7K8fX8cBBDbnpfpz/ryX16REchT7E8AfzT8/D01N2vx5TRFEkUB9xXE1DiwVjmAfUjDF6Ymjl7HZ+SBqGjtRXt2C40F3YbNmJ9rbW2UcElfv+GV6riUH8zeRIA0NIl619WvwEB2gZCgXTVklMvQHMoxUxUWnpCQLYHt6eoSK6uvqsHrVDjx5noLaxm4EnbmPmdNskJycJs/iAtLT0/HmzVu9DOaCX754KbqFQBEZmp6mV1qcPOfAawJFyXCKMfZhf46vZPi7d2+RlZUl97IPAU4ZrvpnZKTLNefP8bn4Fy+6kZKaLBRORZuZmSHjav3zkZenrV8Tofm4fP4O1q/fj8aOl4hLzIOV1R5scvTCq1c90p8wItLU/ElovKZY5/qVDiUny/pfvtR0AE0wZYVwYmSbnhc9smCyVpNOhBDwlPkaC+WK0mJra2/Bovn2uHwjFJW1bbC22gNL8x0yqfj4OD0LE3AEMBuBo0QQEUygsLW2tog4InA5YYoNJUIyMtNRX69ZTvUNNSgr18RRaxu5MkeuP+AtsrLT8f7DO7x58/ojEUIAKRHG8QlwNgJYP34Lx9fWr80/Xi8Cb16/jyWLNqG5+zWC70Rg3twNuHHlCbq7O0SxstG64ZgUQSSK2NgYvbWjIUiJwFxBClsf2tGKHXgzFVBdXa0oR2KI33ESnCixKUqpplp+Z5+u7g4cPXQeTht8UVBci9kzHbBn11F0dWmihCzI+9hPe26tPIuAr6mp1Sm5ajQ2NshYJAbqlYaGejS3NOHlq260tDYhLiEBN27cw5nTN7HV2Rd21u7YssEPdlZ7YGWxA3vdj8PfNwjHjp1BbFwiaqjcO1pkfiSi3jFrxGzW5lUn6+Ca+ZkvXnO9HJ/z53M4L8512aKNCI/KRFp2BaZOtER8bLKIQzV/9uf8uU4+g8/SRFgNqqqrUFtbJ8/SYKmZpX3oDBATVBKKQiin0vVKskIogY3yMTo6Cj06JUMOIYW9e/ceK5Y4Y8VSN6xa5irYzsnJlj6cUK+S6xAKJ/vRrKTCVRROcaI4pKq6AoVF+cgvKIGXZyDmGjtg1jgnzBm5DYtGe8JiWgDsZp+D3azzsJ97GevmXIDFlACYjveBydjdmDJkPWZNXg9P95O4c+cxUnQKl0o2OiZaxieFk0KJBDaKBjU+15+apilJrl/NPy4mDbOM1mHR/C3Y7noQeXnZegqnSawonEYGxR4bAR0XpylpEnpUtBr/vYzXhx5hTW2tWB6Ux5S3HR2dwkZkEw5Am5Vsy8m8fv1GlArZiN+TWmvrqnD+7B306fMl3HcekcFoa1PM1NXVy328n/0p2pSM5/ekSooziox3718jNy8Hp4OuwXTBZswYsRlmUw7BdelteJg/hadFKPatDoOXZTh81kbBxyYGvjax8LGJxn6rSPlt7+oQ7F3zFNuX34XV9ADMHe2GudMc4L7rkJjTbGXlZQJkw/G5fpqmXDMBRV9DdFx9vcw5Ny9XLJ7F85zwoz4DcPXSfbx7/06oXPWnjuGapX9XlzyLnMD+2TnZAnCOaajj+jAswEEIMFIzlRllMamVQCMC+DB24GR5HzFNy6Sru0tj6doqnDpxHT/556HYsztABtT65wiFqf6cLK/LSjU/g8/iM0vLSvD27RvcDn6GmRPXYc5wN2wyuQJ3i0dwWngZFtOOwXzKEVhND8SG+Zexbekd7F71GHssnslrt9kTuJnehdP8y7CeEYhVUw5i5eRDWDf3DHaZPcC25cEwGe2O2ZOdhFBqaymCqkUZUqxSaXMuNBJIvXwnt/NawYZWWkNDHVYsdsZ//dtwnAm8gZevXgqFE5liZTU2CszYl0gg0RJ+Wv9ClFeUa+sv06wkIkMvgqiIFAsSCcpOJYAUC9J0pNhQdjLFiVJSPl6B+PE/DkbQyZuor6/WsyCBruz89rZ2hIeHC3exP5/79u1LlBRXYfoUS4z8eg08Vj+Sl4XRMcwZ6Y5Zw3Zi/qg9WDbRD2uMjsNu9lk4LbgC58XB2Lr0jrxcFgdj44IrsJtzFmuMA7Bskh8WjN6LmcN2YtawXdJ364qbcF/zEPNG78KsKbZ4HhqH/PwcUZyy/qREfRiE4jNNJ4K4fhoTbC9e9GDWtLX45X+OweUL95GTmynIkv6JiYJINiJEGSlcv/ITiPCoqCj9+omEPgwQKbuewOSX5ASKI4oaYpCsRazxYfye71VVlfI9lVRDYx1OBlzDj/9pMDx2HcPLVy/keeQUUpC65mLYn3Y2HbBXr17g/JnbMDXZgiOHr8Fk3A5YzwzErOG7MHvEbswfvVf/Wjx+P1ZMPog1RgFYOzMI6+ach8O8i/Kyn3seNrOCBEG8Z/G4/dJn3qg98pozwh0zh+3A8skHYGV8HItnOYvvsmPrIXR0tqG6hvNqFCVKJPCacyZB8jteU5fRKFi+aAt+/u8jEBR4Q3QKFSxhQ1iQ27X+mp/CNRN+hC/vIbfwWcrTJmeIGUrZzEbWIVUSQ5T1hmYY5ZmhkmJ8hI0WRWlZEc6dvoX//KfB2OK0T76nkiVVcDJsomRLtf6FRYUoKS3Cnp0BWLvaHU3NHSivacLY79bA+Hs3zB/t+RHw54/eg/lj9mLR+P1YNtEfZlOOYLXRcVganxSxtMb4BMymHsGySf5yj+qjIcBD3rXvPDH4Dxbw2nseb95/wHbXACxbtBnJKZqSJnC4Zq6f849P6DWDqQ/q6mqwZtV2/OxfhwnBsdHRowJno58SFx8n+oNimIZNrxmcguISTQeRUzKzsuT6o2CcFn1kMKlA5DeviS0VjNOCSYno6qIj8UIGVgg6cuAC/uMfB8l7Y2ONPphWXV0j4owLam5pFufn3bs32Oy4H3Zr96CotBY5RZVIK6zAolnOmNDXCSZjvX6AAO21YLQnTMbuw+Lx3lg26QAWjNiHGYN2Yck4X5hO9JXfDO9XHDBvlPbZZIwXhvzJHIFnHqK4tgndr9/hoN9lmMx2QGoKDYYavP/wQSLChsFIzp9W04uXPVgwex3++z9G4tL5+0JE9Q0N+PBBC8aRS7hmCWZmZemCibW6YNxbvSNIPWAQjGv1IsaViOBNBDpFRLsuGEWx1KILHlEsaSKrWa7Jot3dXThx9Cp+8qOh8PM+jbb2Znken0WbmJwj1/V1aGyqw3aXA9jk6Iv65m4EnrmHyMRc5FXUYd/eC/j+92ZY+AMELBizD/NG7MW079wwb6QnFo3bj8nfbcaSuW5w2XQMRqM2YObQnTAZsw/zR3lixvc7BVkfPWPsPhj/dRsmD7JDfHoxknNKEZ9RhNbuVwg8fgvzptvLXDs7tVCCmj/Xx2sSH8XVskWb8PN/G4Wzp24JwpSIZijDMJinAnPsrwKbWjBQE1HkNn0wTtmv1ORUEioYR69OiRDDYJz4CTpPjoMVFOTgwpk7+Nm/DofrZh/5/u3bd/IsxYIUQc3NDbhy4QHMlruhres14tMLkZBZjISMIqTlliEqPhdj+1nC6DtXPRcsGOOJWcN3Y9oIR1iYemDqUCcYfbcNSxZsRUFFLbpef8CdhzGYNNgec4d7YMpQRyye7YapgzdhwaheJPB5f/3NCjhvOIbi2mYkZZUgLr0QMan5aO7sga/3RaxZuU08buWJ03hQ609KSkRtXTXWr/XET/95BI4cvKhbfwVidXa+gt/fM1KSDYKJhJ+KBPTp6upi9kgwY2hm8mayCjmAJiSvORmyJsPE/MwHkQI4IJMVVMKH/S9IgKqwkKHqLnkmWZImW0ZaLmYZ2SAsPA3nLj4SCoxNK0B8ehESM4pRWNWI7S6B6P/LRXrKXzB6H8YMsMO5y4/R0vMaBw9dQ//fmmK3RxCqmzpQXtWCzOIqLJrtgmF/scbBI9fR0NmDra4BmNB3I0xG78PCcfsxpZ8Thn2+AuGxOcgurkFSdokgnuNHp+ShtqUDlma7cdD3HNramlBcVCSw4PoYtKMF093diZVLnPHf/zZKzO7KyjK9mUnipEgxNDP5PTkgL0+DH71mmqbkDD6PffswdUdsE7BkGyKDD1OOEtmKZunrV6+EbcTRaGsXKmEAr7GhQeS7v/dp/Mf/G4TDBy5K6IBAZ39SCO/jta2lOy5dfIo9e07j3tM4ocC4tAIkpBcKAtLzK5CSXY4Zox0w7A9rhPpNxuzHmAFrcf9pLHo+AEHnH+C7P62Cqcl2FFU1oPv1B9x7EiscMOab9Th49BrqOjrh4nwUE/o5YdE4b8z4fiv6/twE/r43UFLbgpTsMuEAIiAurVAQwHnkF1dj/gxHPH0SjspKDRZcP7mdRPbq9QusMnUREXTy2DWUUgfoHLW83Dw9/Ah8tX4SMNdPqaKHX7sGPxJnH+YtDfMBDDVwMCLlhyJEsZBhPJ9mWHFJAc6euo3//Kch2OIkboUoJlpU9Q2aFXH4wBlssPfByZO3ce12uIie6GRt4fFphUjKLEZydinyK+rxODQNw/6yAsP+aI75ozwwe/guLJy5FTu2B2Lm+E2YPcwdkwduxrIFO7DN+QSMRm/AjCE7MH/UXkwZugEmM1xgNGQLFo/zxvRBzvjqp3Nga+6DkpoWZBRUITWnvBcBOjEUkZCNzKJKXL4SAmvzXTJn5jP0oZKkZDQ11WOzozf+80fD4Od9Rr7X/AQtVEM/iTkIlQ8wFEFaMFOJoNxeEVRfX89ksi5opAXhVIBMBah4zQdxMhKg0gXjKP+1eHo7Thy7Jo4Yg2KdXVomi/ezHye5Yslm+PpcxOGA60jKKZEFK8qjLkjMKkFKThlSc8pQWtuCGzejMPiTZfj+NysxZ9gOzB62GxP6bcKc4R5i7dCiMR60HeP7btZ/R45ZMGofZg7ZjfkjPTC5nyO+/tl8WCz1QEFZA3JL65CeV6EhILtERGBceoEgICopF2FxmSiorION5V7cuvEUzc2NMnct2NYogT0G/n76z8Nw9NBF8QsoRrhOdZ+CnxZ8ZLCvVn77ONip9aEo0nvCNDUNw6WGniztWTYqmdi4OGGf129eSzCLUUM2KqUf/78hOHrwioSLlSfY0tKEK5duwcpiN3z8L+JhWCIik3IQmZgjCIhN1RBAiiQC0nLLkZ5bgbL6NjwNTYPxSAd8/bN5GPO5FWYN2SZ2PXUDlSpflO8m43hNfbEXc4bvxNQBTvjrr03xzS/nY9vmk5JCLK5oQkZ+FdLzK5GaqyFA44ACxKTkCQJIFHEZ+bh1NxIrlm6R+ROo5H42iiDTRRvxX/86AoHHr6O4ON/ASOnN+ImflKn5SexPTmKjhKFXrJQ8k/l9WO5BDFGOMwZkmLTmNbFFdqJrTlFF7c6BlJ2smVzFYhf/7F+Gw3WLLz58eIdXr1+hoDBPBqaF5LYtAE+jUhGemI3whGxEJuSICBIEUAQRAdllSMspR0ZeJTILqlBe14bcwlps23wCQ/+yAn3/ywRDfrsK476wxZT+GzBt4CYYfbdZ3if1dcDoT60w8H+Wou8vTDB7wkZcvxmF+tYelFQ0IbeoVp5JPfMRAtIMEJCYjbC4DCTnlGDpws24fesROjraZM0SWGusxUYHb/znPw3HQd8L8j1hl5ycJPKeFK2KDnqT/r1FB8rPYmqSypj9JRRBLKqUmZY0bpaH0TJiDFtdE9i85qAURXJNlqqvwZlTwWKGMhTB1EhLawsO+AfCdbM/poyxwPHA24jPLERYfKaGAD0H5CM+rehjBORXIquwWqyVksomNLT2ICGpCJ47L2LuhM0Y8ukK9P/VIvT9bxP0/e8F6PffJhj42yUY298SlqaeuHQxDBW17WhofYGi0gYUFNcjr6gOmQXVGgI+0gFEgCaCyAHP4zORlFMMx/U+uHjulqylrb1VS6y3N8PJfj9++qOROOx/UfwfwoSOq0q69/pJWsqWIkhL+vf6WYQdiZmc8FEw7u8pWXbSB9N0SXsqaXz4gNiYWL2S8fcOEiV8NvA2KqvKYTxpDRZO8MDONQ8wdZgTjp++hZjUPDyPy0R4fJZQf1J2MVLzy0XsUPykZJcKAigmiICc4loUlNahuKwBNfXtaO16jaq6diQmF+H27VicOx2C0ycf48rFMDx7loG8wlq0dL5Ga+drVNS0oaS8CcWlDSgsqUNescYBaXkVguCc0lrkltcho6gKaQUVSC+oQGJWMcITshCfUYjduwOxaPYGHD1yVmL4bB8+vMXqFVvx838djWOHrkgwTp8PSO7NBzCKYAg/FcxUwTjlJ0gwTh+O7uoWeUaTUyWNiQgCmJiiWUqs0Z5lapKd+b0omeoKnDpxEz/90RD4ep2BnfVuWM0KxC3vOgR7V2PJVE8cCbyB6NQ8objU3DI8DEtG4Ln7OBZ4GxeuPsPz2CzklTWgqKpJKF/jgFoUlGgIoBgpq2oWBDS2vkB7zzt0vQJ6XgPdr4H27ndoaOmR38urW1Fa2SIIKCqrFwQUlTegtLZVnh8amYkDh69h40Z/2Nnuwzo7L2xw9IOX7wU8i05HZkkl3PcGYsZQVyydtB/Txq3FrVuPAbyDy8YD+Mk/jcB+z1MSjqGMJ2wICxZ8EWYKfhQzCn6ULDRTDf0sCUdnSkKm5m/sVAKa13yAllDJkQfwPg7Iz69evRStX11TgfOn7+J/fjwKUydaYNkMD1z3KoWHxTN4mD/DjGHOOHXhLtKLKhEQdBtmq7ZjxRIXrLf1xBYnH9iv3SufLVbtwn6fSwiPzRaTsaqpE5V1bSivbhbgE7CkbAK5ur4TNQ1dqGvqQV1TtxQE1DTwuw7tt8ZO1Lf0CLKYq46MyYaf/1VYrfaAxYrt2O12DFcvP0Z4eDIiI1MQfDMMO92OYuXSbQgIuoO9vkEwm34Aj4+04eDmWIzob46YmHh47QnCj/9xKLz2BoreJAES0L0JrVQROQQ05X5DQ6PAj/cQ8KoChQhhf70ISkruTcnxZpUPoIxTdq5iISatmUChzausoA32nvjz/07BV59Oh9vqYHjbhGGraTB2rXgM42GbcPjENVhZekgs5eaNEBSVVaOirhllNU2orG9BeU0jIiJTsG9PIFavcoOlxW647w7ClauhiEvIFzleVd+JxrZXaOt6g46ed+h88QFdL0n979HU9koQUlzaiJS0Ejx6koSTgXexzfUoVpvthLXFTuHOmJgM1Ld0oL69C3nldUjJLUVCVjFS8kpRVtuE7NxSuGw+hP5fzcOSiXtx2DEBF3YXYLf1fSxfvBkOdnvxs38dgUP+F1BYxCI1LR9AS4cyn82waIHBSJUPIGEzn6L8BBFBLC6lYqACoSgilTMeziQzRQ7tWV4TEcQ271O+AeUffYCAIxcxbqANHNcdwvLFbnBZEozNJtfhuuQWdq14hCnfb8B3/ecg8PgNVNQ1ITG7CE8iU/AsJl0UM98vXA/FxeuhiIjNlghpckoeThy9js2OPrC22IXVq3bCysIdttaecLDbD0d7b1GUfF9vux+2Vnthvdodlma75P4N9vvhviMA587cRUZmMbpevkVpVRPuPU3AuashuPs0Acm5ZUjMKcHz+CyZw6OIZDyLTUdhVT22bPTDpIGO8FwTih0rHiDILQMzx23CkO8WSULm2OHLEh0lHAgn+kOEB+Gn8ikM6knxgs4XItwoUQhTfiZB92FlMrXy+w/vhXVUPJzutWFZiFZWoikZlRNmCwuLxoivzbB95T2st/HDSpOdcDK5Aoe5F7HB5Boc51/GiH5muH7jCXLLqnE/LB4PnyciOiUXj8NTsGGDPxbN3whL8+3Y7OAt4QqTuY7YsMFPFOcHAI3NnSiraEB2dikSEnIQFZWGiIhURESkISY2Eykp+cjNYTFWHWrqWtHS/gKdL96g+9V7vPwA3L4fi+XLtsJ0sTPWrd0jFRWrlm3F0sUu8DlwBUk5pYhIykZIdBqeRCbjYXgScksrsGj2RlhMOSr55X2WoTCdtg+Hjl2FycwtWDDDUeL+TG3SgKHnTz+JPhNlPnPchJ9S0hTnSknToqSuZdOLIMOEC2U8lQkbMUzbVhU2MT1H/fD2zRsUFRfAbs0uWM4IwMIxXliywBkLp2+F/ZzzsGPGasEVjO5ri4MHziOntAp3n8XhUXiSWEOnLz7C5PFrxJxraWqTsVR70fMCl8/fx5zp9jgZdB+1zT0oqWxGdV0H6pt60NT2Eq2db9DW/U7ET2vnWzS2vhQ9UFnXgbKqVjFfKbJcnY9JHvfZU00MGLbkxCzYrN6NVSt3IDwxR0xQcubj50mISslB4OmbGP+tHVxNb8NlyS0sGrsP3n5n4edxHUvHe2PVss3o7ukUi0YV97Jp8NMArBw5ihyKIOVnUS/w/j7Nzc1ePd3dImJUyIEcodKPxCA/a9daoawSRWSxZXO2wsL4qMRiFs/bglkTnGA9I0gQwETJjMlrkZxVgIfhiXgUkSTsfv5qCKaMW4O0FI2L3nR1IPRRFM6de4B7waFoq9O4rr6uCdMnWeHS5WeiTEsrm1FR3YrKWk0J1zZ0o67pBeoae1BT3ymlkVTSZVUtgiRHB1842vWWyxdk5IrivXzpEVJjNTORzc/rNObP3SBWWkhUKh6HJ8lcw2JTMWOCPSyNA7HB5IoAffu2Y3DddBjWM07DaPhGkQBd3Z0ieggzwoXvfFE/KLFE+FGakAuUiOLvfRKSPg7GKSVBitfyAVowjlygL67NyUZ2diZa29phPNoei8f5YO4IDyyZ74rp4xxgPjUAtnPOYfwAR+z3Po2YtDxBwOPIZITGZcJo6lpER2ru/YO74Ri0xBv9dz3HgEOZGLAvBqMtj+PcqWD5vaK8BnOM1yO3sE6smYqaVlTpEFDXqCFALCADBJAbbt6MxLKFm+QZL9o7sNHlGAY5XkF/32T080vDAKdbWON4GA2VWrUdxd+27ceFO8mlfEWn5cJipRvmDfeAw/zLWDx2P+xtvbB7R6Akd1ZMO4BtW/xRUKgREiUE4afqpmj/9+YTkv6uku6j2w0j8QtSOp0EiiKyEa+pMCiOlJnKjpR1lHGhoREY1n8VFo7dj7kj92Lp/G2YPnY9lk86CLNpxzBhiC3uPIpESEyqRv2JmdjnewEb7LWIaWxEEhasOwz7yFYM84nHIMsgjPB4hpH3OvHJ6rO4duG+3OezLwh7954VJ0shoKa+S0NAc4/eBOX3lTXtqG/ugdny7cjKyBdPds16f3zllYSJN+oxzPEqBtmcw5iLFfjmeAmmm/vgVVcXWprbMGeGPR6FJ+NxVLLMNyo1F5s2+2Pqd1uk+MtklCdWm+3G/n0XMO07F8nc2VjsxZu3NDOzJYvIvIeCH2FGaUGCJSyZRaQuoK5VPpQkZKhwWfVMtuEPZBmVUKBYInJ4TQSouhfKsJCQcAzra4bF48kBe7B43lYYjVuHpRN8JREyfeI6PItOEcrngiJTcrBipRtCRR5/wNw1+1Hcpu3OGf7dEnzyX2Pwyc9G4U8TXfHX252YYnsSLzs6kZtThBWmLqhveaH5AbWarU8EENj0AzQEdAgyYuPzYbbcVZ4b/jQafZ3v4/tTRfj9n+bii1+Mx5e/HI/PP1+IWRfzMWp/OAJPatzmuM4T/kevIiwhA48ikxCRnA0PzyBM6O8A65lBmDvcHavNduHQoRsw+utWzBuxB/ONHNHR2aGlJGuqBX4MslHREmYk4Lz8PPme8FMJGVZUSEKG+QC1xYaUTbFD4JKFaPMrK8iwdI8KhoG2vNxiDP3aHEsm+GoImOOKqWNtJQkya+guzDbegGdx6XgcqcnU0LgMLJzvhMryWjRWVmOw4yX4572A/74gfPPnmfjs9zMwbOBiHN5/Bocz2zBk52PEhCdJQtxs2VYUljYKlRPQFDl/I4JqO9DY+grXbzyHs5O3zHX7vosYfLYUa3zuYPaYVfjst9Px1Z/n4Ms/GKOquBTNb97BxvWE3MuIrvO2IwJ4zpeWkd+hSxj9jTXMjQIwY+gOWK/Zg+PH72DKQGcJfU8abon4BK2496OUJFO6BvAzLL38KB/Q2sqkfI9oZNqmxJIKFvGaGFRJZSoPYlarduvEw4ehGPb1aiyZ4Id5I/dg0WwiwA4mY/djxpCdmG3khGdxaYIAvp5Fp2Hxgs1orG9BYUYuxro/xsTrtfj1Lybjqz/OxB9/NRUhT3qtlR0XI3Dl+jO5Xr1yO/IK60TOKwQQ8OIJCwK65Hs6ZJcuh2DH1sNaP9dT8Imp0j+Txby//+Vk/OU303D23AM8aXqPpVtOyW/Mcm12OYSIlGw8jU7BvZA4OG3yxeh+Vlgx9QiMvt+GdWu9cCLgDiYN2AKTMd6YNtoOVdXVshVLaoeam0WSEIYKfgzdKF9LgnVUwC3NWlJelaXI/oCyMql3lP0BOVq4mQ8gtrRwtVZWoTbbMc4/5AszHQI8sWj2VkwdYyt53DkjPIQbHkck6WUqKcp0iQvSU3LR2dSMyRvPYnEy8P2ULfjsp2Pwx5+PR1KUppzZ/E/dQWxMBlqb27B6xQ4BcqVOBCkE0CzVrCANAfz8LDQNDjZ75Rk7Pc8iPE1TfmxXz97BL/9lCD792RiMdr2FcffbsH57kPzG8MTe/edEVz2JSJbQtNuOAIz4yhLLJx8Sql9v54Njx4IxaYAzFo71xaThNlJ6kpObLWJbCzcXidjmNWtn6UOp/RWyx023x07KUj4qzDJIGEhhUnz8301JkgsqKsqQGJeOQZ8tx9KJfpg3yhOLZ2/DtDG2khqkZTR8gDkuBYcgNDZNrIrIpGxsdjksZh+b844TmHqjAsvD2zBs7k58/tVSrHQOQv4r4FlZO6wcD4iNffn8PWzZeBBtXW/FDCUXkPKbWl+iue0VGlteoq6hWxBAE7W8qhWmC7egtbUDGWn52OJ+Gj1MDr15j+V2B/BF3+WYuPkSzPOB4dvu4tnDSJnPonkbcONuJEKiUvA8LgOnLtyHybxNmDjQEcsmH8TEAZuwYb0vjh0NxuRvnbF4vB9Gf7caISHPpT+tSFo+PT3dYkXSKVPRYoZ2lBVJAtc7YtwfQPGiUmZEBq/56k2f1Yl4IqbVfa1tLXjy6Dm+/2KFIGD+SE8sme0Go/F2Io444SFfrobbrgBEp+SIc/M0KgV3nsRhlpEdmpta0VLXgLnWvph6vQwWeYB56gcsSgHGBVdhyCo/ZKbmCgLmzViHsPAMqSOqa2IArlvi/UUVTSgsa0RZdRuqGzrR0PICtQ1daO16Cw+PM9jtplVqHz8RjIkuF7A8ugvLE9/DKhMwywaGejzBth3H5Z5L5+5h2bKt4gs8fp6I8PhMuO8LwqRxVjD+3hWmkw5gQn8nbHLyx7GjtzB5oIaAkd+aISkpVZJZKv3K8ANrhlQqUoUfVNqXKV1eswpd7wkbKgnDHRzcWGAYjFM7WNiuXbuNoX01JTxv5F6YztuOGZPXSS3m0om+mPa9C4ynrhNWJkuLaZeSA3fP01LewdbV3AKXHccxfctZzNj3ALN2XIOtcwBK87S50D533xWE9h7GcloQnZiPE6fvY9OWw7BY4wFrGy/Y2Xtjn89FBN+PQWZ+NarqO8QqWmKyBQ/vRchzbl5/igUORzHb4y5m7b2PRZsDcfrkTfmNDuG0ida4FxKPkJg0hMWlI/hBBOwd9mPkX82k4GvpRH+M7+8E5y2HcOToTT0CRg00R2VlDfLycvSesAa/3qIFfUq3U9sfofwswrsPd2wTG7R8qDTIKtTmZCNeE4NSH6+r7+d93ENVXl6KouIyjBhgIVYPK9dWmLhjjrGjVDQvnuAtiPnrX1bA++AlRCXnCAJo3sWk5WPj5oMwXbgZRQXa9qh3L7pRX16Nt12afmmob4KlmZvE6bkvKzopH65uxzFtii3Gj7GA6dJtWLt2P4yn2sPc3B3GRuswbsxqLDRxhv/hm8guqpE05MJ5m3DI74I8k629vhFNVbX6z9cuPcLEMeY4fy1EnDAq3ztPY7DTIxBLFrtgbP91MJ3kjyUT/TCunyNcXI/g4OHrgoAlRMC3a/DsWYTAhSKIuXRKFFpC1KsKtqpuiLJf6QjZH8Bt82QRWjX02tQuRj5EFV5RacgDKisFOVqUlNtsajFy4ArMGe6O+aP2wXyxJ+bPdsL0wW5YPMFHEGA8eBvGDl+N20/jEBavmaRPolIQm1GAA0dviIzduN4b58/cwYO7Ybh49i62bj6AZYtd4H/oGhIyy+DueQZTJ66F0eR1cNt5Co9CU5BdVCt55k0Oh3Do8HWk5VXi0vVQ2Njux/gxVpgx3QFHTtxGZEI+Nm06hNUrd8J//1ncCQ7B3eBnOOB9FubL3LBqxQ7ceRyL2PR8hMakIzw+A5eCw7B8hRsGfWUq61g0wVfE7Lh+DnDdegR+flcw5dstOgRYITycGUPmhyvFYCGQCScqX1pFpHi114C/0ewn9fO6Nx9gYOd/JIIMUmrKT2BHtsSkJIwdshzj+zpg0TgfWCzeh4XzN2LqIBcs4cTH7xfqoS6YOWM9bj2Jk3Avo45cLFN/USl5Umy1a3egZKhcth7F8cB7eBKWigNHbmDKJGtMGL8GO91PIyqhAGXVrRLnKSyug72ND65eSYSL02H0vAIaml6guLJFELRh02GMGmEuiDgccBvXb0fB58BVbHI+hE2bD2KP5xncehgjqcjY1DyExpI4UnHu6lNYWLhjxBBTjOhnjqVSce0jImhsv/VwdT2M/fsvYspAFywZ74sh36xCcbFmxNAHoOPKRvhR2bLRaFGbAqW8patLLCQRQYoDKN81DsgTc4nsRA4g1viblpTXSu/ISiXF2g6XNSu3o///LsOM77djpYmHhHwnf7dFqJ8IIOVMGuSE4YOXY4aRPfb7X8HdkAQNCbEZEhuKSs1HXEYRnoSnCtVaWu3F1ElrMXP6emzfFYhnkVmSYqTpWV3XjtrGLqRnlsLK3BOHDj6Fk70/Wjpe6wJ0XWKeFlc248nzNLhsPS5iavIEa1it3YejJ+/i/rNEhMZl4VF4CoIfxuDijVB4+1/BKrNdGDPCHBYWHpIVG/NXKywax3VoCBjTdx22bjuCPe5nMX3QdswZthuDPl+CyMh4kfWq3JAEzCItTYrky2ZtXjMLJkZNQz2qaqo0DtDrgDcsq6AOSNPvw+VGPSJH9jjpShd5TXGkzlI4deIGPvv5XIz4xArzpm2RdOPEb50E8AoBY/utw8YtB7HLIwiTx1tj3hwnWFrtwXpHP9ja+WDZsu2YNWM9jKbYCtDNzHfB5+A1hEZnIr+kXkxPRjjLGA1lLKi+XQC91sIT82duhZ/3ZcmMUfmKGapLXfK6sKwBz2OyBPErV+7EDKP1MJpkh5nG6zF3jpOMN2WiNYymroO5hTsOHQ9GWHwWbtyPwKiBa/QIICcP/9Ia27Yfw/Ztp2A0cCvGf+2AwV8uRVhYFMorysR0J7wohujckljJCSRoiiHG0aRgq75OtiuJDvj7hVk5Bik1FhZpWtzQCiLw09NTkZmZgyH9TDHot2aYN9UFllbuGPaXNb0ImKTJTnv7/ZL6uxwcBrddJ2Gxxh2mptuwYsUOWK/dC1e3AAScvifJ+oTMEmQV1aCwvFES8kWSlG/UEKALRze3v0ZERBbcXI6jrKIFDS0vBQEMVZRXt0nomkn50oomlFY2Ib+0AUlZZXgYmoyTZx/Ac/957HQPhMe+MzgaeFs4ISwuC1FJOSKSLgc/EwQsnuAnCpgi9a9/XAHnrUex1soLg/+4BiM/s8GwfqZoaGRa8oP4UEoEUQGrpBVFjYIndQL1LYEv+QDlB6iUpKEfwGu151WVKvJ7YpGb23jd3dMBf58g/O4/psFozAah9H6/WYjp32/D4oneWDbZH5MHbsEaCw8kZhVJ2YdWAJWFkOgMeY9OzkdCRjFS88qQWVSlVUUUVSOnSKuK0COgigjQKJtOV0vHG/S8gQC/mp6w4oAqckyz9GFfPoPnO7DMhc9laUpyTpmUobA2SEoUWaCVnCvzY7bu0s2nogNY4j53hDum9N+E/r9diJ17TsF85W4M/oM5BvzvcpiZbpN9A8pnInVTbxJmhvBsaNT5CNyy1NIsL+5F+MgTVp4cT09h0p37W5UnZ7jTnVs203UYleLc4gJcOn8Pk8etwYihyzDgjwsx5gtbTOy3QRAx+qv1WLpoK+LS8/E8PkMQEKMrzCUAkjJLkCyliaVIyS3T6oIKqpBbXKuVlOgQwEoH6gFJyNTR1tfkPXUCryVKWtcpjhrFkFaWoqsLKqpFRkEl0vLKpf6UGzSIACnO1RVmRSawai8L0Wl5OHf1MQZ+ughjPrfD6M9t5cV1HTh+A5bme9H3F0vxp18YIT42HfUNtXoK5+Ek8fEJIo4odghPwlXt/Kcera3jTn9N2nwUC1Klh9pZC71nRRD4fCA5hbsJKYpYesjvG+o1BFVUlCIvPx/Xrz5Gvz8vwJgvtImP+twW3//BAjOnOiIyOUc8zN7SxFytMo7FubrqaNYMpefpEFBUi8JSIqBeZH9MWhGyShvQ8RJoan8lQTgVkKNI6nwFlDZ0IjK5UJy20opmPQfwWQoBrIRQlXEsjWSFnipNDE/IFEIJOHUL/X5vgrFf2GPMl/YY/cU6fPfFYmzdeQJjhppj+IDlOHv6hjhghMfrN29kQzjhRRgylkZ/ideEqyrl5B4BiipKEUGAnBf0d7coaaWHFD2MAfGa3GC4C5Df87vWNm2LTn1DHd5/eIM509ZjwK+XY+xX2uRHfGaDsUPWSL6Vi4yIz9IQkKRDQMbfIoBVbFIZV1KH6oYOPHyShHn/OxVW/ZbAc3cgYjJKUNP+Ck0971Db8QrxmeU44HMZtmMtsegnY+DnHoS61pdaaWJJHXKKapBRqENATi8CflgbSgQkZBZhn/dZ9P21CcZ96YjxXzlg+CeWGPD5XKxa5iKV0aWlZXKMAuFB8d3SQhiUCGAJGwJdRUQJV7VlieKIfZR/9YPzgrR9wDSPtHxApD4Yx7oXFUwyLL3jgKySZiMiM7PScfdOCL7+/Tx8//vVGP+1I8Z+sR5//XwJrt55LtTG0kSpDWVxbooBB2QZIKBQE0EEXn3rCwQcuwWHPkNw+F/GweEfhsH8t8awm7gWW0xcsG6CDVb/bgYc/t8weP9oFPb3GYr18zajRhDQiIJickBNLwcQAZn/BwJYG5pdAls7L/zpxzMw8Ldm+OZ/FuPz38zChfO3ZZ08y0g75kCL9efn5evrqEic9JVI+YQjzw+i5KCSpgGjYE3CpaT5aIsSKZ9YISYJ7O6uLn1G7OMtSszodPzNTnA+nNc8QeRM0HWMHmSGL/5nAfr9aik+/cUMHAq4hYSMQpGzgoAkVZyr7ZBJzi7REJCvIYBym/Kb9v29p0mw+OkEHP3PyTj8s6nw//cJ8OgzEm59hsm737+Mx5GfTMLRn03Gpn8Yht0uAWhof63nAEMd8BEHpH5cnBuRmCX7BGYa2WPx3E1YuWQr1phtQ2ICZfwHca5IzWpPBWFAAPMzrR7CizAgDAkP5TepDJmCIQ0a8QO4Rentu3f6nLBCBhUtbyD7yBYbXfk1lQ2VipTepfMIGe0sCQ6uSu+0qGqphC5OnbwmpYeTRq6GnZ0vYlILEJmoq45WOuD/QACplggoLm8U2W45ZR32/sNIeP/7RBz75VTc/HIuHn67AHf7z8e5P87A/h9PwJF/G4/lv5qK0LhcCchRf4gOKP4/EKDjADWf+IwC+B+5iXmz1gvAuSmjuqYSb9+9lviY8oe4TgKbkoDESLufsGHwUouXFejrP3k/4clrJmPYh8gj3D/aoqRCDeo0k6hoblHqPTLs74UqOAkVLWW5Nt3xmmpNwTA2kp2ts5aqG8ThuXInSiKi3KRBqlMI4JYlQUCeDgHUAUU1ArzC0no5KuxxeAbm/nwSHn8xC61LbdC+ch06zO3RbmaP9hV2KJy4Ehb/+D083c+gueuNgRlab6CEyz5GAMvTZZ9CPu4+icWRE8FYuWonPHYdQV6+VmjLxuimOuaM5iPFM6mYjcBWRxuQwqlkKcIJR4odEiw/81r5WkqfyhYl2c1YrcWotXi1diYQKfmHW2w030DbYqPKFXt9A81XoO9AT6+2rhY8iaWO23HaWrBs8SY4OR8VZ4vOzkdW0A85QKwgIqBegEguaO58g4s3IrBj7AqkGS9D+1IrdC2zRqepNcrmmCNg5CK4ux1HdctLyZyxolpZQXmFH+sAbonSdknmIy41XxB/4VoItu8OwryZDigsKEJNje58pGrtjCH6Pr3xfg1Omp2v2f9MS6ozgXhdI+Wb2jYv8YB18ORpLXxnYv4jT7h3E1nvFhsOqDxhYpInSkk8+/07bROeQekizS42xjrIPZryAUpKS0U5P7wXCqPJdjh14QluPYrVFp/OHTKaP6BHQF6FDgFUwhoCCEx6ty2dbxGRUoI9LkewZ/Y6HJxmiX1Glti5ejeu3Y1Fc/c78RHEEzZAQC4RkP+xCCLiOSY50Nf/Es5fD4OxkT2uXLgnp0FSClA6cP2kZmXXKzjV1tSIxCDnSx1QfJxwPRslBE15Nj5D+QkU44ZBuz6dne0aAgyO1KI9r0oTKyoq9WchKEeN8p+OGvPDfBDjQ7QCVOERkUJ242DyvLw8+Y3nATnY7oHJAhecvRqKy8HPEZGYI06Y7JLMLEEqd8nkVshuRmWGUpHSqWIogvuCufOluecDcirbEJ9dg7TCJtS1v5NSRS0Uwf0BzSgub0Kh7JDRIYAcIJtBNGsrp6QGT8JS4LjBD57eF7Fi5U7YWu5CaWmxPu9NcauVFr4TGW54+iG5gs6qWicVsvKZqCMolihmVFkiZT85hvAikfJaRBC/YOfKigrBMjHMjoz5UMzwN35PeUjO4LuygshaKnLKODf78CA/2UPc2CifqTv4HAb7CvLzpWJtydJtOHPlGU5feorbj+LEHOXmPFIpX5kFlbI/jMDrRUCLxIIqatv0daLNbW/Q1PpKyhS1iol2uUdDQCOKSqiEa5FfXIPs4mpkFVYJEkIi0uHlfREODr44fOI2lphug8UKVxEv3JhN7uWa1Z5pEh3XnJuj7ZMg1VMkZWRo6+S9XCcJlyKdn4k82aT+8qUuWZMhcCBylAjvk5SS5EV5zSa7IGNj5QYihXuGlR9gGKwz3D9ABKhD7zgwWVKxF+/Xjj54qe0g1/UvLSmG+TJnTJu6Dn5HgnH2yjMEXXyEG3ejJAKamFmC7KJqFJTWawE1HeVLQl4HfFo4jIjWNb7QVUZ0y2ctZK1DRE2rcE1RaT0y8ypE2T4KTcalm2Fwdj2GnR6n4X3gKubP2YiN6z2FQlnnyUaCIrcTDmzkApVmJHfQX6J4ZqOIieFhh7pULcUP4cVGbqAPRaSyEbFiqChPWAXjqL2pmYldvqtrYopIoGxT9e0cmC9+r5S0Kj5lHy2AV6+dpaOeJ4pa7Teuld30gQFXMG3CWqxcuQsHAoJx5koIAi88QtClJ7h86zluP47Hs8gMxCUVIC2rDNkF1aITSiq4W0YF5Rgb6hCxQ5HDzXhp2ZWITynE8+hMPAhJxNXb4Th96QlOnnuE05ef4Pz1UKH6Zct3YOJYM1w6f1tOwCVFc+5cB40Qxm1UIS3XyN/5mWsg0LV7ew/r45r5u1LWzU1a7Q8pvXftLNbVjjggwX+0SU95uiyZMDyMw/DEJ3WyLhupXSlhYpkleGw0XZ8/f65XwlJJp9t9GRERLvY0W01tJeJiE2TnyvRptlhgshluu0/h+NkHOH0lBKcv8/VE/wq6/ARnr4QIAC/eCMOlm8/ldfFmmCRVzl17hjOXnyLo0mPt3qtPcZbfXQ3BiXMP4OV/BXbrfTFjuj3mzXLE6ZPBcvBsWnoKmpq1Ilqe/RkRqduIKMdlZujNbDZKBaWECVDDddLWV6f4srGfyoTxeXyuOjOJiKFo68Mj5SnTDeMUEutpaxOZxkE4AMUJkaFiGypWRFZSsSICmH1IEbymKOLZO7zWuIiVd83yzvv4PWVpW1sz8nLzcfXSA6yz9sD0qWsxZ5ajJEiYEfP0uQAv/8vCJUdP3UXA6fs4fuY+Tpx9gJPnHsg7cwlHTt6WPPNe7wvY4XEa9o7+koQxWbAZC+Y6YdG8jdjksA8hj6Nl7xaPHiCxcB6cMwmQn7nFlmvluvgb10+qVfEchpYJJ3K4WufH95YLh/Berpew4z2EHZWv1pfH/HRrO2Q4GBsVBaldHLF3mpmpHBHDwixDM9VQB9AbZB8l3zgJpVPIPZwYGydA6qBVwUbZq0zY9+/fICw0AneCn0itzmwjO4wdZYH58zZi9gwHyWAxtTljmj1mGPOzI2YaO8B4mj2mT7PD3NmOWLp0q6RAVy7djIAjF5CSlI2qSi1k3NyiOZY8ITE6qldfVeq24xIGbKTOj+eYqjcleY+hHCeAuU4qXjbep3Qkq8gJH66ZTczQaAMdoE5LMSw7UZqf15WVVaLZec3OahcgFU6m7ALUdlFSMZOqec3veB8pQoU2OElyltqFqb5Xpe98Hi0K3keT7+3bVygrK8bs6Tawc/CB35FrOHXuIc5eeoozl57i7OUQnLvyDOcuP8PZSyE4c/GJZLr8jl7HgWM3sXChM44cvKQDEK0wplO10hHOQbNGtLnyMymfcyIM1NGVao4MLdOQUOEGEhZ/o8RgX7VOEigJk4jgmggTdWpKe0eHPJfjajDSmaH8pwcigDfzC2pt2rmkcloApFoOymsCnoDl5DgIB+cC2Id6g4jjNSdGjiH7Sel7fr4ghwhQ/fk9J8d+CgE8EZ338Rk8eTc6KlZOqN264wRcth2Dj98l+B24gmMnbiHozAOcP/8E5y88xZmzj+TzyaB7ktrcvPUYTE3dsH9PkC51qlUo8LlUfJwDAcE5cP6cK0UkfycMSO30Ut+8fYsC3X4uAo2Sgc8iAtiXipSnoFBKCNwKi+R3IkojSuYEtO2oalzlBygj528O7aOjwc4EinZ2tKac2Flt0iNVKCXNSRARXCgHERGkiwUZsqaqJ2LjpJRoYmN/JdIoF/kMUi2bvc1eqYDzPnAZN+9F4WlEKsKZu43PQ2xiASKjs3HvQRzOXXyCQ0euY72DD8zMduPbr01w5MB5VFRoz+GaDA0NAo1GhDIl+c45KRFEAvyhmFT1nJopaWCGlpfrNzeyETa9IkhzWPPzNSdVjVut66udFaFTHqQCUjuxw4cR0MrRUnvDeK2qvPhwll5zQCKAwCNweR+fxwmS8jioUsL8ntTHZzMvyoJgIorUwGcT4XwGzV7ea2flgckTbLBi+XZYW+/BWuu9sFrjgbWWe2Br5SlbVDfY+2Kj0wHs3BUIZ9ej2L47UCoujCauloOUqOg5B85bzYFr5DjcpE7RSoXKOfF7fub6eD/7cV6Gc6S4IoIIG95LPSLraW6W30loPJqTn7ke7Wh8bfuq5uj1nrLyUWEWAcMJyj4nnVNBylRmmDoHjYOzEeA0y7gFR/3OxkXwjGa9eZZL/aBZCTRDyaLaszP0Ji4bKV89++XLHjwPC8eEUebYsOkAtrsHIvDsA5y9+BjXbkbg7oM4hISmIjwyE6Hh6XjwNBHXb0XC0/s8NrkcwibXIzCeshb1dQ1IzzDgsJ5uMaWV80mg8DBZzo2NIpB/AqHM0Kzs7I/MUK5fcRFF7fPn4XquoRnK0ARFOhsJkdtYee5QQyPPFKLlqBk13NbKJjlhXpAaCEh2IiJU9FNF8tR7b8RTi/oRYQxK8Zq2NAcnAghMfkfO4POosBkd5TUdGP7GZ/HFseigqIhhW1sLel504dnTSPT/fDZWrtqFQ8dv4sbdCNx+GCOvO49i8eBJAh6FJOHB03g5uO/KjTDs8DgFh43+MFnogn6fz8bz0Fi8eNklx85oANEcR/17Xe9nOlDqICq1dlYxqH5qvUSWErm894cwomSQeI8OhoQLk/VcN4mSwGdIm9yj5wB6anQUDM0wUqeianKIiodTqZASVOkdP1Ouqc98sKF81WRmbwSQFhd/V5TC83aydAfBAu9x7MhZLJ63HoMHLMGqVbvR7/N5GNJ/OVYs2gUbq33YtPEQduw4id27g7Bzxyls3ngEq1d5YK7xRkwbtQFDvjTHZ78zgsMGf8lsWZntwJOHUYiLi0dzi+Y4UvzRiVJzosh5/jxU0rJs5MQfzlHJda1pcGGjyNH2BWiczaaqIBTnvejpQUREpF63cfe8OGI8qkBzkbUjhrmjg/KeWOt1KrRzMwlYsijlGa95T2+grVW2tLIvqYDPIvZp33ORytETZ0cnZzUK5GGpb0UUnA68gdUr3OBguw+PHkRj+9YjYv977LsAo4nrMfgTC8wYvB3T/uqKSd9u1l4DtmDKQFfZp7xkgjeMB7li8kgbXLrxDIeO38L0qTZyuJLLRl8sXeCE3dsPIyoyHp2dbXjz9pVUKfCERM5VOVXKYlO6gABW6+c19QHXqQDO+2kNKbFFmCmnjYhWBQ+ECWFNGJDzxBHT/UGaTERhk2EF/sjGMAIxT04gVkkJNFnZOAFqeFUBxoAbrSjl2IkuiI4SL5KNQOb9POZFm3gxwsIi5UyI5Ys2w8neCzHRGWju6EFxdSMa27tx9dJjzDa2xxbXozAz24WhX1vItljz6QFYMeWgnKC7fPJBmBkdxZhv7DB9sg0eR6Rgq1sAFsxaj6TkPDR1vJCzQfPyK3D00BWYmjhJTWtEmHaiFTNfPPNCAS8hvnf9XCstILX+1JRUATYlAgFMUUSA0hAhApTZq0IQtKJo5iq9wT4flaWo/5Ch/FMP5UDKuiGw+JkOEwHNawKekyIV8DMnq5mp2n+o0P7ns0hNb9+RuhuQmJCgH7SwoBjuOw/BdOFGfPmnGVhn5YmCggrUtXRIzD6WZ8hla2fI1bV1IT2jSBBkbeMJ30NXYTTeHiO/tMHC0fuwcsohOax7yOerYWO9ByFR6Vi6xAUb1+9HbUMbSmqa5TlyKGBeGcrruYu+C9evhuDPv5qKiaPN4b7jKKKjklBSWoyXuv+EoU7kmrg2+gMUSap0n4DmOskdvIe6gJ/J7QQ8gc5jPnnN5/A3RgKICCJY9GSjzg/gRm0qDAKLJiiBxwHINsQ+J0AvkOFUPkyFFYgQcgvv5QTIKeoz2ZcbFQhwfiaV5eRkiXw9E3gb82etxyH/i3JGz5/+dzKu3ghDVkmVJGdYKMWKNQKNmTG+Smta0NLeDa89gViy0Fk2U2zfeRxjBq3GsM8tMbTvMgScCEZIeBpmT7dH0MlgdHS/QWF5g9QCpRdUSsUdnxudwixcIeJS8vH1JzPh5OArJ6osX+QMkznrcP9umBAJrReuUVE1jQj6AYQHv+OLCKCMp3giXEi0RBKtPBIp4Uci5W/UrUQIpQP7UjpQLEksiDepmA+pmCakcip4SrrKiFF7U8QoEURM0+xS2TMOSvZSeQNSPpUvayfra5uE9dfb7pWTT94CePg0EZ/8ZgpOnX8ohbssD+FBqqwRYtYqPb9cqiNyimvkxKue1+/x8H4UjKdYIyDwLjLyK3Dm/D0kZxTh6PFg+fus+PgcdL54J4n8vJJaOVWXOWaFAAKfhWDcEf/5H6dj08aD6HrzAS3tPXj0MAaWZtuxcK4DsrM1x6mnhxZU+0cZP8KLCtrQYCFcaIyw8V5VN6XyAarvDw0SyQeoWL4Wy9aUEM0mIoUYI6sop4JUrhwQ3q+ipOp3Ffnjb2Ka1VUjOioRc41tcfP6M7x5D+1wpcZuXLn+HJ/91hhXg8ORkFX0NwhIUwgoUSWKDbJTsrS8HpZmO7He3kdOQFy/zkcAV1XTKiWLzISxqloQUFQthV56BKQVIC6jEE8jU/HNp3NgY+WJstpWyTF0vniPV2+B+/ejMMfYDvduP8fr1zyWTKts47q4RoFRaalQseQKqqsELrwmzMgZLMnhNQ0VGjF8hsC5oV7v1EplnDJDmXygWUZxomE1T6q/lBlK5amyPPI5KVFvllGmcVe4MkNJGXwWHZCKslrMn2mP+IQcNLS/QEFpg2S4Khs6cSLwLj77nTFuP4hBfGaRJh4+4oAKPQeQorVy82Y5noDccNDvIqaTGw5fw8s3kDpRw2S8hoAqEWOCgEwdAtIKEBaTgYFfmcB08TYUVjZp2beKJtmPUN3MrFoLVi11wZlTV5GTq4UgKEpY6abyxSQ2+UeRV7rcQUaGfjcRG6VDbz7gBcIjIvTSQXn8ffh3rqRmveNh4IjxpZwvXhPjdFR4jzJdq6u0chZxzHQnyKpnsXB38lgLzJhig5t3IqXsPLe0Vs4F5eF5/v5X8PkfjXE/JEEqFGKS/y8E1EpaUfLCkppkNqwDnS8/oL3ztQCfG7X5PX/XI4AnJRYaIEA4oFAO7GZtKv0MhrXJJeo0xfyKBukTfDcKq0zdMGOyNVpbtXwBS07E2aquFipnKYpaa69DqcGML8JIk/W9pxDzfvalg0dCleNqVLaKjaYhsaxsWlpChq44ZZ06SZaNWFZxcuoNUgQLV9m4RfRXPxmLT35tjD/9aqr8/dPTiDRtD0BJNfZ6nhEr6GFokr5Cje96BLA8RVcjqhBACmf5OcVFFQ92LahCXkE1GlteoK37LcprtLJ0HlcpCDDkAIWA1HypSx05eAXGjbCQkhjuXcgqrca5y08wc5o9Pv2NMX71kwkYP8JCMz9TUz/a80UON9QBdEw/dto0acHvxGnVhViUdKBIYuvT2t7uRc1O7KhgFGUXKZnswmvlkPGlHBJeV1Rq/71I5FGuUYxpzka9yH6eqP7bn09G389M8PnvZ+KLP83AuJEWuHw9FKmFZXDddgRffzJL8rbqAFXhALGCdAhQ+wRK61Fc3oDm9lfIzK2Ay+Yj8q9Ik4fZY+JQGzm+/tSpe6hv6ZZzhUQJy4aMv4MA3ThTJqzFoH6L8DAsQUoSXV2PYvTwlej72Rx88utp+PQ3MzFrmp38x4wWhONZD6oaukWonva9odNG4uM1X4Qfv9OcVO0/ediHfZVEMdABlaKdFVaptXkeMv96g000u64qQD6nper/R4ZmLGWfYSgiPSMVt24+wl9+NxN/+d0sjBqySk7K8vK/hHGjLHD6ymNscT2Mbz6djTuP4xCvO0T7bxCgE0GsD+IJuk+eJWLM96sx6FMzDPncAv1/b4oBf1yOUX3tMOQvllizyl12x5RVt2gIKNZOy/0hB8RnFGPhAhd8+aeZuHA9RP6CcRZ3VJ64DWeXw5g4eg3+5ydj4ecThELdEcxa8DFarwPEYjSohqAprrKDVJ0JiQkSzGPjSZO8t9dC1A7ukINbldVCEaJsW9rsFEP8wx06IRyEWpuJd4ostQmNJinzvuxLbqCtTO5hXx75u3TBRnz+hxkSr9+3/6wcWeNz8AomjbfC8mVuGPDlfNy8Hy2moXaauoEO0COgBuU1LUjOKMbo7y0wpp+dHJNjNMwRgYduw3fPJYzsb4Vx363D4E/XYL2ttxxr8HfNUEFAgZijlpZ7RQcZT7XBzOn2uHIrHDt3nZAN214+FzBptJn+z6NJXOVlZbJORlIJI1I1P9MCIhz4d7b8TCDzfsKAyOJvvIe/aT6StlmPhMu/NJfiXBWfZmfKeQJXw6p2jDFlGTuo5AmdE/oBtIxydSV47MPzR4lEWk8vX3bLLpI//HKyFEI9j83AbvdARCblYr2jL/7yOyMM/tYUV25H6BCg4wDWieaUiGLUNmrUyCHc/P+BwX+xwIh+/LelHXh4o1cX7Vp/GUO/sMTEQQ4Y/s1qPHwcj/K61l4doNsTxudTz3AOG5wO4A+/mIgh3y3FyXMPsWPXcTyLTJX9C9zNudZihzyba6EuVOEEUj4PLySns1EspSQn65M3/F2F1enEaqEIlRNuEWmh9wPUFiVSLR0uyictFFGsS69pf+lEyiYmFSL42TA0of6miZ+pEyiycrJzYW2+A25bA7B4wRbkldfiyo1Q7PM+j3uhSZg8fi36f2GC89ee6fdq6TmA50jnlYsnm1dWj9zCahiNscfwr9fAaOgWzBvuifhIDfFsux0vYM6wPRg5wAojv7aBm2sAapu7eq0gtScsrVBqQUOi0yV9+elvpsLd8yy27z6JyzdDcfN+JKaMs8Kp0/flfzEP+p2W3LRKKxKY5HSuk4TLzyoUQQNGpVQJI5qg5AT+xnsIX30oQldbpE/KEyMEGhHAzuQEApnAJWYJbJUnfv36FfJyc+Xh/My/NCRbsQ/vJQIKCvJwLugGFs52RGR8FrbvOIk15u7IKqmB576zOH/1iWxXHfj1Qpy68FjEgmYFGThiREBBJYoqGxEdl4MJ39tiVH8rzBuzC7OHeCDQ76EA/+WLVzCf6YfF430xdqAdRve1g7X5XlQ1tAsHaAW5JXpHjCbvo+epUhM6cdxq+Xds/0OXEfwgGmOGr8D5y09wJTgMrlsOCRfwv2Iohgh4SgBSPtfJqDCr5ag3CQeKZsKEv5FwqRNI+fyN97AvpYOWlNdOpzSIhmpmKIFIs1MpFiKBYkYfDUxN1eeDGaRKSkzUhyaopNhXKZpH9yNga70HJ0/fxvFTd7BmtQccHX3lmALGcg6fvIUxw81w9OQd3X8J/MATziUCKlBY0Yj4xDxMGroOo/pbY/ZIN5hy7+7EfYgKSUfA/juYNdQdy6f4Y/QAK4zpZw87Sy9UN3YY+AGKA/ifNUV4EJqEtWv3SOCO4jD4fjSmTrCUhP/hgBs4euImbt+Lhsf2AFkLYaBKSygdWIClMn78nspXiSASozLNVS7asCyFZqleBBUVFXnRBKUSpsKgKCKl80YClAkEKmLKd/5OOcgJ8DPZilgnwPmZ3EMKoIJpa2vHJof9GP7Xpbh5L1Ioyn3PaRhPscNer7Pw8DojacZJ49bA9/A1RCTk6LzUXiuIlCtFuiW1stlivtFGjPzGWqh88Vhf7Ha4hPKSWiRE5mCl0QHMHb0bIwesxogvLbF/3znUNnf+HQQUCgfceRoPS0sPjBq2EkcCb2HBPCdsczsO34OXcersPZQ1tMF0qSscbT2lnIXrVkc18J1ER72ptiAxEkrDhflnBuVI1IQfKZ8w4T28l33YlzB/aXhYB81Q5kqVGUrA8q/8GJplo4xXxanaZ+6S18xQyjyGLVQBKourHGx3YavzQTx5noypE63h5OSPi9eewdn1CAYPXCKh5TWW7uKN+hy8KudGSJRSt2HPEAE5RdVi3/v4XMT3n1lg3Hd2GPmVLby33UBP5ysU5dZg8SQPDPpiBYyGumDkQHMkJOfLWdGihA3N0HQtFHHzYTRWr3HHoH4LsXihsyBgn/cFnDx9RwyAVSu2Y63VHjk39Njhc7IuSgUWVXGnOxuJk6alClxSWqh/12Yj5WdnaWYorUbmRgxDEYU0Qzs7O2WbKlOSuXl5wgkq10lMks1I+cQoB+Q1f+eDqER4TW4hwigTGTns6GyXqrRtWw/j6q3nAtSdu05i/pwNWL7cTSrbBnwxDyMGmWL0kOXYu/8CnkSmiiWkRBA3TiglnFNcjeKKRpTXtmKpyVYM/9IGE753wOAvVmHSkHUY0W81hvddDeOhLhj45+U4evSG7CHIK67RI4D7AVQ0lBtDrt4Jx/LlWxF8/THmTrfF8uU7sHP3STx+noz5sx1hZb1X9i/s3HFCchFdneTwZqFmZv5IqAzCcd3kfMKFO18oHfgbX7ym0uVvSjqwHKXD4B/MpS6I8qlFpwOKS0q0AJNBjYvaIcMWbxCKoF7gn/oY5j0ZxmhtbYKLkx+OnQjG8VO34eN/AXcex+De01i4bj2GebM34g+/moQj/pewf+8pbNh8EA/CkmR/LhHAd0EAdYAOAQyWVde3y6a71at2Y/AXZhjx9VqM+sYOY/ray0EaI781x7GjN+Q4G3rNvcE4tS2pWMQPjyI4f/0ZlptuRXJSGpYtcsLSJVtxKOAmpvBUFUtPEZMHj13H07BkrLXYqZPjWkiZ8DKUFipw2RuKSNYHLvkdTdZ83V/dUkyxCoPxITbJCatwtPrvdQlH64pWyQmkbAk3NzejnKHWioqPwtMqHE3uoZMSFhqNGVMsMXOaLXx8L4p5dyjgOrz9LspxZTvcT2HqJFucOn4NPvtOwcHJX8rRuWVU4jIKAUoEMR8gmzQaZV8AAXz3bhRcNh3BmhUeWGuxD157zyElrUj+W4CHcxBR+T9EQHYxknNL8fB5shT0rli6VUxlty3+mDdnI2Ya22H+HCf4Hb6KkMhUBN+LwmqLXRJKb2rSApH6cLSucFn7zHC0VnqpwtHUEwpGhqFs9vk/w9Hh4b3haB4uqoWj5aOuKqK3KoCyzjAczTAG2ZHN3todN4LDEBmfIzF7loLbrN0LR0c/LFvmJon2CaMsEBEeg+0u/rB38Jd9AXTI6Izxf8YoMmiCcqNGHnMB5Q3642qq6zvkDxxevAW6Xr5H9yvINU9Q5Oa88motIlpQVifIIxK5O5KynX9REnie1dR3sXyxC7Izc6SeZ9XSLRjw1Vy47TgBD8/TMJnnhIXzt8Bjz2msMd+By+cfyNpY2aDyxUSCtgFFq/BjFlD9+TObFo7WpAO9YtZEqaoIfTiaW5QIdGKEmFOySQux8tQPdeqfFnZWIVb1WQvBqj8o00K2Fsu3Ij23DIVVTWjqfisH6V24EgIX12MigqaMt8JB3zOyZ8zL4wSsbfbh+v0onLnyFJeCn+PW4zjcfhyHO0/i5Y8W7oUk4v6zJDwITcbDsFQ84ut5Gh6HG7zC0vH4eZr89jAsRe699ywJ954misVz+0k8bjyIxtlrITh3LUTGWjTfCfm5BejoaEXQiesY/O1SLFywGRbmu3DxegjyyutR3/lK/kLFzNQF5eVa4dr/X+hebcTgi1TfC6N6VFZUSjWE2klJbumTlpZ2RLnUbESGCqqx8QEUM6oR60SUakVFxfIw1chJzk7+WGfDo95PYveOIOzacRq7dp6G85YAuLicgMmczYgM14JWEaEpmDxuLRzWH8A6Wz/Y2fjCxnq/vGzXesNmrTds1/rCdq2PvGz4uw0/+8F2rb/uN91nG8OXP+xs/LXrtf6wsfaDzVpfrLP1h/26A7C29Mbc6ev1Yffga2H44g9zYGXpDY+9F3H46G0cP/EI586E4srlCCye64z8PM2CIeBUJQibinSqRqSomig2VbaoGqWL4pr/D3B0yklYzExBAAAAAElFTkSuQmCC" alt="" />Enviar</button>
        </form>
      </section>
    </section>
  </main>

  <div class="modal-backdrop" id="nameGate">
      <div class="modal">
        <div class="brand">
        <img class="mark" src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAGAAAABgCAYAAADimHc4AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAFP4SURBVHhefb0HWFRZnj7s/ndn4+zM7OzMzk6e6e7paGh7bHPOgjlgwgQoICCCGADFBIpIMqMi5qyYI4LknHPOOUezvt/z/m6douyZ/c7z1FO3qu6555xfTudUn6SkJOOKqiqvDx8+eLW3t3vFxcV5vXz50uvt27deCQlxXo2NjfJbSkqKV1lZmVzn5+d7ZWamy3VNTY1cv3//3utc0E2vmJgYrw8f3nu9ffvaKyEhQd8/NTX1B/0z9f35bMPxe3p6vN6/5/gf9y8pKZHrwqJCr8xsrX9VVZVXkmH/BM6/x+v169decQb909PT9f3zCwu9srOz9eMnJSXJdWdnp6yZ43P9SUm9/bnGxsZar0Xz1nt9+tuJXunpWv+KigqZZ2//BH1/w/kbjl9YWKhff5/09PQjL16+gGpFxUX665KSYrx7906uKyoq0NnZKdfNzU2oqamR6zdv3qCmphq73I7B1mofHGw9P+r//v17ua6qqtL3b2pqQm2t1v/FixcoLy/X9yksLNRfl5aW4v373vHb29vluqWlBXV1dXLNuZcZ9C8q6p1/cXHv+JWVlfr+zc0tqK+vl+uXL1/KOIZ9VCstLdGP39XdjoN+Z7HNJQCHDgXDzfmQfN/d043Ssr/fv6SkRA+/6upqdHR0yHVLS7N+/n1IuR8+fEBGRgbKyspkwgUFBcjOzpZrAi4lJUUexIUnJiagq6sLr169QlJSEnpedOLWjRCYrdiOVx8AZ+fDcN28H7W1VdK/sqoSqala/7a2NunT3d0tgE9ISBBA8D7D8fPy8mR8zouITktLw9u3b2V89u/q7pbxk1OS0dDQIH0yMzNRVkaEvUd+fj5ycnLkmgvn+CSU5uZmJCYmyvjSPzkZ9fV1fzM+kaj6c/2ZWZloamzG/Nn2yCmsQWvPO5gudcbF8zfQ2dku96VnpOsRTiLKzc3V909NTdWvn/AjIb5+/VqIqg/ZllREbFVXV8lNnAg/85oAKCouFuppqK9Hfl6ecAB/Ky4uQldXO2ytduP85adoan+Ju0/iYTLXHjW11dKHACgqKkRbexsaGxtRWFgggOQrPz9POIH3aeNX68cn9bW2tqG2tlYWRuohsAlcNT4BRQSyP+/hglpbW4Wj+FLjk6D4Pfur8dvaWlFQkC/r47M4nhqf3MI5qPVXV1fiwb0wmJtvR11LN8pqWrHJ+SB2bvPHq9c9MjfOhWNyHM6jvLxMxhf4FRXJs7h+zoXz53383Icyl4OxsUNERIRQB7EXGRkhN7GRchR7Z2dnIT09TWOtmhrMMl6LkMg01DR04N6TBCyc54jmphYkJMSjsUnrT2rv7Z8tVM1WUVGOmJgYueZCIiMjdeO/lfEJNDZSvmJvbfwMua6qqkZMrNafiwqPCBdq5/yjoiL1oiY+Pl4vanJzc4Qr2CorKxAXF6cbX1s/xRK5Lzo6Wr9+X69AODr5oq65G/nFtdi9Jwh7dp1Ae3sLYmNjtf6dnTJncjfH1/o3yW/a+jXxmpubJ1zB1oeKgwtWFEQkaBRYivaOXgoicCi78wtIgc1CRaSmqsoqrDR1xfOYTFTVdeBZeBrmTF+HhPgUEVWUdWRJPpfAIGuyb3t7m1Azv+Oz+SyOz2u+y/jt7cLC7K8fX8cBBDbnpfpz/ryX16REchT7E8AfzT8/D01N2vx5TRFEkUB9xXE1DiwVjmAfUjDF6Ymjl7HZ+SBqGjtRXt2C40F3YbNmJ9rbW2UcElfv+GV6riUH8zeRIA0NIl619WvwEB2gZCgXTVklMvQHMoxUxUWnpCQLYHt6eoSK6uvqsHrVDjx5noLaxm4EnbmPmdNskJycJs/iAtLT0/HmzVu9DOaCX754KbqFQBEZmp6mV1qcPOfAawJFyXCKMfZhf46vZPi7d2+RlZUl97IPAU4ZrvpnZKTLNefP8bn4Fy+6kZKaLBRORZuZmSHjav3zkZenrV8Tofm4fP4O1q/fj8aOl4hLzIOV1R5scvTCq1c90p8wItLU/ElovKZY5/qVDiUny/pfvtR0AE0wZYVwYmSbnhc9smCyVpNOhBDwlPkaC+WK0mJra2/Bovn2uHwjFJW1bbC22gNL8x0yqfj4OD0LE3AEMBuBo0QQEUygsLW2tog4InA5YYoNJUIyMtNRX69ZTvUNNSgr18RRaxu5MkeuP+AtsrLT8f7DO7x58/ojEUIAKRHG8QlwNgJYP34Lx9fWr80/Xi8Cb16/jyWLNqG5+zWC70Rg3twNuHHlCbq7O0SxstG64ZgUQSSK2NgYvbWjIUiJwFxBClsf2tGKHXgzFVBdXa0oR2KI33ESnCixKUqpplp+Z5+u7g4cPXQeTht8UVBci9kzHbBn11F0dWmihCzI+9hPe26tPIuAr6mp1Sm5ajQ2NshYJAbqlYaGejS3NOHlq260tDYhLiEBN27cw5nTN7HV2Rd21u7YssEPdlZ7YGWxA3vdj8PfNwjHjp1BbFwiaqjcO1pkfiSi3jFrxGzW5lUn6+Ca+ZkvXnO9HJ/z53M4L8512aKNCI/KRFp2BaZOtER8bLKIQzV/9uf8uU4+g8/SRFgNqqqrUFtbJ8/SYKmZpX3oDBATVBKKQiin0vVKskIogY3yMTo6Cj06JUMOIYW9e/ceK5Y4Y8VSN6xa5irYzsnJlj6cUK+S6xAKJ/vRrKTCVRROcaI4pKq6AoVF+cgvKIGXZyDmGjtg1jgnzBm5DYtGe8JiWgDsZp+D3azzsJ97GevmXIDFlACYjveBydjdmDJkPWZNXg9P95O4c+cxUnQKl0o2OiZaxieFk0KJBDaKBjU+15+apilJrl/NPy4mDbOM1mHR/C3Y7noQeXnZegqnSawonEYGxR4bAR0XpylpEnpUtBr/vYzXhx5hTW2tWB6Ux5S3HR2dwkZkEw5Am5Vsy8m8fv1GlArZiN+TWmvrqnD+7B306fMl3HcekcFoa1PM1NXVy328n/0p2pSM5/ekSooziox3718jNy8Hp4OuwXTBZswYsRlmUw7BdelteJg/hadFKPatDoOXZTh81kbBxyYGvjax8LGJxn6rSPlt7+oQ7F3zFNuX34XV9ADMHe2GudMc4L7rkJjTbGXlZQJkw/G5fpqmXDMBRV9DdFx9vcw5Ny9XLJ7F85zwoz4DcPXSfbx7/06oXPWnjuGapX9XlzyLnMD+2TnZAnCOaajj+jAswEEIMFIzlRllMamVQCMC+DB24GR5HzFNy6Sru0tj6doqnDpxHT/556HYsztABtT65wiFqf6cLK/LSjU/g8/iM0vLSvD27RvcDn6GmRPXYc5wN2wyuQJ3i0dwWngZFtOOwXzKEVhND8SG+Zexbekd7F71GHssnslrt9kTuJnehdP8y7CeEYhVUw5i5eRDWDf3DHaZPcC25cEwGe2O2ZOdhFBqaymCqkUZUqxSaXMuNBJIvXwnt/NawYZWWkNDHVYsdsZ//dtwnAm8gZevXgqFE5liZTU2CszYl0gg0RJ+Wv9ClFeUa+sv06wkIkMvgqiIFAsSCcpOJYAUC9J0pNhQdjLFiVJSPl6B+PE/DkbQyZuor6/WsyCBruz89rZ2hIeHC3exP5/79u1LlBRXYfoUS4z8eg08Vj+Sl4XRMcwZ6Y5Zw3Zi/qg9WDbRD2uMjsNu9lk4LbgC58XB2Lr0jrxcFgdj44IrsJtzFmuMA7Bskh8WjN6LmcN2YtawXdJ364qbcF/zEPNG78KsKbZ4HhqH/PwcUZyy/qREfRiE4jNNJ4K4fhoTbC9e9GDWtLX45X+OweUL95GTmynIkv6JiYJINiJEGSlcv/ITiPCoqCj9+omEPgwQKbuewOSX5ASKI4oaYpCsRazxYfye71VVlfI9lVRDYx1OBlzDj/9pMDx2HcPLVy/keeQUUpC65mLYn3Y2HbBXr17g/JnbMDXZgiOHr8Fk3A5YzwzErOG7MHvEbswfvVf/Wjx+P1ZMPog1RgFYOzMI6+ach8O8i/Kyn3seNrOCBEG8Z/G4/dJn3qg98pozwh0zh+3A8skHYGV8HItnOYvvsmPrIXR0tqG6hvNqFCVKJPCacyZB8jteU5fRKFi+aAt+/u8jEBR4Q3QKFSxhQ1iQ27X+mp/CNRN+hC/vIbfwWcrTJmeIGUrZzEbWIVUSQ5T1hmYY5ZmhkmJ8hI0WRWlZEc6dvoX//KfB2OK0T76nkiVVcDJsomRLtf6FRYUoKS3Cnp0BWLvaHU3NHSivacLY79bA+Hs3zB/t+RHw54/eg/lj9mLR+P1YNtEfZlOOYLXRcVganxSxtMb4BMymHsGySf5yj+qjIcBD3rXvPDH4Dxbw2nseb95/wHbXACxbtBnJKZqSJnC4Zq6f849P6DWDqQ/q6mqwZtV2/OxfhwnBsdHRowJno58SFx8n+oNimIZNrxmcguISTQeRUzKzsuT6o2CcFn1kMKlA5DeviS0VjNOCSYno6qIj8UIGVgg6cuAC/uMfB8l7Y2ONPphWXV0j4owLam5pFufn3bs32Oy4H3Zr96CotBY5RZVIK6zAolnOmNDXCSZjvX6AAO21YLQnTMbuw+Lx3lg26QAWjNiHGYN2Yck4X5hO9JXfDO9XHDBvlPbZZIwXhvzJHIFnHqK4tgndr9/hoN9lmMx2QGoKDYYavP/wQSLChsFIzp9W04uXPVgwex3++z9G4tL5+0JE9Q0N+PBBC8aRS7hmCWZmZemCibW6YNxbvSNIPWAQjGv1IsaViOBNBDpFRLsuGEWx1KILHlEsaSKrWa7Jot3dXThx9Cp+8qOh8PM+jbb2Znken0WbmJwj1/V1aGyqw3aXA9jk6Iv65m4EnrmHyMRc5FXUYd/eC/j+92ZY+AMELBizD/NG7MW079wwb6QnFo3bj8nfbcaSuW5w2XQMRqM2YObQnTAZsw/zR3lixvc7BVkfPWPsPhj/dRsmD7JDfHoxknNKEZ9RhNbuVwg8fgvzptvLXDs7tVCCmj/Xx2sSH8XVskWb8PN/G4Wzp24JwpSIZijDMJinAnPsrwKbWjBQE1HkNn0wTtmv1ORUEioYR69OiRDDYJz4CTpPjoMVFOTgwpk7+Nm/DofrZh/5/u3bd/IsxYIUQc3NDbhy4QHMlruhres14tMLkZBZjISMIqTlliEqPhdj+1nC6DtXPRcsGOOJWcN3Y9oIR1iYemDqUCcYfbcNSxZsRUFFLbpef8CdhzGYNNgec4d7YMpQRyye7YapgzdhwaheJPB5f/3NCjhvOIbi2mYkZZUgLr0QMan5aO7sga/3RaxZuU08buWJ03hQ609KSkRtXTXWr/XET/95BI4cvKhbfwVidXa+gt/fM1KSDYKJhJ+KBPTp6upi9kgwY2hm8mayCjmAJiSvORmyJsPE/MwHkQI4IJMVVMKH/S9IgKqwkKHqLnkmWZImW0ZaLmYZ2SAsPA3nLj4SCoxNK0B8ehESM4pRWNWI7S6B6P/LRXrKXzB6H8YMsMO5y4/R0vMaBw9dQ//fmmK3RxCqmzpQXtWCzOIqLJrtgmF/scbBI9fR0NmDra4BmNB3I0xG78PCcfsxpZ8Thn2+AuGxOcgurkFSdokgnuNHp+ShtqUDlma7cdD3HNramlBcVCSw4PoYtKMF093diZVLnPHf/zZKzO7KyjK9mUnipEgxNDP5PTkgL0+DH71mmqbkDD6PffswdUdsE7BkGyKDD1OOEtmKZunrV6+EbcTRaGsXKmEAr7GhQeS7v/dp/Mf/G4TDBy5K6IBAZ39SCO/jta2lOy5dfIo9e07j3tM4ocC4tAIkpBcKAtLzK5CSXY4Zox0w7A9rhPpNxuzHmAFrcf9pLHo+AEHnH+C7P62Cqcl2FFU1oPv1B9x7EiscMOab9Th49BrqOjrh4nwUE/o5YdE4b8z4fiv6/twE/r43UFLbgpTsMuEAIiAurVAQwHnkF1dj/gxHPH0SjspKDRZcP7mdRPbq9QusMnUREXTy2DWUUgfoHLW83Dw9/Ah8tX4SMNdPqaKHX7sGPxJnH+YtDfMBDDVwMCLlhyJEsZBhPJ9mWHFJAc6euo3//Kch2OIkboUoJlpU9Q2aFXH4wBlssPfByZO3ce12uIie6GRt4fFphUjKLEZydinyK+rxODQNw/6yAsP+aI75ozwwe/guLJy5FTu2B2Lm+E2YPcwdkwduxrIFO7DN+QSMRm/AjCE7MH/UXkwZugEmM1xgNGQLFo/zxvRBzvjqp3Nga+6DkpoWZBRUITWnvBcBOjEUkZCNzKJKXL4SAmvzXTJn5jP0oZKkZDQ11WOzozf+80fD4Od9Rr7X/AQtVEM/iTkIlQ8wFEFaMFOJoNxeEVRfX89ksi5opAXhVIBMBah4zQdxMhKg0gXjKP+1eHo7Thy7Jo4Yg2KdXVomi/ezHye5Yslm+PpcxOGA60jKKZEFK8qjLkjMKkFKThlSc8pQWtuCGzejMPiTZfj+NysxZ9gOzB62GxP6bcKc4R5i7dCiMR60HeP7btZ/R45ZMGofZg7ZjfkjPTC5nyO+/tl8WCz1QEFZA3JL65CeV6EhILtERGBceoEgICopF2FxmSiorION5V7cuvEUzc2NMnct2NYogT0G/n76z8Nw9NBF8QsoRrhOdZ+CnxZ8ZLCvVn77ONip9aEo0nvCNDUNw6WGniztWTYqmdi4OGGf129eSzCLUUM2KqUf/78hOHrwioSLlSfY0tKEK5duwcpiN3z8L+JhWCIik3IQmZgjCIhN1RBAiiQC0nLLkZ5bgbL6NjwNTYPxSAd8/bN5GPO5FWYN2SZ2PXUDlSpflO8m43hNfbEXc4bvxNQBTvjrr03xzS/nY9vmk5JCLK5oQkZ+FdLzK5GaqyFA44ACxKTkCQJIFHEZ+bh1NxIrlm6R+ROo5H42iiDTRRvxX/86AoHHr6O4ON/ASOnN+ImflKn5SexPTmKjhKFXrJQ8k/l9WO5BDFGOMwZkmLTmNbFFdqJrTlFF7c6BlJ2smVzFYhf/7F+Gw3WLLz58eIdXr1+hoDBPBqaF5LYtAE+jUhGemI3whGxEJuSICBIEUAQRAdllSMspR0ZeJTILqlBe14bcwlps23wCQ/+yAn3/ywRDfrsK476wxZT+GzBt4CYYfbdZ3if1dcDoT60w8H+Wou8vTDB7wkZcvxmF+tYelFQ0IbeoVp5JPfMRAtIMEJCYjbC4DCTnlGDpws24fesROjraZM0SWGusxUYHb/znPw3HQd8L8j1hl5ycJPKeFK2KDnqT/r1FB8rPYmqSypj9JRRBLKqUmZY0bpaH0TJiDFtdE9i85qAURXJNlqqvwZlTwWKGMhTB1EhLawsO+AfCdbM/poyxwPHA24jPLERYfKaGAD0H5CM+rehjBORXIquwWqyVksomNLT2ICGpCJ47L2LuhM0Y8ukK9P/VIvT9bxP0/e8F6PffJhj42yUY298SlqaeuHQxDBW17WhofYGi0gYUFNcjr6gOmQXVGgI+0gFEgCaCyAHP4zORlFMMx/U+uHjulqylrb1VS6y3N8PJfj9++qOROOx/UfwfwoSOq0q69/pJWsqWIkhL+vf6WYQdiZmc8FEw7u8pWXbSB9N0SXsqaXz4gNiYWL2S8fcOEiV8NvA2KqvKYTxpDRZO8MDONQ8wdZgTjp++hZjUPDyPy0R4fJZQf1J2MVLzy0XsUPykZJcKAigmiICc4loUlNahuKwBNfXtaO16jaq6diQmF+H27VicOx2C0ycf48rFMDx7loG8wlq0dL5Ga+drVNS0oaS8CcWlDSgsqUNescYBaXkVguCc0lrkltcho6gKaQUVSC+oQGJWMcITshCfUYjduwOxaPYGHD1yVmL4bB8+vMXqFVvx838djWOHrkgwTp8PSO7NBzCKYAg/FcxUwTjlJ0gwTh+O7uoWeUaTUyWNiQgCmJiiWUqs0Z5lapKd+b0omeoKnDpxEz/90RD4ep2BnfVuWM0KxC3vOgR7V2PJVE8cCbyB6NQ8objU3DI8DEtG4Ln7OBZ4GxeuPsPz2CzklTWgqKpJKF/jgFoUlGgIoBgpq2oWBDS2vkB7zzt0vQJ6XgPdr4H27ndoaOmR38urW1Fa2SIIKCqrFwQUlTegtLZVnh8amYkDh69h40Z/2Nnuwzo7L2xw9IOX7wU8i05HZkkl3PcGYsZQVyydtB/Txq3FrVuPAbyDy8YD+Mk/jcB+z1MSjqGMJ2wICxZ8EWYKfhQzCn6ULDRTDf0sCUdnSkKm5m/sVAKa13yAllDJkQfwPg7Iz69evRStX11TgfOn7+J/fjwKUydaYNkMD1z3KoWHxTN4mD/DjGHOOHXhLtKLKhEQdBtmq7ZjxRIXrLf1xBYnH9iv3SufLVbtwn6fSwiPzRaTsaqpE5V1bSivbhbgE7CkbAK5ur4TNQ1dqGvqQV1TtxQE1DTwuw7tt8ZO1Lf0CLKYq46MyYaf/1VYrfaAxYrt2O12DFcvP0Z4eDIiI1MQfDMMO92OYuXSbQgIuoO9vkEwm34Aj4+04eDmWIzob46YmHh47QnCj/9xKLz2BoreJAES0L0JrVQROQQ05X5DQ6PAj/cQ8KoChQhhf70ISkruTcnxZpUPoIxTdq5iISatmUChzausoA32nvjz/07BV59Oh9vqYHjbhGGraTB2rXgM42GbcPjENVhZekgs5eaNEBSVVaOirhllNU2orG9BeU0jIiJTsG9PIFavcoOlxW647w7ClauhiEvIFzleVd+JxrZXaOt6g46ed+h88QFdL0n979HU9koQUlzaiJS0Ejx6koSTgXexzfUoVpvthLXFTuHOmJgM1Ld0oL69C3nldUjJLUVCVjFS8kpRVtuE7NxSuGw+hP5fzcOSiXtx2DEBF3YXYLf1fSxfvBkOdnvxs38dgUP+F1BYxCI1LR9AS4cyn82waIHBSJUPIGEzn6L8BBFBLC6lYqACoSgilTMeziQzRQ7tWV4TEcQ271O+AeUffYCAIxcxbqANHNcdwvLFbnBZEozNJtfhuuQWdq14hCnfb8B3/ecg8PgNVNQ1ITG7CE8iU/AsJl0UM98vXA/FxeuhiIjNlghpckoeThy9js2OPrC22IXVq3bCysIdttaecLDbD0d7b1GUfF9vux+2Vnthvdodlma75P4N9vvhviMA587cRUZmMbpevkVpVRPuPU3AuashuPs0Acm5ZUjMKcHz+CyZw6OIZDyLTUdhVT22bPTDpIGO8FwTih0rHiDILQMzx23CkO8WSULm2OHLEh0lHAgn+kOEB+Gn8ikM6knxgs4XItwoUQhTfiZB92FlMrXy+w/vhXVUPJzutWFZiFZWoikZlRNmCwuLxoivzbB95T2st/HDSpOdcDK5Aoe5F7HB5Boc51/GiH5muH7jCXLLqnE/LB4PnyciOiUXj8NTsGGDPxbN3whL8+3Y7OAt4QqTuY7YsMFPFOcHAI3NnSiraEB2dikSEnIQFZWGiIhURESkISY2Eykp+cjNYTFWHWrqWtHS/gKdL96g+9V7vPwA3L4fi+XLtsJ0sTPWrd0jFRWrlm3F0sUu8DlwBUk5pYhIykZIdBqeRCbjYXgScksrsGj2RlhMOSr55X2WoTCdtg+Hjl2FycwtWDDDUeL+TG3SgKHnTz+JPhNlPnPchJ9S0hTnSknToqSuZdOLIMOEC2U8lQkbMUzbVhU2MT1H/fD2zRsUFRfAbs0uWM4IwMIxXliywBkLp2+F/ZzzsGPGasEVjO5ri4MHziOntAp3n8XhUXiSWEOnLz7C5PFrxJxraWqTsVR70fMCl8/fx5zp9jgZdB+1zT0oqWxGdV0H6pt60NT2Eq2db9DW/U7ET2vnWzS2vhQ9UFnXgbKqVjFfKbJcnY9JHvfZU00MGLbkxCzYrN6NVSt3IDwxR0xQcubj50mISslB4OmbGP+tHVxNb8NlyS0sGrsP3n5n4edxHUvHe2PVss3o7ukUi0YV97Jp8NMArBw5ihyKIOVnUS/w/j7Nzc1ePd3dImJUyIEcodKPxCA/a9daoawSRWSxZXO2wsL4qMRiFs/bglkTnGA9I0gQwETJjMlrkZxVgIfhiXgUkSTsfv5qCKaMW4O0FI2L3nR1IPRRFM6de4B7waFoq9O4rr6uCdMnWeHS5WeiTEsrm1FR3YrKWk0J1zZ0o67pBeoae1BT3ymlkVTSZVUtgiRHB1842vWWyxdk5IrivXzpEVJjNTORzc/rNObP3SBWWkhUKh6HJ8lcw2JTMWOCPSyNA7HB5IoAffu2Y3DddBjWM07DaPhGkQBd3Z0ieggzwoXvfFE/KLFE+FGakAuUiOLvfRKSPg7GKSVBitfyAVowjlygL67NyUZ2diZa29phPNoei8f5YO4IDyyZ74rp4xxgPjUAtnPOYfwAR+z3Po2YtDxBwOPIZITGZcJo6lpER2ru/YO74Ri0xBv9dz3HgEOZGLAvBqMtj+PcqWD5vaK8BnOM1yO3sE6smYqaVlTpEFDXqCFALCADBJAbbt6MxLKFm+QZL9o7sNHlGAY5XkF/32T080vDAKdbWON4GA2VWrUdxd+27ceFO8mlfEWn5cJipRvmDfeAw/zLWDx2P+xtvbB7R6Akd1ZMO4BtW/xRUKgREiUE4afqpmj/9+YTkv6uku6j2w0j8QtSOp0EiiKyEa+pMCiOlJnKjpR1lHGhoREY1n8VFo7dj7kj92Lp/G2YPnY9lk86CLNpxzBhiC3uPIpESEyqRv2JmdjnewEb7LWIaWxEEhasOwz7yFYM84nHIMsgjPB4hpH3OvHJ6rO4duG+3OezLwh7954VJ0shoKa+S0NAc4/eBOX3lTXtqG/ugdny7cjKyBdPds16f3zllYSJN+oxzPEqBtmcw5iLFfjmeAmmm/vgVVcXWprbMGeGPR6FJ+NxVLLMNyo1F5s2+2Pqd1uk+MtklCdWm+3G/n0XMO07F8nc2VjsxZu3NDOzJYvIvIeCH2FGaUGCJSyZRaQuoK5VPpQkZKhwWfVMtuEPZBmVUKBYInJ4TQSouhfKsJCQcAzra4bF48kBe7B43lYYjVuHpRN8JREyfeI6PItOEcrngiJTcrBipRtCRR5/wNw1+1Hcpu3OGf7dEnzyX2Pwyc9G4U8TXfHX252YYnsSLzs6kZtThBWmLqhveaH5AbWarU8EENj0AzQEdAgyYuPzYbbcVZ4b/jQafZ3v4/tTRfj9n+bii1+Mx5e/HI/PP1+IWRfzMWp/OAJPatzmuM4T/kevIiwhA48ikxCRnA0PzyBM6O8A65lBmDvcHavNduHQoRsw+utWzBuxB/ONHNHR2aGlJGuqBX4MslHREmYk4Lz8PPme8FMJGVZUSEKG+QC1xYaUTbFD4JKFaPMrK8iwdI8KhoG2vNxiDP3aHEsm+GoImOOKqWNtJQkya+guzDbegGdx6XgcqcnU0LgMLJzvhMryWjRWVmOw4yX4572A/74gfPPnmfjs9zMwbOBiHN5/Bocz2zBk52PEhCdJQtxs2VYUljYKlRPQFDl/I4JqO9DY+grXbzyHs5O3zHX7vosYfLYUa3zuYPaYVfjst9Px1Z/n4Ms/GKOquBTNb97BxvWE3MuIrvO2IwJ4zpeWkd+hSxj9jTXMjQIwY+gOWK/Zg+PH72DKQGcJfU8abon4BK2496OUJFO6BvAzLL38KB/Q2sqkfI9oZNqmxJIKFvGaGFRJZSoPYlarduvEw4ehGPb1aiyZ4Id5I/dg0WwiwA4mY/djxpCdmG3khGdxaYIAvp5Fp2Hxgs1orG9BYUYuxro/xsTrtfj1Lybjqz/OxB9/NRUhT3qtlR0XI3Dl+jO5Xr1yO/IK60TOKwQQ8OIJCwK65Hs6ZJcuh2DH1sNaP9dT8Imp0j+Txby//+Vk/OU303D23AM8aXqPpVtOyW/Mcm12OYSIlGw8jU7BvZA4OG3yxeh+Vlgx9QiMvt+GdWu9cCLgDiYN2AKTMd6YNtoOVdXVshVLaoeam0WSEIYKfgzdKF9LgnVUwC3NWlJelaXI/oCyMql3lP0BOVq4mQ8gtrRwtVZWoTbbMc4/5AszHQI8sWj2VkwdYyt53DkjPIQbHkck6WUqKcp0iQvSU3LR2dSMyRvPYnEy8P2ULfjsp2Pwx5+PR1KUppzZ/E/dQWxMBlqb27B6xQ4BcqVOBCkE0CzVrCANAfz8LDQNDjZ75Rk7Pc8iPE1TfmxXz97BL/9lCD792RiMdr2FcffbsH57kPzG8MTe/edEVz2JSJbQtNuOAIz4yhLLJx8Sql9v54Njx4IxaYAzFo71xaThNlJ6kpObLWJbCzcXidjmNWtn6UOp/RWyx023x07KUj4qzDJIGEhhUnz8301JkgsqKsqQGJeOQZ8tx9KJfpg3yhOLZ2/DtDG2khqkZTR8gDkuBYcgNDZNrIrIpGxsdjksZh+b844TmHqjAsvD2zBs7k58/tVSrHQOQv4r4FlZO6wcD4iNffn8PWzZeBBtXW/FDCUXkPKbWl+iue0VGlteoq6hWxBAE7W8qhWmC7egtbUDGWn52OJ+Gj1MDr15j+V2B/BF3+WYuPkSzPOB4dvu4tnDSJnPonkbcONuJEKiUvA8LgOnLtyHybxNmDjQEcsmH8TEAZuwYb0vjh0NxuRvnbF4vB9Gf7caISHPpT+tSFo+PT3dYkXSKVPRYoZ2lBVJAtc7YtwfQPGiUmZEBq/56k2f1Yl4IqbVfa1tLXjy6Dm+/2KFIGD+SE8sme0Go/F2Io444SFfrobbrgBEp+SIc/M0KgV3nsRhlpEdmpta0VLXgLnWvph6vQwWeYB56gcsSgHGBVdhyCo/ZKbmCgLmzViHsPAMqSOqa2IArlvi/UUVTSgsa0RZdRuqGzrR0PICtQ1daO16Cw+PM9jtplVqHz8RjIkuF7A8ugvLE9/DKhMwywaGejzBth3H5Z5L5+5h2bKt4gs8fp6I8PhMuO8LwqRxVjD+3hWmkw5gQn8nbHLyx7GjtzB5oIaAkd+aISkpVZJZKv3K8ANrhlQqUoUfVNqXKV1eswpd7wkbKgnDHRzcWGAYjFM7WNiuXbuNoX01JTxv5F6YztuOGZPXSS3m0om+mPa9C4ynrhNWJkuLaZeSA3fP01LewdbV3AKXHccxfctZzNj3ALN2XIOtcwBK87S50D533xWE9h7GcloQnZiPE6fvY9OWw7BY4wFrGy/Y2Xtjn89FBN+PQWZ+NarqO8QqWmKyBQ/vRchzbl5/igUORzHb4y5m7b2PRZsDcfrkTfmNDuG0ida4FxKPkJg0hMWlI/hBBOwd9mPkX82k4GvpRH+M7+8E5y2HcOToTT0CRg00R2VlDfLycvSesAa/3qIFfUq3U9sfofwswrsPd2wTG7R8qDTIKtTmZCNeE4NSH6+r7+d93ENVXl6KouIyjBhgIVYPK9dWmLhjjrGjVDQvnuAtiPnrX1bA++AlRCXnCAJo3sWk5WPj5oMwXbgZRQXa9qh3L7pRX16Nt12afmmob4KlmZvE6bkvKzopH65uxzFtii3Gj7GA6dJtWLt2P4yn2sPc3B3GRuswbsxqLDRxhv/hm8guqpE05MJ5m3DI74I8k629vhFNVbX6z9cuPcLEMeY4fy1EnDAq3ztPY7DTIxBLFrtgbP91MJ3kjyUT/TCunyNcXI/g4OHrgoAlRMC3a/DsWYTAhSKIuXRKFFpC1KsKtqpuiLJf6QjZH8Bt82QRWjX02tQuRj5EFV5RacgDKisFOVqUlNtsajFy4ArMGe6O+aP2wXyxJ+bPdsL0wW5YPMFHEGA8eBvGDl+N20/jEBavmaRPolIQm1GAA0dviIzduN4b58/cwYO7Ybh49i62bj6AZYtd4H/oGhIyy+DueQZTJ66F0eR1cNt5Co9CU5BdVCt55k0Oh3Do8HWk5VXi0vVQ2Njux/gxVpgx3QFHTtxGZEI+Nm06hNUrd8J//1ncCQ7B3eBnOOB9FubL3LBqxQ7ceRyL2PR8hMakIzw+A5eCw7B8hRsGfWUq61g0wVfE7Lh+DnDdegR+flcw5dstOgRYITycGUPmhyvFYCGQCScqX1pFpHi114C/0ewn9fO6Nx9gYOd/JIIMUmrKT2BHtsSkJIwdshzj+zpg0TgfWCzeh4XzN2LqIBcs4cTH7xfqoS6YOWM9bj2Jk3Avo45cLFN/USl5Umy1a3egZKhcth7F8cB7eBKWigNHbmDKJGtMGL8GO91PIyqhAGXVrRLnKSyug72ND65eSYSL02H0vAIaml6guLJFELRh02GMGmEuiDgccBvXb0fB58BVbHI+hE2bD2KP5xncehgjqcjY1DyExpI4UnHu6lNYWLhjxBBTjOhnjqVSce0jImhsv/VwdT2M/fsvYspAFywZ74sh36xCcbFmxNAHoOPKRvhR2bLRaFGbAqW8patLLCQRQYoDKN81DsgTc4nsRA4g1viblpTXSu/ISiXF2g6XNSu3o///LsOM77djpYmHhHwnf7dFqJ8IIOVMGuSE4YOXY4aRPfb7X8HdkAQNCbEZEhuKSs1HXEYRnoSnCtVaWu3F1ElrMXP6emzfFYhnkVmSYqTpWV3XjtrGLqRnlsLK3BOHDj6Fk70/Wjpe6wJ0XWKeFlc248nzNLhsPS5iavIEa1it3YejJ+/i/rNEhMZl4VF4CoIfxuDijVB4+1/BKrNdGDPCHBYWHpIVG/NXKywax3VoCBjTdx22bjuCPe5nMX3QdswZthuDPl+CyMh4kfWq3JAEzCItTYrky2ZtXjMLJkZNQz2qaqo0DtDrgDcsq6AOSNPvw+VGPSJH9jjpShd5TXGkzlI4deIGPvv5XIz4xArzpm2RdOPEb50E8AoBY/utw8YtB7HLIwiTx1tj3hwnWFrtwXpHP9ja+WDZsu2YNWM9jKbYCtDNzHfB5+A1hEZnIr+kXkxPRjjLGA1lLKi+XQC91sIT82duhZ/3ZcmMUfmKGapLXfK6sKwBz2OyBPErV+7EDKP1MJpkh5nG6zF3jpOMN2WiNYymroO5hTsOHQ9GWHwWbtyPwKiBa/QIICcP/9Ia27Yfw/Ztp2A0cCvGf+2AwV8uRVhYFMorysR0J7wohujckljJCSRoiiHG0aRgq75OtiuJDvj7hVk5Bik1FhZpWtzQCiLw09NTkZmZgyH9TDHot2aYN9UFllbuGPaXNb0ImKTJTnv7/ZL6uxwcBrddJ2Gxxh2mptuwYsUOWK/dC1e3AAScvifJ+oTMEmQV1aCwvFES8kWSlG/UEKALRze3v0ZERBbcXI6jrKIFDS0vBQEMVZRXt0nomkn50oomlFY2Ib+0AUlZZXgYmoyTZx/Ac/957HQPhMe+MzgaeFs4ISwuC1FJOSKSLgc/EwQsnuAnCpgi9a9/XAHnrUex1soLg/+4BiM/s8GwfqZoaGRa8oP4UEoEUQGrpBVFjYIndQL1LYEv+QDlB6iUpKEfwGu151WVKvJ7YpGb23jd3dMBf58g/O4/psFozAah9H6/WYjp32/D4oneWDbZH5MHbsEaCw8kZhVJ2YdWAJWFkOgMeY9OzkdCRjFS88qQWVSlVUUUVSOnSKuK0COgigjQKJtOV0vHG/S8gQC/mp6w4oAqckyz9GFfPoPnO7DMhc9laUpyTpmUobA2SEoUWaCVnCvzY7bu0s2nogNY4j53hDum9N+E/r9diJ17TsF85W4M/oM5BvzvcpiZbpN9A8pnInVTbxJmhvBsaNT5CNyy1NIsL+5F+MgTVp4cT09h0p37W5UnZ7jTnVs203UYleLc4gJcOn8Pk8etwYihyzDgjwsx5gtbTOy3QRAx+qv1WLpoK+LS8/E8PkMQEKMrzCUAkjJLkCyliaVIyS3T6oIKqpBbXKuVlOgQwEoH6gFJyNTR1tfkPXUCryVKWtcpjhrFkFaWoqsLKqpFRkEl0vLKpf6UGzSIACnO1RVmRSawai8L0Wl5OHf1MQZ+ughjPrfD6M9t5cV1HTh+A5bme9H3F0vxp18YIT42HfUNtXoK5+Ek8fEJIo4odghPwlXt/Kcera3jTn9N2nwUC1Klh9pZC71nRRD4fCA5hbsJKYpYesjvG+o1BFVUlCIvPx/Xrz5Gvz8vwJgvtImP+twW3//BAjOnOiIyOUc8zN7SxFytMo7FubrqaNYMpefpEFBUi8JSIqBeZH9MWhGyShvQ8RJoan8lQTgVkKNI6nwFlDZ0IjK5UJy20opmPQfwWQoBrIRQlXEsjWSFnipNDE/IFEIJOHUL/X5vgrFf2GPMl/YY/cU6fPfFYmzdeQJjhppj+IDlOHv6hjhghMfrN29kQzjhRRgylkZ/ideEqyrl5B4BiipKEUGAnBf0d7coaaWHFD2MAfGa3GC4C5Df87vWNm2LTn1DHd5/eIM509ZjwK+XY+xX2uRHfGaDsUPWSL6Vi4yIz9IQkKRDQMbfIoBVbFIZV1KH6oYOPHyShHn/OxVW/ZbAc3cgYjJKUNP+Ck0971Db8QrxmeU44HMZtmMtsegnY+DnHoS61pdaaWJJHXKKapBRqENATi8CflgbSgQkZBZhn/dZ9P21CcZ96YjxXzlg+CeWGPD5XKxa5iKV0aWlZXKMAuFB8d3SQhiUCGAJGwJdRUQJV7VlieKIfZR/9YPzgrR9wDSPtHxApD4Yx7oXFUwyLL3jgKySZiMiM7PScfdOCL7+/Tx8//vVGP+1I8Z+sR5//XwJrt55LtTG0kSpDWVxbooBB2QZIKBQE0EEXn3rCwQcuwWHPkNw+F/GweEfhsH8t8awm7gWW0xcsG6CDVb/bgYc/t8weP9oFPb3GYr18zajRhDQiIJickBNLwcQAZn/BwJYG5pdAls7L/zpxzMw8Ldm+OZ/FuPz38zChfO3ZZ08y0g75kCL9efn5evrqEic9JVI+YQjzw+i5KCSpgGjYE3CpaT5aIsSKZ9YISYJ7O6uLn1G7OMtSszodPzNTnA+nNc8QeRM0HWMHmSGL/5nAfr9aik+/cUMHAq4hYSMQpGzgoAkVZyr7ZBJzi7REJCvIYBym/Kb9v29p0mw+OkEHP3PyTj8s6nw//cJ8OgzEm59hsm737+Mx5GfTMLRn03Gpn8Yht0uAWhof63nAEMd8BEHpH5cnBuRmCX7BGYa2WPx3E1YuWQr1phtQ2ICZfwHca5IzWpPBWFAAPMzrR7CizAgDAkP5TepDJmCIQ0a8QO4Rentu3f6nLBCBhUtbyD7yBYbXfk1lQ2VipTepfMIGe0sCQ6uSu+0qGqphC5OnbwmpYeTRq6GnZ0vYlILEJmoq45WOuD/QACplggoLm8U2W45ZR32/sNIeP/7RBz75VTc/HIuHn67AHf7z8e5P87A/h9PwJF/G4/lv5qK0LhcCchRf4gOKP4/EKDjADWf+IwC+B+5iXmz1gvAuSmjuqYSb9+9lviY8oe4TgKbkoDESLufsGHwUouXFejrP3k/4clrJmPYh8gj3D/aoqRCDeo0k6hoblHqPTLs74UqOAkVLWW5Nt3xmmpNwTA2kp2ts5aqG8ThuXInSiKi3KRBqlMI4JYlQUCeDgHUAUU1ArzC0no5KuxxeAbm/nwSHn8xC61LbdC+ch06zO3RbmaP9hV2KJy4Ehb/+D083c+gueuNgRlab6CEyz5GAMvTZZ9CPu4+icWRE8FYuWonPHYdQV6+VmjLxuimOuaM5iPFM6mYjcBWRxuQwqlkKcIJR4odEiw/81r5WkqfyhYl2c1YrcWotXi1diYQKfmHW2w030DbYqPKFXt9A81XoO9AT6+2rhY8iaWO23HaWrBs8SY4OR8VZ4vOzkdW0A85QKwgIqBegEguaO58g4s3IrBj7AqkGS9D+1IrdC2zRqepNcrmmCNg5CK4ux1HdctLyZyxolpZQXmFH+sAbonSdknmIy41XxB/4VoItu8OwryZDigsKEJNje58pGrtjCH6Pr3xfg1Omp2v2f9MS6ozgXhdI+Wb2jYv8YB18ORpLXxnYv4jT7h3E1nvFhsOqDxhYpInSkk8+/07bROeQekizS42xjrIPZryAUpKS0U5P7wXCqPJdjh14QluPYrVFp/OHTKaP6BHQF6FDgFUwhoCCEx6ty2dbxGRUoI9LkewZ/Y6HJxmiX1Glti5ejeu3Y1Fc/c78RHEEzZAQC4RkP+xCCLiOSY50Nf/Es5fD4OxkT2uXLgnp0FSClA6cP2kZmXXKzjV1tSIxCDnSx1QfJxwPRslBE15Nj5D+QkU44ZBuz6dne0aAgyO1KI9r0oTKyoq9WchKEeN8p+OGvPDfBDjQ7QCVOERkUJ242DyvLw8+Y3nATnY7oHJAhecvRqKy8HPEZGYI06Y7JLMLEEqd8nkVshuRmWGUpHSqWIogvuCufOluecDcirbEJ9dg7TCJtS1v5NSRS0Uwf0BzSgub0Kh7JDRIYAcIJtBNGsrp6QGT8JS4LjBD57eF7Fi5U7YWu5CaWmxPu9NcauVFr4TGW54+iG5gs6qWicVsvKZqCMolihmVFkiZT85hvAikfJaRBC/YOfKigrBMjHMjoz5UMzwN35PeUjO4LuygshaKnLKODf78CA/2UPc2CifqTv4HAb7CvLzpWJtydJtOHPlGU5feorbj+LEHOXmPFIpX5kFlbI/jMDrRUCLxIIqatv0daLNbW/Q1PpKyhS1iol2uUdDQCOKSqiEa5FfXIPs4mpkFVYJEkIi0uHlfREODr44fOI2lphug8UKVxEv3JhN7uWa1Z5pEh3XnJuj7ZMg1VMkZWRo6+S9XCcJlyKdn4k82aT+8qUuWZMhcCBylAjvk5SS5EV5zSa7IGNj5QYihXuGlR9gGKwz3D9ABKhD7zgwWVKxF+/Xjj54qe0g1/UvLSmG+TJnTJu6Dn5HgnH2yjMEXXyEG3ejJAKamFmC7KJqFJTWawE1HeVLQl4HfFo4jIjWNb7QVUZ0y2ctZK1DRE2rcE1RaT0y8ypE2T4KTcalm2Fwdj2GnR6n4X3gKubP2YiN6z2FQlnnyUaCIrcTDmzkApVmJHfQX6J4ZqOIieFhh7pULcUP4cVGbqAPRaSyEbFiqChPWAXjqL2pmYldvqtrYopIoGxT9e0cmC9+r5S0Kj5lHy2AV6+dpaOeJ4pa7Teuld30gQFXMG3CWqxcuQsHAoJx5koIAi88QtClJ7h86zluP47Hs8gMxCUVIC2rDNkF1aITSiq4W0YF5Rgb6hCxQ5HDzXhp2ZWITynE8+hMPAhJxNXb4Th96QlOnnuE05ef4Pz1UKH6Zct3YOJYM1w6f1tOwCVFc+5cB40Qxm1UIS3XyN/5mWsg0LV7ew/r45r5u1LWzU1a7Q8pvXftLNbVjjggwX+0SU95uiyZMDyMw/DEJ3WyLhupXSlhYpkleGw0XZ8/f65XwlJJp9t9GRERLvY0W01tJeJiE2TnyvRptlhgshluu0/h+NkHOH0lBKcv8/VE/wq6/ARnr4QIAC/eCMOlm8/ldfFmmCRVzl17hjOXnyLo0mPt3qtPcZbfXQ3BiXMP4OV/BXbrfTFjuj3mzXLE6ZPBcvBsWnoKmpq1Ilqe/RkRqduIKMdlZujNbDZKBaWECVDDddLWV6f4srGfyoTxeXyuOjOJiKFo68Mj5SnTDeMUEutpaxOZxkE4AMUJkaFiGypWRFZSsSICmH1IEbymKOLZO7zWuIiVd83yzvv4PWVpW1sz8nLzcfXSA6yz9sD0qWsxZ5ajJEiYEfP0uQAv/8vCJUdP3UXA6fs4fuY+Tpx9gJPnHsg7cwlHTt6WPPNe7wvY4XEa9o7+koQxWbAZC+Y6YdG8jdjksA8hj6Nl7xaPHiCxcB6cMwmQn7nFlmvluvgb10+qVfEchpYJJ3K4WufH95YLh/Berpew4z2EHZWv1pfH/HRrO2Q4GBsVBaldHLF3mpmpHBHDwixDM9VQB9AbZB8l3zgJpVPIPZwYGydA6qBVwUbZq0zY9+/fICw0AneCn0itzmwjO4wdZYH58zZi9gwHyWAxtTljmj1mGPOzI2YaO8B4mj2mT7PD3NmOWLp0q6RAVy7djIAjF5CSlI2qSi1k3NyiOZY8ITE6qldfVeq24xIGbKTOj+eYqjcleY+hHCeAuU4qXjbep3Qkq8gJH66ZTczQaAMdoE5LMSw7UZqf15WVVaLZec3OahcgFU6m7ALUdlFSMZOqec3veB8pQoU2OElyltqFqb5Xpe98Hi0K3keT7+3bVygrK8bs6Tawc/CB35FrOHXuIc5eeoozl57i7OUQnLvyDOcuP8PZSyE4c/GJZLr8jl7HgWM3sXChM44cvKQDEK0wplO10hHOQbNGtLnyMymfcyIM1NGVao4MLdOQUOEGEhZ/o8RgX7VOEigJk4jgmggTdWpKe0eHPJfjajDSmaH8pwcigDfzC2pt2rmkcloApFoOymsCnoDl5DgIB+cC2Id6g4jjNSdGjiH7Sel7fr4ghwhQ/fk9J8d+CgE8EZ338Rk8eTc6KlZOqN264wRcth2Dj98l+B24gmMnbiHozAOcP/8E5y88xZmzj+TzyaB7ktrcvPUYTE3dsH9PkC51qlUo8LlUfJwDAcE5cP6cK0UkfycMSO30Ut+8fYsC3X4uAo2Sgc8iAtiXipSnoFBKCNwKi+R3IkojSuYEtO2oalzlBygj528O7aOjwc4EinZ2tKac2Flt0iNVKCXNSRARXCgHERGkiwUZsqaqJ2LjpJRoYmN/JdIoF/kMUi2bvc1eqYDzPnAZN+9F4WlEKsKZu43PQ2xiASKjs3HvQRzOXXyCQ0euY72DD8zMduPbr01w5MB5VFRoz+GaDA0NAo1GhDIl+c45KRFEAvyhmFT1nJopaWCGlpfrNzeyETa9IkhzWPPzNSdVjVut66udFaFTHqQCUjuxw4cR0MrRUnvDeK2qvPhwll5zQCKAwCNweR+fxwmS8jioUsL8ntTHZzMvyoJgIorUwGcT4XwGzV7ea2flgckTbLBi+XZYW+/BWuu9sFrjgbWWe2Br5SlbVDfY+2Kj0wHs3BUIZ9ej2L47UCoujCauloOUqOg5B85bzYFr5DjcpE7RSoXKOfF7fub6eD/7cV6Gc6S4IoIIG95LPSLraW6W30loPJqTn7ke7Wh8bfuq5uj1nrLyUWEWAcMJyj4nnVNBylRmmDoHjYOzEeA0y7gFR/3OxkXwjGa9eZZL/aBZCTRDyaLaszP0Ji4bKV89++XLHjwPC8eEUebYsOkAtrsHIvDsA5y9+BjXbkbg7oM4hISmIjwyE6Hh6XjwNBHXb0XC0/s8NrkcwibXIzCeshb1dQ1IzzDgsJ5uMaWV80mg8DBZzo2NIpB/AqHM0Kzs7I/MUK5fcRFF7fPn4XquoRnK0ARFOhsJkdtYee5QQyPPFKLlqBk13NbKJjlhXpAaCEh2IiJU9FNF8tR7b8RTi/oRYQxK8Zq2NAcnAghMfkfO4POosBkd5TUdGP7GZ/HFseigqIhhW1sLel504dnTSPT/fDZWrtqFQ8dv4sbdCNx+GCOvO49i8eBJAh6FJOHB03g5uO/KjTDs8DgFh43+MFnogn6fz8bz0Fi8eNklx85oANEcR/17Xe9nOlDqICq1dlYxqH5qvUSWErm894cwomSQeI8OhoQLk/VcN4mSwGdIm9yj5wB6anQUDM0wUqeianKIiodTqZASVOkdP1Ouqc98sKF81WRmbwSQFhd/V5TC83aydAfBAu9x7MhZLJ63HoMHLMGqVbvR7/N5GNJ/OVYs2gUbq33YtPEQduw4id27g7Bzxyls3ngEq1d5YK7xRkwbtQFDvjTHZ78zgsMGf8lsWZntwJOHUYiLi0dzi+Y4UvzRiVJzosh5/jxU0rJs5MQfzlHJda1pcGGjyNH2BWiczaaqIBTnvejpQUREpF63cfe8OGI8qkBzkbUjhrmjg/KeWOt1KrRzMwlYsijlGa95T2+grVW2tLIvqYDPIvZp33ORytETZ0cnZzUK5GGpb0UUnA68gdUr3OBguw+PHkRj+9YjYv977LsAo4nrMfgTC8wYvB3T/uqKSd9u1l4DtmDKQFfZp7xkgjeMB7li8kgbXLrxDIeO38L0qTZyuJLLRl8sXeCE3dsPIyoyHp2dbXjz9pVUKfCERM5VOVXKYlO6gABW6+c19QHXqQDO+2kNKbFFmCmnjYhWBQ+ECWFNGJDzxBHT/UGaTERhk2EF/sjGMAIxT04gVkkJNFnZOAFqeFUBxoAbrSjl2IkuiI4SL5KNQOb9POZFm3gxwsIi5UyI5Ys2w8neCzHRGWju6EFxdSMa27tx9dJjzDa2xxbXozAz24WhX1vItljz6QFYMeWgnKC7fPJBmBkdxZhv7DB9sg0eR6Rgq1sAFsxaj6TkPDR1vJCzQfPyK3D00BWYmjhJTWtEmHaiFTNfPPNCAS8hvnf9XCstILX+1JRUATYlAgFMUUSA0hAhApTZq0IQtKJo5iq9wT4flaWo/5Ch/FMP5UDKuiGw+JkOEwHNawKekyIV8DMnq5mp2n+o0P7ns0hNb9+RuhuQmJCgH7SwoBjuOw/BdOFGfPmnGVhn5YmCggrUtXRIzD6WZ8hla2fI1bV1IT2jSBBkbeMJ30NXYTTeHiO/tMHC0fuwcsohOax7yOerYWO9ByFR6Vi6xAUb1+9HbUMbSmqa5TlyKGBeGcrruYu+C9evhuDPv5qKiaPN4b7jKKKjklBSWoyXuv+EoU7kmrg2+gMUSap0n4DmOskdvIe6gJ/J7QQ8gc5jPnnN5/A3RgKICCJY9GSjzg/gRm0qDAKLJiiBxwHINsQ+J0AvkOFUPkyFFYgQcgvv5QTIKeoz2ZcbFQhwfiaV5eRkiXw9E3gb82etxyH/i3JGz5/+dzKu3ghDVkmVJGdYKMWKNQKNmTG+Smta0NLeDa89gViy0Fk2U2zfeRxjBq3GsM8tMbTvMgScCEZIeBpmT7dH0MlgdHS/QWF5g9QCpRdUSsUdnxudwixcIeJS8vH1JzPh5OArJ6osX+QMkznrcP9umBAJrReuUVE1jQj6AYQHv+OLCKCMp3giXEi0RBKtPBIp4Uci5W/UrUQIpQP7UjpQLEksiDepmA+pmCakcip4SrrKiFF7U8QoEURM0+xS2TMOSvZSeQNSPpUvayfra5uE9dfb7pWTT94CePg0EZ/8ZgpOnX8ohbssD+FBqqwRYtYqPb9cqiNyimvkxKue1+/x8H4UjKdYIyDwLjLyK3Dm/D0kZxTh6PFg+fus+PgcdL54J4n8vJJaOVWXOWaFAAKfhWDcEf/5H6dj08aD6HrzAS3tPXj0MAaWZtuxcK4DsrM1x6mnhxZU+0cZP8KLCtrQYCFcaIyw8V5VN6XyAarvDw0SyQeoWL4Wy9aUEM0mIoUYI6sop4JUrhwQ3q+ipOp3Ffnjb2Ka1VUjOioRc41tcfP6M7x5D+1wpcZuXLn+HJ/91hhXg8ORkFX0NwhIUwgoUSWKDbJTsrS8HpZmO7He3kdOQFy/zkcAV1XTKiWLzISxqloQUFQthV56BKQVIC6jEE8jU/HNp3NgY+WJstpWyTF0vniPV2+B+/ejMMfYDvduP8fr1zyWTKts47q4RoFRaalQseQKqqsELrwmzMgZLMnhNQ0VGjF8hsC5oV7v1EplnDJDmXygWUZxomE1T6q/lBlK5amyPPI5KVFvllGmcVe4MkNJGXwWHZCKslrMn2mP+IQcNLS/QEFpg2S4Khs6cSLwLj77nTFuP4hBfGaRJh4+4oAKPQeQorVy82Y5noDccNDvIqaTGw5fw8s3kDpRw2S8hoAqEWOCgEwdAtIKEBaTgYFfmcB08TYUVjZp2beKJtmPUN3MrFoLVi11wZlTV5GTq4UgKEpY6abyxSQ2+UeRV7rcQUaGfjcRG6VDbz7gBcIjIvTSQXn8ffh3rqRmveNh4IjxpZwvXhPjdFR4jzJdq6u0chZxzHQnyKpnsXB38lgLzJhig5t3IqXsPLe0Vs4F5eF5/v5X8PkfjXE/JEEqFGKS/y8E1EpaUfLCkppkNqwDnS8/oL3ztQCfG7X5PX/XI4AnJRYaIEA4oFAO7GZtKv0MhrXJJeo0xfyKBukTfDcKq0zdMGOyNVpbtXwBS07E2aquFipnKYpaa69DqcGML8JIk/W9pxDzfvalg0dCleNqVLaKjaYhsaxsWlpChq44ZZ06SZaNWFZxcuoNUgQLV9m4RfRXPxmLT35tjD/9aqr8/dPTiDRtD0BJNfZ6nhEr6GFokr5Cje96BLA8RVcjqhBACmf5OcVFFQ92LahCXkE1GlteoK37LcprtLJ0HlcpCDDkAIWA1HypSx05eAXGjbCQkhjuXcgqrca5y08wc5o9Pv2NMX71kwkYP8JCMz9TUz/a80UON9QBdEw/dto0acHvxGnVhViUdKBIYuvT2t7uRc1O7KhgFGUXKZnswmvlkPGlHBJeV1Rq/71I5FGuUYxpzka9yH6eqP7bn09G389M8PnvZ+KLP83AuJEWuHw9FKmFZXDddgRffzJL8rbqAFXhALGCdAhQ+wRK61Fc3oDm9lfIzK2Ay+Yj8q9Ik4fZY+JQGzm+/tSpe6hv6ZZzhUQJy4aMv4MA3ThTJqzFoH6L8DAsQUoSXV2PYvTwlej72Rx88utp+PQ3MzFrmp38x4wWhONZD6oaukWonva9odNG4uM1X4Qfv9OcVO0/ediHfZVEMdABlaKdFVaptXkeMv96g000u64qQD6nper/R4ZmLGWfYSgiPSMVt24+wl9+NxN/+d0sjBqySk7K8vK/hHGjLHD6ymNscT2Mbz6djTuP4xCvO0T7bxCgE0GsD+IJuk+eJWLM96sx6FMzDPncAv1/b4oBf1yOUX3tMOQvllizyl12x5RVt2gIKNZOy/0hB8RnFGPhAhd8+aeZuHA9RP6CcRZ3VJ64DWeXw5g4eg3+5ydj4ecThELdEcxa8DFarwPEYjSohqAprrKDVJ0JiQkSzGPjSZO8t9dC1A7ukINbldVCEaJsW9rsFEP8wx06IRyEWpuJd4ostQmNJinzvuxLbqCtTO5hXx75u3TBRnz+hxkSr9+3/6wcWeNz8AomjbfC8mVuGPDlfNy8Hy2moXaauoEO0COgBuU1LUjOKMbo7y0wpp+dHJNjNMwRgYduw3fPJYzsb4Vx363D4E/XYL2ttxxr8HfNUEFAgZijlpZ7RQcZT7XBzOn2uHIrHDt3nZAN214+FzBptJn+z6NJXOVlZbJORlIJI1I1P9MCIhz4d7b8TCDzfsKAyOJvvIe/aT6StlmPhMu/NJfiXBWfZmfKeQJXw6p2jDFlGTuo5AmdE/oBtIxydSV47MPzR4lEWk8vX3bLLpI//HKyFEI9j83AbvdARCblYr2jL/7yOyMM/tYUV25H6BCg4wDWieaUiGLUNmrUyCHc/P+BwX+xwIh+/LelHXh4o1cX7Vp/GUO/sMTEQQ4Y/s1qPHwcj/K61l4doNsTxudTz3AOG5wO4A+/mIgh3y3FyXMPsWPXcTyLTJX9C9zNudZihzyba6EuVOEEUj4PLySns1EspSQn65M3/F2F1enEaqEIlRNuEWmh9wPUFiVSLR0uyictFFGsS69pf+lEyiYmFSL42TA0of6miZ+pEyiycrJzYW2+A25bA7B4wRbkldfiyo1Q7PM+j3uhSZg8fi36f2GC89ee6fdq6TmA50jnlYsnm1dWj9zCahiNscfwr9fAaOgWzBvuifhIDfFsux0vYM6wPRg5wAojv7aBm2sAapu7eq0gtScsrVBqQUOi0yV9+elvpsLd8yy27z6JyzdDcfN+JKaMs8Kp0/flfzEP+p2W3LRKKxKY5HSuk4TLzyoUQQNGpVQJI5qg5AT+xnsIX30oQldbpE/KEyMEGhHAzuQEApnAJWYJbJUnfv36FfJyc+Xh/My/NCRbsQ/vJQIKCvJwLugGFs52RGR8FrbvOIk15u7IKqmB576zOH/1iWxXHfj1Qpy68FjEgmYFGThiREBBJYoqGxEdl4MJ39tiVH8rzBuzC7OHeCDQ76EA/+WLVzCf6YfF430xdqAdRve1g7X5XlQ1tAsHaAW5JXpHjCbvo+epUhM6cdxq+Xds/0OXEfwgGmOGr8D5y09wJTgMrlsOCRfwv2Iohgh4SgBSPtfJqDCr5ag3CQeKZsKEv5FwqRNI+fyN97AvpYOWlNdOpzSIhmpmKIFIs1MpFiKBYkYfDUxN1eeDGaRKSkzUhyaopNhXKZpH9yNga70HJ0/fxvFTd7BmtQccHX3lmALGcg6fvIUxw81w9OQd3X8J/MATziUCKlBY0Yj4xDxMGroOo/pbY/ZIN5hy7+7EfYgKSUfA/juYNdQdy6f4Y/QAK4zpZw87Sy9UN3YY+AGKA/ifNUV4EJqEtWv3SOCO4jD4fjSmTrCUhP/hgBs4euImbt+Lhsf2AFkLYaBKSygdWIClMn78nspXiSASozLNVS7asCyFZqleBBUVFXnRBKUSpsKgKCKl80YClAkEKmLKd/5OOcgJ8DPZilgnwPmZ3EMKoIJpa2vHJof9GP7Xpbh5L1Ioyn3PaRhPscNer7Pw8DojacZJ49bA9/A1RCTk6LzUXiuIlCtFuiW1stlivtFGjPzGWqh88Vhf7Ha4hPKSWiRE5mCl0QHMHb0bIwesxogvLbF/3znUNnf+HQQUCgfceRoPS0sPjBq2EkcCb2HBPCdsczsO34OXcersPZQ1tMF0qSscbT2lnIXrVkc18J1ER72ptiAxEkrDhflnBuVI1IQfKZ8w4T28l33YlzB/aXhYB81Q5kqVGUrA8q/8GJplo4xXxanaZ+6S18xQyjyGLVQBKourHGx3YavzQTx5noypE63h5OSPi9eewdn1CAYPXCKh5TWW7uKN+hy8KudGSJRSt2HPEAE5RdVi3/v4XMT3n1lg3Hd2GPmVLby33UBP5ysU5dZg8SQPDPpiBYyGumDkQHMkJOfLWdGihA3N0HQtFHHzYTRWr3HHoH4LsXihsyBgn/cFnDx9RwyAVSu2Y63VHjk39Njhc7IuSgUWVXGnOxuJk6alClxSWqh/12Yj5WdnaWYorUbmRgxDEYU0Qzs7O2WbKlOSuXl5wgkq10lMks1I+cQoB+Q1f+eDqER4TW4hwigTGTns6GyXqrRtWw/j6q3nAtSdu05i/pwNWL7cTSrbBnwxDyMGmWL0kOXYu/8CnkSmiiWkRBA3TiglnFNcjeKKRpTXtmKpyVYM/9IGE753wOAvVmHSkHUY0W81hvddDeOhLhj45+U4evSG7CHIK67RI4D7AVQ0lBtDrt4Jx/LlWxF8/THmTrfF8uU7sHP3STx+noz5sx1hZb1X9i/s3HFCchFdneTwZqFmZv5IqAzCcd3kfMKFO18oHfgbX7ym0uVvSjqwHKXD4B/MpS6I8qlFpwOKS0q0AJNBjYvaIcMWbxCKoF7gn/oY5j0ZxmhtbYKLkx+OnQjG8VO34eN/AXcex+De01i4bj2GebM34g+/moQj/pewf+8pbNh8EA/CkmR/LhHAd0EAdYAOAQyWVde3y6a71at2Y/AXZhjx9VqM+sYOY/ray0EaI781x7GjN+Q4G3rNvcE4tS2pWMQPjyI4f/0ZlptuRXJSGpYtcsLSJVtxKOAmpvBUFUtPEZMHj13H07BkrLXYqZPjWkiZ8DKUFipw2RuKSNYHLvkdTdZ83V/dUkyxCoPxITbJCatwtPrvdQlH64pWyQmkbAk3NzejnKHWioqPwtMqHE3uoZMSFhqNGVMsMXOaLXx8L4p5dyjgOrz9LspxZTvcT2HqJFucOn4NPvtOwcHJX8rRuWVU4jIKAUoEMR8gmzQaZV8AAXz3bhRcNh3BmhUeWGuxD157zyElrUj+W4CHcxBR+T9EQHYxknNL8fB5shT0rli6VUxlty3+mDdnI2Ya22H+HCf4Hb6KkMhUBN+LwmqLXRJKb2rSApH6cLSucFn7zHC0VnqpwtHUEwpGhqFs9vk/w9Hh4b3haB4uqoWj5aOuKqK3KoCyzjAczTAG2ZHN3todN4LDEBmfIzF7loLbrN0LR0c/LFvmJon2CaMsEBEeg+0u/rB38Jd9AXTI6Izxf8YoMmiCcqNGHnMB5Q3642qq6zvkDxxevAW6Xr5H9yvINU9Q5Oa88motIlpQVifIIxK5O5KynX9REnie1dR3sXyxC7Izc6SeZ9XSLRjw1Vy47TgBD8/TMJnnhIXzt8Bjz2msMd+By+cfyNpY2aDyxUSCtgFFq/BjFlD9+TObFo7WpAO9YtZEqaoIfTiaW5QIdGKEmFOySQux8tQPdeqfFnZWIVb1WQvBqj8o00K2Fsu3Ij23DIVVTWjqfisH6V24EgIX12MigqaMt8JB3zOyZ8zL4wSsbfbh+v0onLnyFJeCn+PW4zjcfhyHO0/i5Y8W7oUk4v6zJDwITcbDsFQ84ut5Gh6HG7zC0vH4eZr89jAsRe699ywJ954misVz+0k8bjyIxtlrITh3LUTGWjTfCfm5BejoaEXQiesY/O1SLFywGRbmu3DxegjyyutR3/lK/kLFzNQF5eVa4dr/X+hebcTgi1TfC6N6VFZUSjWE2klJbumTlpZ2RLnUbESGCqqx8QEUM6oR60SUakVFxfIw1chJzk7+WGfDo95PYveOIOzacRq7dp6G85YAuLicgMmczYgM14JWEaEpmDxuLRzWH8A6Wz/Y2fjCxnq/vGzXesNmrTds1/rCdq2PvGz4uw0/+8F2rb/uN91nG8OXP+xs/LXrtf6wsfaDzVpfrLP1h/26A7C29Mbc6ev1Yffga2H44g9zYGXpDY+9F3H46G0cP/EI586E4srlCCye64z8PM2CIeBUJQibinSqRqSomig2VbaoGqWL4pr/D3B0yklYzExBAAAAAElFTkSuQmCC" alt="Asistente MADEVAL" />
        <h2>Antes de empezar</h2>
      </div>
      <p class="file-meta">Ingresa con tu usuario asignado para iniciar.</p>
      <input class="text-input" id="gateName" type="text" placeholder="Usuario" autocomplete="username" />
      <input class="text-input" id="gatePassword" type="password" placeholder="Contraseña" autocomplete="current-password" />
      <button class="primary" type="button" id="saveNameBtn"><i data-lucide="check"></i>Continuar</button>
      <div class="file-meta" id="gateMessage"></div>
    </div>
  </div>

  <div class="modal-backdrop" id="adminLogin">
    <div class="modal">
      <div class="panel-header" style="padding:0 0 10px;">
        <h2>Super usuario</h2>
        <button class="icon-button" type="button" id="adminLoginClose" title="Cerrar"><i data-lucide="x"></i></button>
      </div>
      <input class="text-input" id="adminUser" type="text" placeholder="Usuario" autocomplete="username" />
      <input class="text-input" id="adminPassword" type="password" placeholder="Contraseña" autocomplete="current-password" />
      <button class="primary" type="button" id="adminLoginBtn"><i data-lucide="lock-keyhole"></i>Ingresar</button>
      <div class="file-meta" id="adminLoginMessage"></div>
    </div>
  </div>

  <section class="admin-panel" id="adminPanel">
    <div class="panel-header">
      <h2>Opciones de super usuario</h2>
      <button class="icon-button" type="button" id="adminClose" title="Cerrar"><i data-lucide="x"></i></button>
    </div>
    <div class="admin-body">
      <div class="control-box">
        <strong>Reporte de consultas</strong>
        <div class="file-meta">Descarga el historial ordenado por usuario, preguntas, respuestas, fecha y calificacion.</div>
        <div class="btn-row">
          <input class="text-input" id="reportStart" type="date" title="Fecha inicial" />
          <input class="text-input" id="reportEnd" type="date" title="Fecha final" />
          <select class="text-input" id="reportUser" title="Filtrar por usuario">
            <option value="">Todos los usuarios</option>
          </select>
        </div>
        <div class="btn-row">
          <button class="secondary" type="button" id="downloadReportBtn"><i data-lucide="file-spreadsheet"></i>Excel editable</button>
          <button class="secondary" type="button" id="downloadReportPdfBtn"><i data-lucide="file-text"></i>PDF</button>
        </div>
        <div class="report-list" id="reportList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Resumen por usuario</strong>
        <div class="file-meta">Consultas, promedio de calificacion y pendientes de calificar.</div>
        <div class="report-list" id="summaryList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Actividad y temas</strong>
        <div class="file-meta">Usuarios mas activos y temas mas preguntados.</div>
        <div class="report-list" id="activityList"></div>
        <div class="report-list" id="topicList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Preguntas sin respuesta</strong>
        <div class="file-meta">Consultas donde el agente necesita mejor base o reglas.</div>
        <div class="report-list" id="unansweredList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Base de conocimiento</strong>
        <div class="file-meta">Estado de archivos locales, indice y carpeta de informes TXT.</div>
        <div class="report-list" id="knowledgeStatus"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Permisos por rol</strong>
        <div class="file-meta">Referencia rapida para saber que puede hacer cada tipo de usuario.</div>
        <div class="report-list" id="permissionList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Usuarios autorizados</strong>
        <div class="file-meta">Invitado usa el chat. Super usuario usa el chat, reportes y administracion.</div>
        <input class="text-input" id="managedUsername" type="text" placeholder="Usuario" autocomplete="off" />
        <input class="text-input" id="managedPassword" type="password" placeholder="Nueva contrasena" autocomplete="new-password" />
        <div class="btn-row">
          <select class="text-input" id="managedRole">
            <option value="invitado">Invitado</option>
            <option value="super_usuario">Super usuario</option>
          </select>
          <select class="text-input" id="managedActive">
            <option value="1">Activo</option>
            <option value="0">Desactivado</option>
          </select>
        </div>
        <button class="primary" type="button" id="saveManagedUserBtn"><i data-lucide="user-plus"></i>Guardar usuario</button>
        <div class="file-meta" id="managedUserMessage"></div>
        <div class="report-list" id="adminUsersList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Mejorar respuestas</strong>
        <div class="file-meta">Respuestas con calificacion baja para revisar la base o ajustar reglas.</div>
        <div class="report-list" id="lowRatingList"></div>
      </div>
      <div class="control-box" style="margin-top: 12px;">
        <strong>Auditoria de accesos y cambios</strong>
        <div class="file-meta">Registro interno de ingresos y cambios realizados por super usuarios.</div>
        <div class="report-list" id="auditList"></div>
      </div>
    </div>
  </section>

  <script>
    const filesEl = document.querySelector("#files");
    const fileCountEl = document.querySelector("#fileCount");
    const messagesEl = document.querySelector("#messages");
    const reportListEl = document.querySelector("#reportList");
    const summaryListEl = document.querySelector("#summaryList");
    const unansweredListEl = document.querySelector("#unansweredList");
    const knowledgeStatusEl = document.querySelector("#knowledgeStatus");
    const adminUsersListEl = document.querySelector("#adminUsersList");
    const lowRatingListEl = document.querySelector("#lowRatingList");
    const reportStartEl = document.querySelector("#reportStart");
    const reportEndEl = document.querySelector("#reportEnd");
    const reportUserEl = document.querySelector("#reportUser");
    const activityListEl = document.querySelector("#activityList");
    const topicListEl = document.querySelector("#topicList");
    const permissionListEl = document.querySelector("#permissionList");
    const auditListEl = document.querySelector("#auditList");
    const managedUsernameEl = document.querySelector("#managedUsername");
    const managedPasswordEl = document.querySelector("#managedPassword");
    const managedRoleEl = document.querySelector("#managedRole");
    const managedActiveEl = document.querySelector("#managedActive");
    const saveManagedUserBtn = document.querySelector("#saveManagedUserBtn");
    const managedUserMessageEl = document.querySelector("#managedUserMessage");
    const adminFileCountEl = document.querySelector("#adminFileCount");
    const nameGate = document.querySelector("#nameGate");
    const gateNameEl = document.querySelector("#gateName");
    const gatePasswordEl = document.querySelector("#gatePassword");
    const gateMessageEl = document.querySelector("#gateMessage");
    const saveNameBtn = document.querySelector("#saveNameBtn");
    const localClockEl = document.querySelector("#localClock");
    const logoutUserBtn = document.querySelector("#logoutUserBtn");
    const adminOpen = document.querySelector("#adminOpen");
    const adminLogin = document.querySelector("#adminLogin");
    const adminLoginClose = document.querySelector("#adminLoginClose");
    const adminLoginBtn = document.querySelector("#adminLoginBtn");
    const adminUserEl = document.querySelector("#adminUser");
    const adminPasswordEl = document.querySelector("#adminPassword");
    const adminLoginMessage = document.querySelector("#adminLoginMessage");
    const adminPanel = document.querySelector("#adminPanel");
    const adminClose = document.querySelector("#adminClose");
    const downloadReportBtn = document.querySelector("#downloadReportBtn");
    const downloadReportPdfBtn = document.querySelector("#downloadReportPdfBtn");
    const askForm = document.querySelector("#askForm");
    const questionEl = document.querySelector("#question");
    let adminToken = sessionStorage.getItem("kb_admin_token") || "";
    let activeUserName = localStorage.getItem("kb_user_name") || "";

    function icons() {
      if (window.lucide) window.lucide.createIcons();
    }

    function addMessage(text, role) {
      const el = document.createElement("div");
      el.className = `message ${role}`;
      el.textContent = text;
      messagesEl.appendChild(el);
      messagesEl.scrollTop = messagesEl.scrollHeight;
      return el;
    }

    function appendInlineText(parent, text) {
      const parts = text.split(/(\*\*[^*]+\*\*)/g);
      for (const part of parts) {
        if (part.startsWith("**") && part.endsWith("**")) {
          const strong = document.createElement("strong");
          strong.textContent = part.slice(2, -2);
          parent.appendChild(strong);
        } else if (part) {
          const urlParts = part.split(/(https?:\/\/[^\s|)]+)/g);
          for (const urlPart of urlParts) {
            if (/^https?:\/\//.test(urlPart)) {
              const link = document.createElement("a");
              link.href = urlPart;
              link.target = "_blank";
              link.rel = "noopener";
              link.textContent = urlPart;
              parent.appendChild(link);
            } else if (urlPart) {
              parent.appendChild(document.createTextNode(urlPart));
            }
          }
        }
      }
    }

    function isTableSeparator(line) {
      return /^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$/.test(line);
    }

    function parseTableRow(line) {
      return line.trim().replace(/^\|/, "").replace(/\|$/, "").split("|").map((cell) => cell.trim());
    }

    function appendParagraph(container, lines) {
      const text = lines.join("\n").trim();
      if (!text) return;
      const p = document.createElement("p");
      appendInlineText(p, text);
      container.appendChild(p);
    }

    function appendTable(container, lines) {
      const headers = parseTableRow(lines[0]);
      const rows = lines.slice(2).map(parseTableRow).filter((row) => row.length);
      const wrap = document.createElement("div");
      wrap.className = "message-table-wrap";
      const table = document.createElement("table");
      const thead = document.createElement("thead");
      const headRow = document.createElement("tr");
      for (const header of headers) {
        const th = document.createElement("th");
        appendInlineText(th, header);
        headRow.appendChild(th);
      }
      thead.appendChild(headRow);
      table.appendChild(thead);
      const tbody = document.createElement("tbody");
      for (const row of rows) {
        const tr = document.createElement("tr");
        for (const cell of row) {
          const td = document.createElement("td");
          appendInlineText(td, cell);
          tr.appendChild(td);
        }
        tbody.appendChild(tr);
      }
      table.appendChild(tbody);
      wrap.appendChild(table);
      container.appendChild(wrap);
    }

    function renderAssistantText(container, text) {
      const lines = text.split("\n");
      let paragraph = [];
      for (let index = 0; index < lines.length; index += 1) {
        const line = lines[index];
        const next = lines[index + 1] || "";
        if (line.includes("|") && isTableSeparator(next)) {
          appendParagraph(container, paragraph);
          paragraph = [];
          const tableLines = [line, next];
          index += 2;
          while (index < lines.length && lines[index].includes("|")) {
            tableLines.push(lines[index]);
            index += 1;
          }
          index -= 1;
          appendTable(container, tableLines);
          continue;
        }
        paragraph.push(line);
      }
      appendParagraph(container, paragraph);
    }

    async function rateAnswer(interactionId, rating, box) {
      if (!interactionId) return;
      const buttons = box.querySelectorAll(".rating-btn");
      buttons.forEach((button) => {
        button.disabled = true;
        button.classList.toggle("selected", Number(button.dataset.rating) === rating);
      });
      const status = box.querySelector(".rating-status");
      status.textContent = "Guardando...";
      const response = await fetch("/api/rate", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ interaction_id: interactionId, rating, ...browserLocalMeta() })
      });
      const data = await response.json();
      status.textContent = data.ok ? "Gracias, calificacion guardada." : (data.message || "No se pudo guardar.");
      if (adminPanel.classList.contains("active")) await refreshReport();
    }

    function appendRatingBox(parent, interactionId) {
      if (!interactionId) return;
      const box = document.createElement("div");
      box.className = "rating-box";
      const label = document.createElement("span");
      label.textContent = "Califica esta respuesta:";
      box.appendChild(label);
      for (let rating = 1; rating <= 5; rating += 1) {
        const button = document.createElement("button");
        button.type = "button";
        button.className = "rating-btn";
        button.dataset.rating = String(rating);
        button.textContent = String(rating);
        button.title = `${rating} de 5`;
        button.addEventListener("click", () => rateAnswer(interactionId, rating, box));
        box.appendChild(button);
      }
      const status = document.createElement("span");
      status.className = "rating-status";
      box.appendChild(status);
      parent.appendChild(box);
    }

    function addAssistantMessage(text, images = [], interactionId = null) {
      const el = document.createElement("div");
      el.className = "message assistant";
      const textEl = document.createElement("div");
      renderAssistantText(textEl, text);
      el.appendChild(textEl);
      if (images.length) {
        const grid = document.createElement("div");
        grid.className = "image-grid";
        for (const image of images) {
          const item = document.createElement("a");
          item.className = "image-result";
          item.href = image.url;
          item.target = "_blank";
          item.rel = "noopener";
          const img = document.createElement("img");
          img.src = image.url;
          img.alt = image.name;
          const caption = document.createElement("span");
          caption.textContent = image.name;
          item.appendChild(img);
          item.appendChild(caption);
          grid.appendChild(item);
        }
        el.appendChild(grid);
      }
      appendRatingBox(el, interactionId);
      messagesEl.appendChild(el);
      messagesEl.scrollTop = messagesEl.scrollHeight;
      return el;
    }

    function humanSize(bytes) {
      if (bytes < 1024) return `${bytes} B`;
      if (bytes < 1024 * 1024) return `${(bytes / 1024).toFixed(1)} KB`;
      return `${(bytes / 1024 / 1024).toFixed(1)} MB`;
    }

    async function refreshFiles() {
      const response = await fetch("/api/files");
      const data = await response.json();
      if (fileCountEl) fileCountEl.textContent = data.files.length;
      if (adminFileCountEl) adminFileCountEl.textContent = data.files.length;
      if (!filesEl) return;
      filesEl.innerHTML = "";
      for (const file of data.files) {
        const row = document.createElement("div");
        row.className = "file";
        row.innerHTML = `
          <div class="file-icon"><i data-lucide="${file.is_image ? "image" : "file-text"}"></i></div>
          <div>
            <div class="file-name"></div>
            <div class="file-meta">${humanSize(file.size)} - ${file.extension}</div>
          </div>`;
        row.querySelector(".file-name").textContent = file.path || file.name;
        filesEl.appendChild(row);
      }
      icons();
    }

    async function refreshReport() {
      if (!adminToken) return;
      const response = await fetch(`/api/report?${reportQuery()}`, { headers: { Authorization: `Bearer ${adminToken}` } });
      if (!response.ok) return;
      const data = await response.json();
      renderAdminOverview(data);
      reportListEl.innerHTML = "";
      if (!data.interactions.length) {
        const empty = document.createElement("div");
        empty.className = "file-meta";
        empty.textContent = "Todavia no hay historial.";
        reportListEl.appendChild(empty);
        return;
      }
      for (const item of data.interactions) {
        const row = document.createElement("div");
        row.className = "report-item";
        const date = new Date(item.created_at);
        row.innerHTML = `<strong></strong><span></span>`;
        row.querySelector("strong").textContent = `${item.user_name} - ${date.toLocaleString()}`;
        const rating = item.rating ? `Calificacion: ${item.rating}/5` : "Calificacion pendiente";
        row.querySelector("span").textContent = `${item.question} - ${rating}`;
        reportListEl.appendChild(row);
      }
    }

    function renderList(container, items, emptyText, renderItem) {
      container.innerHTML = "";
      if (!items || !items.length) {
        const empty = document.createElement("div");
        empty.className = "file-meta";
        empty.textContent = emptyText;
        container.appendChild(empty);
        return;
      }
      for (const item of items) {
        const row = document.createElement("div");
        row.className = "report-item";
        renderItem(row, item);
        container.appendChild(row);
      }
    }

    function renderAdminOverview(data) {
      const selectedUser = reportUserEl.value;
      reportUserEl.innerHTML = "<option value=\"\">Todos los usuarios</option>";
      for (const user of data.admin_users || []) {
        const option = document.createElement("option");
        option.value = user.username;
        option.textContent = user.username;
        reportUserEl.appendChild(option);
      }
      reportUserEl.value = selectedUser;
      renderList(summaryListEl, data.summary || [], "Todavia no hay resumen.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = item.user_name;
        row.querySelector("span").textContent = `${item.consultas} consultas | promedio ${item.promedio || "sin calificar"} | pendientes ${item.pendientes} | temas: ${item.temas || "sin datos"}`;
      });
      renderList(activityListEl, data.activity || [], "Todavia no hay actividad.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = item.user_name;
        row.querySelector("span").textContent = `${item.consultas} consultas | ultima actividad: ${item.last_seen || "sin registro"}`;
      });
      renderList(topicListEl, data.topics || [], "Todavia no hay temas frecuentes.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = item.topic;
        row.querySelector("span").textContent = `${item.count} menciones`;
      });
      renderList(unansweredListEl, data.unanswered || [], "No hay preguntas sin respuesta.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = `${item.user_name} - ${new Date(item.created_at).toLocaleString()}`;
        row.querySelector("span").textContent = item.question;
      });
      const status = data.knowledge || {};
      renderList(knowledgeStatusEl, [status], "Sin datos de base.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = item.index_status || "Estado no disponible";
        row.querySelector("span").textContent = `${item.files || 0} archivos | TXT local: ${item.txt_report_dir || ""}`;
      });
      renderList(permissionListEl, data.permissions || [], "Sin permisos configurados.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = item.role_label;
        row.querySelector("span").textContent = item.description;
      });
      renderList(adminUsersListEl, data.admin_users || [], "No hay usuarios configurados.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span><div class=\"btn-row\" style=\"margin-top:8px;\"><button class=\"secondary\" type=\"button\">Editar</button></div>";
        row.querySelector("strong").textContent = item.username;
        row.querySelector("span").textContent = `${item.role_label || item.role} | ${item.status}`;
        row.querySelector("button").addEventListener("click", () => {
          managedUsernameEl.value = item.username;
          managedRoleEl.value = item.role;
          managedActiveEl.value = item.active ? "1" : "0";
          managedPasswordEl.value = "";
          managedPasswordEl.placeholder = "Nueva contrasena opcional";
          managedUserMessageEl.textContent = "Usuario cargado para editar.";
        });
      });
      renderList(lowRatingListEl, data.low_ratings || [], "No hay respuestas mal calificadas.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = `${item.user_name} - ${item.rating}/5`;
        row.querySelector("span").textContent = item.question;
      });
      renderList(auditListEl, data.audit || [], "Todavia no hay auditoria.", (row, item) => {
        row.innerHTML = "<strong></strong><span></span>";
        row.querySelector("strong").textContent = `${item.created_at} - ${item.actor}`;
        row.querySelector("span").textContent = `${item.action_label}: ${item.target_user || ""} ${item.detail || ""}`.trim();
      });
    }

    function reportQuery() {
      const params = new URLSearchParams({ token: adminToken });
      if (reportStartEl.value) params.set("start", reportStartEl.value);
      if (reportEndEl.value) params.set("end", reportEndEl.value);
      if (reportUserEl.value) params.set("user", reportUserEl.value);
      return params.toString();
    }

    async function saveManagedUser() {
      if (!adminToken) return;
      managedUserMessageEl.textContent = "Guardando...";
      const response = await fetch("/api/admin/users", {
        method: "POST",
        headers: {
          "Content-Type": "application/json",
          Authorization: `Bearer ${adminToken}`
        },
        body: JSON.stringify({
          username: managedUsernameEl.value.trim(),
          password: managedPasswordEl.value,
          role: managedRoleEl.value,
          active: managedActiveEl.value === "1",
          ...browserLocalMeta()
        })
      });
      const data = await response.json();
      managedUserMessageEl.textContent = data.ok ? "Usuario guardado." : (data.message || "No se pudo guardar.");
      if (data.ok) {
        managedPasswordEl.value = "";
        await refreshReport();
      }
    }

    function currentUserName() {
      return activeUserName.trim();
    }

    function padTimePart(value) {
      return String(value).padStart(2, "0");
    }

    function browserLocalMeta() {
      const now = new Date();
      return {
        client_timestamp: [
          now.getFullYear(),
          padTimePart(now.getMonth() + 1),
          padTimePart(now.getDate())
        ].join("-") + "T" + [
          padTimePart(now.getHours()),
          padTimePart(now.getMinutes()),
          padTimePart(now.getSeconds())
        ].join(":"),
        client_timezone: Intl.DateTimeFormat().resolvedOptions().timeZone || ""
      };
    }

    function updateLocalClock() {
      const now = new Date();
      const dateParts = new Intl.DateTimeFormat("es-EC", {
        day: "numeric",
        month: "numeric",
        year: "numeric"
      }).formatToParts(now).reduce((acc, part) => {
        acc[part.type] = part.value;
        return acc;
      }, {});
      const timeText = new Intl.DateTimeFormat("es-EC", {
        hour: "2-digit",
        minute: "2-digit",
        hour12: false
      }).format(now);
      localClockEl.textContent = `${dateParts.day}/${dateParts.month}/${dateParts.year} ${timeText}`;
    }

    async function logoutUser() {
      if (adminToken) {
        await fetch("/api/admin/logout", {
          method: "POST",
          headers: { "Content-Type": "application/json", Authorization: `Bearer ${adminToken}` },
          body: JSON.stringify(browserLocalMeta())
        }).catch(() => {});
      }
      adminToken = "";
      activeUserName = "";
      sessionStorage.removeItem("kb_admin_token");
      localStorage.removeItem("kb_user_name");
      adminPanel.classList.remove("active");
      adminLogin.classList.remove("active");
      gateNameEl.value = "";
      gatePasswordEl.value = "";
      gateMessageEl.textContent = "";
      nameGate.classList.add("active");
      setTimeout(() => gateNameEl.focus(), 50);
    }

    updateLocalClock();
    setInterval(updateLocalClock, 30000);
    gateNameEl.value = activeUserName;
    if (!activeUserName) {
      nameGate.classList.add("active");
      setTimeout(() => gateNameEl.focus(), 50);
    }

    function randomGreeting(name) {
      const greetings = [
        `Hola ${name}, hoy es un gran dia. En que puedo ayudarte?`,
        `Hola ${name}, listo para ayudarte. Que necesitas revisar?`,
        `Buen dia ${name}. En que puedo apoyarte?`,
        `Hola ${name}, estoy atento a tu requerimiento.`,
        `Bienvenido ${name}. Como puedo ayudarte hoy?`,
        `Hola ${name}, cuentame que necesitas consultar.`,
        `Buen dia ${name}. Estoy listo para ayudarte.`,
        `Hola ${name}, dime tu requerimiento y lo reviso.`,
        `Saludos ${name}. En que puedo asistirte?`,
        `Hola ${name}, podemos empezar cuando gustes.`,
        `Bienvenido ${name}. Que deseas consultar?`,
        `Hola ${name}, estoy listo para revisar tu solicitud.`,
        `Buen dia ${name}. Cual es tu requerimiento?`,
        `Hola ${name}, dime como puedo ayudarte.`,
        `Saludos ${name}. Que necesitas validar?`,
        `Hola ${name}, estoy disponible para ayudarte.`,
        `Bienvenido ${name}. Que revisamos hoy?`,
        `Hola ${name}, comparte tu consulta y avanzamos.`,
        `Buen dia ${name}. Como puedo apoyarte ahora?`,
        `Hola ${name}, listo para atender tu consulta.`,
        `Saludos ${name}. Que informacion necesitas?`,
        `Hola ${name}, dime que deseas revisar.`,
        `Bienvenido ${name}. Estoy atento a tu consulta.`,
        `Hola ${name}, en que te puedo apoyar?`,
        `Buen dia ${name}. Revisemos tu requerimiento.`,
        `Hola ${name}, que necesitas resolver?`,
        `Saludos ${name}. Estoy listo para ayudarte.`,
        `Hola ${name}, cuentame tu consulta.`,
        `Bienvenido ${name}. Que necesitas buscar?`,
        `Hola ${name}, dime tu requerimiento y empezamos.`
      ];
      return greetings[Math.floor(Math.random() * greetings.length)];
    }

    async function saveUserName() {
      const name = gateNameEl.value.trim();
      if (!name) {
        gateNameEl.focus();
        return;
      }
      if (!gatePasswordEl.value) {
        gatePasswordEl.focus();
        return;
      }
      gateMessageEl.textContent = "Validando usuario...";
      const response = await fetch("/api/user/login", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ username: name, password: gatePasswordEl.value, ...browserLocalMeta() })
      });
      const data = await response.json();
      if (!data.ok) {
        gateMessageEl.textContent = data.message || "No pude validar el usuario.";
        return;
      }
      activeUserName = data.username || name;
      localStorage.setItem("kb_user_name", activeUserName);
      gatePasswordEl.value = "";
      gateMessageEl.textContent = "";
      nameGate.classList.remove("active");
      addAssistantMessage(randomGreeting(activeUserName));
      questionEl.focus();
    }

    saveNameBtn.addEventListener("click", saveUserName);
    gateNameEl.addEventListener("keydown", (event) => {
      if (event.key === "Enter") saveUserName();
    });
    gatePasswordEl.addEventListener("keydown", (event) => {
      if (event.key === "Enter") saveUserName();
    });
    logoutUserBtn.addEventListener("click", logoutUser);

    adminOpen.addEventListener("click", () => {
      if (adminToken) {
        adminPanel.classList.add("active");
        refreshReport();
      } else {
        adminLogin.classList.add("active");
        setTimeout(() => adminUserEl.focus(), 50);
      }
    });

    function closeAdminSession() {
      const token = adminToken;
      adminToken = "";
      sessionStorage.removeItem("kb_admin_token");
      adminPanel.classList.remove("active");
      adminLogin.classList.remove("active");
      adminPasswordEl.value = "";
      adminLoginMessage.textContent = "";
      if (token) {
        fetch("/api/admin/logout", {
          method: "POST",
          headers: { "Content-Type": "application/json", Authorization: `Bearer ${token}` },
          body: JSON.stringify(browserLocalMeta())
        }).catch(() => {});
      }
    }

    adminLoginClose.addEventListener("click", closeAdminSession);
    adminClose.addEventListener("click", closeAdminSession);

    async function loginAdmin() {
      adminLoginMessage.textContent = "";
      const response = await fetch("/api/admin/login", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ username: adminUserEl.value.trim(), password: adminPasswordEl.value, ...browserLocalMeta() })
      });
      const data = await response.json();
      if (!data.ok) {
        adminLoginMessage.textContent = data.message || "No se pudo ingresar.";
        return;
      }
      adminToken = data.token;
      sessionStorage.setItem("kb_admin_token", adminToken);
      adminLogin.classList.remove("active");
      adminPanel.classList.add("active");
      await refreshReport();
    }

    adminLoginBtn.addEventListener("click", loginAdmin);
    adminPasswordEl.addEventListener("keydown", (event) => {
      if (event.key === "Enter") loginAdmin();
    });
    saveManagedUserBtn.addEventListener("click", saveManagedUser);
    reportStartEl.addEventListener("change", refreshReport);
    reportEndEl.addEventListener("change", refreshReport);
    reportUserEl.addEventListener("change", refreshReport);

    downloadReportBtn.addEventListener("click", () => {
      if (!adminToken) return;
      window.location.href = `/api/report.xlsx?${reportQuery()}`;
    });

    downloadReportPdfBtn.addEventListener("click", () => {
      if (!adminToken) return;
      window.location.href = `/api/report.pdf?${reportQuery()}`;
    });

    askForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      const question = questionEl.value.trim();
      if (!question) return;
      const userName = currentUserName();
      if (!userName) {
        nameGate.classList.add("active");
        gateNameEl.focus();
        return;
      }
      addMessage(question, "user");
      questionEl.value = "";
      const waiting = addAssistantMessage("Revisando la base y el contexto...");
      const response = await fetch("/api/ask", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ question, user_name: userName, ...browserLocalMeta() })
      });
      const data = await response.json();
      waiting.remove();
      addAssistantMessage(data.answer || data.message, data.images || [], data.interaction_id || null);
      if (adminPanel.classList.contains("active")) await refreshReport();
    });

    questionEl.addEventListener("keydown", (event) => {
      if (event.key === "Enter" && !event.shiftKey) {
        event.preventDefault();
        askForm.requestSubmit();
      }
    });

    refreshFiles();
    refreshReport();
    icons();
  </script>
</body>
</html>
"""


def json_response(handler: BaseHTTPRequestHandler, payload: dict, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def page_response(handler: BaseHTTPRequestHandler) -> None:
    body = PAGE.encode("utf-8")
    handler.send_response(200)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def local_now() -> datetime:
    return datetime.now().replace(tzinfo=None)


def client_now(payload: dict | None = None) -> datetime:
    if not payload:
        return local_now()
    raw = str(payload.get("client_timestamp", "")).strip()[:32]
    if not raw:
        return local_now()
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return local_now()
    return parsed.replace(tzinfo=None)


def hash_password(password: str, salt: str = "") -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return f"{salt}${digest}"


def verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash or "$" not in stored_hash:
        return False
    salt, expected = stored_hash.split("$", 1)
    return hash_password(password, salt).split("$", 1)[1] == expected


def normalize_role(role: str) -> str:
    value = str(role or "").strip().lower().replace(" ", "_").replace("-", "_")
    return value if value in USER_ROLES else "invitado"


def parse_active_flag(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() not in {"0", "false", "no", "off", "inactivo", "desactivado"}


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS interactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                date TEXT NOT NULL,
                user_name TEXT NOT NULL,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                sources TEXT NOT NULL,
                images TEXT NOT NULL,
                rating INTEGER,
                rating_note TEXT,
                rated_at TEXT
            )
            """
        )
        columns = {row[1] for row in connection.execute("PRAGMA table_info(interactions)").fetchall()}
        if "rating" not in columns:
            connection.execute("ALTER TABLE interactions ADD COLUMN rating INTEGER")
        if "rating_note" not in columns:
            connection.execute("ALTER TABLE interactions ADD COLUMN rating_note TEXT")
        if "rated_at" not in columns:
            connection.execute("ALTER TABLE interactions ADD COLUMN rated_at TEXT")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS app_users (
                username TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                date TEXT NOT NULL,
                actor TEXT NOT NULL,
                action TEXT NOT NULL,
                target_user TEXT,
                detail TEXT
            )
            """
        )
        existing_users = {
            str(row[0]).upper()
            for row in connection.execute("SELECT username FROM app_users").fetchall()
        }
        now = local_now().isoformat(timespec="seconds")
        for username, passwords in ADMIN_USERS.items():
            if username.upper() in existing_users:
                continue
            password = sorted(passwords, key=len)[0]
            connection.execute(
                """
                INSERT INTO app_users (username, password_hash, role, active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (username, hash_password(password), "super_usuario", now, now),
            )


def list_app_users() -> list[dict]:
    init_db()
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT username, role, active, created_at, updated_at
            FROM app_users
            ORDER BY username COLLATE NOCASE
            """
        ).fetchall()
    return [dict(row) for row in rows]


def get_app_user(username: str) -> dict | None:
    init_db()
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT username, password_hash, role, active, created_at, updated_at
            FROM app_users
            WHERE UPPER(username) = UPPER(?)
            """,
            (username,),
        ).fetchone()
    return dict(row) if row else None


def legacy_admin_password_ok(username: str, password: str) -> bool:
    candidates = ADMIN_USERS.get(str(username or "").upper(), set())
    normalized_password = str(password or "").replace("ñ", "n").replace("Ñ", "N")
    normalized_candidates = {
        str(candidate).replace("ñ", "n").replace("Ñ", "N")
        for candidate in candidates
    }
    return password in candidates or normalized_password in normalized_candidates


def upsert_app_user(username: str, password: str = "", role: str = "invitado", active: bool = True, event_time: datetime | None = None) -> dict:
    clean_username = str(username or "").strip()
    if not clean_username:
        raise UserFacingError("Ingresa un nombre de usuario.")
    clean_role = normalize_role(role)
    now = (event_time or local_now()).isoformat(timespec="seconds")
    existing = get_app_user(clean_username)
    if existing is None and not password:
        raise UserFacingError("Ingresa una contrasena para el usuario nuevo.")
    with sqlite3.connect(DB_PATH) as connection:
        if existing is None:
            connection.execute(
                """
                INSERT INTO app_users (username, password_hash, role, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (clean_username, hash_password(password), clean_role, 1 if active else 0, now, now),
            )
        else:
            if password:
                connection.execute(
                    """
                    UPDATE app_users
                    SET password_hash = ?, role = ?, active = ?, updated_at = ?
                    WHERE UPPER(username) = UPPER(?)
                    """,
                    (hash_password(password), clean_role, 1 if active else 0, now, clean_username),
                )
            else:
                connection.execute(
                    """
                    UPDATE app_users
                    SET role = ?, active = ?, updated_at = ?
                    WHERE UPPER(username) = UPPER(?)
                    """,
                    (clean_role, 1 if active else 0, now, clean_username),
                )
    return get_app_user(clean_username) or {}


def log_audit_event(actor: str, action: str, target_user: str = "", detail: str = "", event_time: datetime | None = None) -> None:
    init_db()
    now = event_time or local_now()
    with sqlite3.connect(DB_PATH) as connection:
        connection.execute(
            """
            INSERT INTO audit_events (created_at, date, actor, action, target_user, detail)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                now.isoformat(timespec="seconds"),
                now.date().isoformat(),
                str(actor or "sistema"),
                str(action or "evento"),
                str(target_user or ""),
                str(detail or ""),
            ),
        )


def recent_audit_events(limit: int = 30) -> list[dict]:
    init_db()
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT created_at, actor, action, target_user, detail
            FROM audit_events
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        item["action_label"] = ACTION_LABELS.get(item["action"], item["action"])
        result.append(item)
    return result


def log_interaction(user_name: str, question: str, answer: str, sources: list[dict], images: list[dict], event_time: datetime | None = None) -> int:
    init_db()
    now = event_time or local_now()
    source_names = [str(source.get("source", "")) for source in sources if source.get("source")]
    created_at = now.isoformat(timespec="seconds")
    date_text = now.date().isoformat()
    with sqlite3.connect(DB_PATH) as connection:
        cursor = connection.execute(
            """
            INSERT INTO interactions (created_at, date, user_name, question, answer, sources, images)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                created_at,
                date_text,
                user_name,
                question,
                answer,
                json.dumps(source_names, ensure_ascii=False),
                json.dumps(images, ensure_ascii=False),
            ),
        )
        interaction_id = int(cursor.lastrowid)
    append_txt_report(interaction_id, created_at, date_text, user_name, question, answer, source_names, images)
    return interaction_id


def append_txt_report(
    interaction_id: int,
    created_at: str,
    date_text: str,
    user_name: str,
    question: str,
    answer: str,
    sources: list[str],
    images: list[dict],
) -> None:
    TXT_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = TXT_REPORT_DIR / f"informe_{date_text}.txt"
    image_names = [str(image.get("name", "")) for image in images if image.get("name")]
    block = [
        "=" * 90,
        f"ID: {interaction_id}",
        f"Fecha/hora: {created_at}",
        f"Usuario: {user_name}",
        "",
        "Pregunta:",
        question,
        "",
        "Respuesta:",
        answer,
        "",
        f"Fuentes internas registradas: {', '.join(sources) if sources else 'Sin fuentes registradas'}",
        f"Imagenes: {', '.join(image_names) if image_names else 'Sin imagenes'}",
        "Calificacion: Pendiente",
        "",
    ]
    with path.open("a", encoding="utf-8") as report:
        report.write("\n".join(block))


def save_rating(interaction_id: int, rating: int, note: str = "", event_time: datetime | None = None) -> bool:
    init_db()
    rating = max(1, min(5, int(rating)))
    rated_at = (event_time or local_now()).isoformat(timespec="seconds")
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT id, created_at, date, user_name, question
            FROM interactions
            WHERE id = ?
            """,
            (interaction_id,),
        ).fetchone()
        if row is None:
            return False
        connection.execute(
            """
            UPDATE interactions
            SET rating = ?, rating_note = ?, rated_at = ?
            WHERE id = ?
            """,
            (rating, note, rated_at, interaction_id),
        )
    append_rating_txt(dict(row), rating, note, rated_at)
    return True


def append_rating_txt(row: dict, rating: int, note: str, rated_at: str) -> None:
    TXT_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = TXT_REPORT_DIR / f"informe_{row['date']}.txt"
    block = [
        "-" * 90,
        f"Calificacion registrada para ID: {row['id']}",
        f"Fecha/hora calificacion: {rated_at}",
        f"Usuario: {row['user_name']}",
        f"Pregunta: {row['question']}",
        f"Calificacion: {rating}/5",
        f"Comentario: {note or 'Sin comentario'}",
        "",
    ]
    with path.open("a", encoding="utf-8") as report:
        report.write("\n".join(block))


def filtered_interactions(start_date: str = "", end_date: str = "", user_name: str = "", limit: int | None = None) -> list[dict]:
    init_db()
    conditions = []
    params: list[object] = []
    if start_date:
        conditions.append("date >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("date <= ?")
        params.append(end_date)
    if user_name:
        conditions.append("UPPER(user_name) = UPPER(?)")
        params.append(user_name)
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    limit_sql = "LIMIT ?" if limit else ""
    if limit:
        params.append(limit)
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            f"""
            SELECT id, created_at, date, user_name, question, answer, sources, rating, rating_note, rated_at
            FROM interactions
            {where}
            ORDER BY id DESC
            {limit_sql}
            """,
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def recent_interactions(limit: int = 20, user_name: str = "") -> list[dict]:
    return filtered_interactions(user_name=user_name, limit=limit)


def recent_user_history(user_name: str, limit: int = 20) -> list[dict]:
    init_db()
    if not user_name:
        return []
    with sqlite3.connect(DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT question, answer, sources
            FROM interactions
            WHERE user_name = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (user_name, limit),
        ).fetchall()
    return [dict(row) for row in reversed(rows)]


def all_interactions(start_date: str = "", end_date: str = "", user_name: str = "") -> list[dict]:
    init_db()
    rows = filtered_interactions(start_date=start_date, end_date=end_date, user_name=user_name)
    return sorted(rows, key=lambda row: (str(row["user_name"]).lower(), str(row["created_at"])), reverse=False)


def report_overview(start_date: str = "", end_date: str = "", user_name: str = "") -> dict:
    from collections import Counter, defaultdict

    rows = filtered_interactions(start_date=start_date, end_date=end_date, user_name=user_name, limit=250)
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[str(row["user_name"])].append(row)
    summary = []
    global_words = []
    ignored_words = {
        "para", "como", "dame", "quiero", "consulta", "muestra", "sobre",
        "necesito", "puedes", "favor", "empresa", "respuesta", "informacion",
    }
    for group_user, group in sorted(grouped.items(), key=lambda item: item[0].lower()):
        ratings = [int(row["rating"]) for row in group if row["rating"]]
        questions = " ".join(str(row["question"]) for row in group).lower()
        words = [
            word
            for word in re.findall(r"[a-zA-Z0-9áéíóúÁÉÍÓÚñÑ]{4,}", questions)
            if word.lower() not in ignored_words
        ]
        global_words.extend(words)
        common = ", ".join(word for word, _ in Counter(words).most_common(4))
        summary.append(
            {
                "user_name": group_user,
                "consultas": len(group),
                "promedio": round(sum(ratings) / len(ratings), 2) if ratings else "",
                "pendientes": len(group) - len(ratings),
                "temas": common,
            }
        )
    activity = [
        {
            "user_name": group_user,
            "consultas": len(group),
            "last_seen": max(str(row["created_at"]) for row in group).replace("T", " "),
        }
        for group_user, group in grouped.items()
    ]
    activity.sort(key=lambda item: (-int(item["consultas"]), item["user_name"].lower()))
    topics = [
        {"topic": word, "count": count}
        for word, count in Counter(global_words).most_common(12)
    ]
    unanswered = [
        row
        for row in rows
        if "no encuentro esa informacion" in normalize_text_for_report(row["answer"])
        or "no pude responder" in normalize_text_for_report(row["answer"])
    ][:20]
    low_ratings = [row for row in rows if row["rating"] and int(row["rating"]) <= 2][:20]
    fp = knowledge_fingerprint()
    knowledge = {
        "files": fp["file_count"],
        "index_status": "Indice listo" if index_is_current(fp) else "Indice actualizandose o pendiente",
        "last_update": datetime.fromtimestamp(fp["newest_mtime"] / 1_000_000_000).isoformat(timespec="seconds") if fp["newest_mtime"] else "",
        "txt_report_dir": str(TXT_REPORT_DIR),
    }
    return {
        "summary": summary,
        "activity": activity[:10],
        "topics": topics,
        "unanswered": unanswered,
        "low_ratings": low_ratings,
        "knowledge": knowledge,
        "permissions": ROLE_PERMISSIONS,
        "audit": recent_audit_events(),
        "admin_users": [
            {
                "username": user["username"],
                "role": user["role"],
                "role_label": ROLE_LABELS.get(user["role"], "Invitado"),
                "active": bool(user["active"]),
                "status": "Activo" if user["active"] else "Desactivado",
            }
            for user in list_app_users()
        ],
    }

def normalize_text_for_report(text: str) -> str:
    normalized = unicodedata.normalize("NFD", str(text or "").lower())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def report_filters_from_path(path: str) -> tuple[str, str, str]:
    query = parse_qs(urlparse(path).query)
    return (
        str(query.get("start", [""])[0]).strip()[:10],
        str(query.get("end", [""])[0]).strip()[:10],
        str(query.get("user", [""])[0]).strip()[:80],
    )


def create_report_xlsx(start_date: str = "", end_date: str = "", user_name: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as error:
        raise UserFacingError("Falta openpyxl para generar el reporte Excel.") from error

    from io import BytesIO
    from collections import Counter

    rows = all_interactions(start_date=start_date, end_date=end_date, user_name=user_name)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    detail = workbook.create_sheet("Detalle")
    unanswered_sheet = workbook.create_sheet("Sin respuesta")
    improve_sheet = workbook.create_sheet("Mejorar respuestas")
    activity_sheet = workbook.create_sheet("Actividad y temas")
    audit_sheet = workbook.create_sheet("Auditoria")
    config_sheet = workbook.create_sheet("Base y usuarios")
    header_fill = PatternFill("solid", fgColor="DFF4EA")

    summary.append(["Usuario", "Fecha", "Consultas", "Promedio calificacion", "Pendientes"])
    counts = Counter((row["user_name"], row["date"]) for row in rows)
    for (user_name, date_text), count in sorted(counts.items(), key=lambda item: (item[0][0].lower(), item[0][1])):
        group = [row for row in rows if row["user_name"] == user_name and row["date"] == date_text]
        ratings = [int(row["rating"]) for row in group if row["rating"]]
        average = round(sum(ratings) / len(ratings), 2) if ratings else "Sin calificar"
        pending = count - len(ratings)
        summary.append([user_name, date_text, count, average, pending])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="00665C")
        cell.fill = header_fill
    for index, width in enumerate([28, 18, 14, 22, 14], start=1):
        summary.column_dimensions[get_column_letter(index)].width = width

    headers = ["Usuario", "Fecha", "Hora", "Pregunta", "Respuesta", "Calificacion", "Comentario calificacion", "Calificado en", "Fuentes internas"]
    detail.append(headers)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="00665C")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        try:
            sources = ", ".join(json.loads(row["sources"] or "[]"))
        except json.JSONDecodeError:
            sources = row["sources"] or ""
        created = str(row["created_at"])
        time_text = created.split("T", 1)[1] if "T" in created else created
        rating = f"{row['rating']}/5" if row["rating"] else "Pendiente"
        detail.append([row["user_name"], row["date"], time_text, row["question"], row["answer"], rating, row["rating_note"] or "", row["rated_at"] or "", sources])

    widths = [24, 16, 14, 48, 86, 16, 36, 22, 52]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.freeze_panes = "A2"
    summary.freeze_panes = "A2"
    for sheet in (summary, detail):
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in detail["A"]:
        cell.font = Font(bold=cell.row == 1)

    unanswered_rows = [
        row
        for row in rows
        if "no encuentro esa informacion" in normalize_text_for_report(row["answer"])
        or "no pude responder" in normalize_text_for_report(row["answer"])
    ]
    unanswered_sheet.append(["Usuario", "Fecha", "Pregunta", "Respuesta"])
    for row in unanswered_rows:
        unanswered_sheet.append([row["user_name"], row["created_at"], row["question"], row["answer"]])

    improve_sheet.append(["Usuario", "Fecha", "Pregunta", "Respuesta", "Calificacion", "Comentario"])
    for row in rows:
        if row["rating"] and int(row["rating"]) <= 2:
            improve_sheet.append([row["user_name"], row["created_at"], row["question"], row["answer"], row["rating"], row["rating_note"] or ""])

    overview = report_overview(start_date=start_date, end_date=end_date, user_name=user_name)
    activity_sheet.append(["Tipo", "Dato", "Detalle"])
    for item in overview["activity"]:
        activity_sheet.append(["Usuario activo", item["user_name"], f"{item['consultas']} consultas | ultima actividad: {item['last_seen']}"])
    for item in overview["topics"]:
        activity_sheet.append(["Tema frecuente", item["topic"], f"{item['count']} menciones"])

    audit_sheet.append(["Fecha", "Actor", "Accion", "Usuario objetivo", "Detalle"])
    for item in overview["audit"]:
        audit_sheet.append([item["created_at"], item["actor"], item["action_label"], item.get("target_user", ""), item.get("detail", "")])

    fp = knowledge_fingerprint()
    config_sheet.append(["Dato", "Valor"])
    config_sheet.append(["Archivos detectados", fp["file_count"]])
    config_sheet.append(["Estado indice", "Indice listo" if index_is_current(fp) else "Indice actualizandose o pendiente"])
    config_sheet.append(["Carpeta reportes TXT", str(TXT_REPORT_DIR)])
    config_sheet.append(["Filtro usuario", user_name or "Todos"])
    config_sheet.append(
        [
            "Usuarios configurados",
            ", ".join(
                f"{user['username']} ({ROLE_LABELS.get(user['role'], 'Invitado')}, {'activo' if user['active'] else 'desactivado'})"
                for user in list_app_users()
            ),
        ]
    )

    for sheet in (unanswered_sheet, improve_sheet, activity_sheet, audit_sheet, config_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="00665C")
            cell.fill = header_fill
        for index, width in enumerate([24, 22, 58, 86, 16, 36], start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def pdf_escape(text: str) -> str:
    return str(text or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def wrap_pdf_text(text: str, width: int = 95, max_lines: int = 6) -> list[str]:
    words = " ".join(str(text or "").split()).split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= width:
            current = candidate
            continue
        if current:
            lines.append(current)
        current = word
        if len(lines) >= max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    if len(lines) == max_lines and len(" ".join(words)) > len(" ".join(lines)):
        lines[-1] = lines[-1].rstrip(".") + "..."
    return lines or [""]


def create_report_pdf(start_date: str = "", end_date: str = "", user_name: str = "") -> bytes:
    rows = all_interactions(start_date=start_date, end_date=end_date, user_name=user_name)
    now = local_now().strftime("%Y-%m-%d %H:%M")
    pages: list[list[str]] = []
    current: list[str] = []
    y = 800

    def new_page() -> None:
        nonlocal current, y
        if current:
            pages.append(current)
        current = [
            "BT /F1 16 Tf 50 810 Td (Reporte de consultas - Asistente MADEVAL) Tj ET",
            f"BT /F1 9 Tf 50 794 Td (Generado: {pdf_escape(now)}) Tj ET",
        ]
        if start_date or end_date:
            current.append(f"BT /F1 9 Tf 50 780 Td (Rango: {pdf_escape(start_date or 'inicio')} a {pdf_escape(end_date or 'hoy')}) Tj ET")
            y = 760
        else:
            y = 770
        if user_name:
            current.append(f"BT /F1 9 Tf 50 {y} Td (Usuario: {pdf_escape(user_name)}) Tj ET")
            y -= 14

    def add_line(text: str, size: int = 9, indent: int = 50, gap: int = 13) -> None:
        nonlocal y
        if y < 58:
            new_page()
        current.append(f"BT /F1 {size} Tf {indent} {y} Td ({pdf_escape(text)}) Tj ET")
        y -= gap

    new_page()
    if not rows:
        add_line("Todavia no hay consultas registradas.")
    for row in rows:
        created = str(row["created_at"]).replace("T", " ")
        rating = f"{row['rating']}/5" if row["rating"] else "Pendiente"
        add_line(f"Usuario: {row['user_name']} | Fecha: {created} | Calificacion: {rating}", size=10, gap=15)
        add_line("Pregunta:", size=9, gap=12)
        for line in wrap_pdf_text(row["question"], width=100, max_lines=3):
            add_line(line, indent=64, gap=12)
        add_line("Respuesta:", size=9, gap=12)
        for line in wrap_pdf_text(row["answer"], width=100, max_lines=6):
            add_line(line, indent=64, gap=12)
        add_line("-" * 112, size=8, gap=14)
    if current:
        pages.append(current)

    objects = ["<< /Type /Catalog /Pages 2 0 R >>"]
    kids = []
    font_object_id = 3 + len(pages) * 2
    for index, lines in enumerate(pages):
        page_id = 3 + index * 2
        content_id = page_id + 1
        kids.append(f"{page_id} 0 R")
        stream = "\n".join(lines).encode("latin-1", errors="replace")
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] "
            f"/Resources << /Font << /F1 {font_object_id} 0 R >> >> /Contents {content_id} 0 R >>"
        )
        objects.append(f"<< /Length {len(stream)} >>\nstream\n{stream.decode('latin-1')}\nendstream")
    objects.insert(1, f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(pages)} >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    body = bytearray(b"%PDF-1.4\n")
    offsets = []
    for object_id, payload in enumerate(objects, start=1):
        offsets.append(len(body))
        body.extend(f"{object_id} 0 obj\n{payload}\nendobj\n".encode("latin-1", errors="replace"))
    xref_start = len(body)
    body.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
    for offset in offsets:
        body.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    body.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode("ascii")
    )
    return bytes(body)


def list_files() -> list[dict]:
    KNOWLEDGE_DIR.mkdir(exist_ok=True)
    files = []
    for path in sorted(KNOWLEDGE_DIR.rglob("*")):
        if not path.is_file() or path.name.startswith(".") or path.name.startswith("~$") or path.name == "urls.txt":
            continue
        relative_path = path.relative_to(KNOWLEDGE_DIR)
        files.append(
            {
                "name": path.name,
                "path": str(relative_path).replace("\\", "/"),
                "size": path.stat().st_size,
                "extension": path.suffix.lower() or "archivo",
                "is_image": path.suffix.lower() in IMAGE_FILE_TYPES,
            }
        )
    return files


def safe_relative_path(name: str) -> Path | None:
    raw_parts = unquote(name).replace("\\", "/").split("/")
    safe_parts = []
    for part in raw_parts:
        cleaned = "".join(char for char in part.strip() if char.isalnum() or char in "._- ")
        if not cleaned or cleaned in {".", ".."}:
            continue
        safe_parts.append(cleaned)
    if not safe_parts:
        return None
    return Path(*safe_parts)


def admin_token_from(handler: BaseHTTPRequestHandler) -> str:
    auth_header = handler.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header.removeprefix("Bearer ").strip()
    query = parse_qs(urlparse(handler.path).query)
    return str(query.get("token", [""])[0]).strip()


def current_session(handler: BaseHTTPRequestHandler) -> dict:
    token = admin_token_from(handler)
    session = ADMIN_SESSIONS.get(token, {}) if token else {}
    username = session.get("username")
    if not username:
        return {}
    user = get_app_user(str(username))
    if not user or not user["active"]:
        ADMIN_SESSIONS.pop(token, None)
        return {}
    return {"username": user["username"], "role": user["role"], "active": bool(user["active"])}


def is_super_admin_request(handler: BaseHTTPRequestHandler) -> bool:
    session = current_session(handler)
    return bool(session.get("active")) and session.get("role") == "super_usuario"


def is_admin_request(handler: BaseHTTPRequestHandler) -> bool:
    return is_super_admin_request(handler)


def read_json_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length)
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    return json.loads(text or "{}")


def knowledge_fingerprint() -> dict:
    KNOWLEDGE_DIR.mkdir(exist_ok=True)
    digest = hashlib.sha256()
    file_count = 0
    total_size = 0
    newest_mtime = 0
    for path in sorted(KNOWLEDGE_DIR.rglob("*")):
        if not path.is_file() or path.name.startswith(".") or path.name.startswith("~$") or path.name == "urls.txt":
            continue
        if path.suffix.lower() not in ALLOWED_EXTENSIONS:
            continue
        stat = path.stat()
        relative = str(path.relative_to(KNOWLEDGE_DIR)).replace("\\", "/")
        digest.update(f"{relative}|{stat.st_size}|{stat.st_mtime_ns}\n".encode("utf-8", errors="ignore"))
        file_count += 1
        total_size += stat.st_size
        newest_mtime = max(newest_mtime, stat.st_mtime_ns)
    return {
        "signature": digest.hexdigest(),
        "file_count": file_count,
        "total_size": total_size,
        "newest_mtime": newest_mtime,
    }


def index_is_current(fingerprint: dict) -> bool:
    if not fingerprint["file_count"] or not INDEX_PATH.exists():
        return False
    if not INDEX_META_PATH.exists():
        return INDEX_PATH.stat().st_mtime_ns >= int(fingerprint.get("newest_mtime") or 0)
    try:
        current = json.loads(INDEX_META_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return INDEX_PATH.stat().st_mtime_ns >= int(fingerprint.get("newest_mtime") or 0)
    return current.get("signature") == fingerprint.get("signature")


def write_index_metadata(index: dict, fingerprint: dict) -> None:
    INDEX_META_PATH.parent.mkdir(exist_ok=True)
    payload = {
        **fingerprint,
        "chunks": len(index.get("chunks", [])),
        "updated_at": local_now().isoformat(timespec="seconds"),
    }
    INDEX_META_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def ensure_index_in_background(force: bool = False) -> None:
    fingerprint = knowledge_fingerprint()
    if not force and index_is_current(fingerprint):
        return
    if not fingerprint["file_count"] or INDEX_BUILD_STATE["running"]:
        return

    def rebuild_index() -> None:
        INDEX_BUILD_STATE["running"] = True
        try:
            index = build_index()
            write_index_metadata(index, fingerprint)
            print(f"Indice actualizado automaticamente con {len(index['chunks'])} fragmentos.")
        except Exception as error:
            print(f"No se pudo actualizar el indice automaticamente: {error}")
        finally:
            INDEX_BUILD_STATE["running"] = False

    threading.Thread(target=rebuild_index, daemon=True).start()


class AppHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        if self.path in {"/", "/index.html"}:
            page_response(self)
            return
        if self.path == "/api/files":
            json_response(self, {"files": list_files(), "index_exists": INDEX_PATH.exists()})
            return
        if self.path.startswith("/api/report.xlsx"):
            self.handle_report_xlsx()
            return
        if self.path.startswith("/api/report.pdf"):
            self.handle_report_pdf()
            return
        if urlparse(self.path).path == "/api/report":
            if not is_admin_request(self):
                json_response(self, {"ok": False, "message": "Acceso restringido."}, status=403)
                return
            start_date, end_date, user_name = report_filters_from_path(self.path)
            json_response(
                self,
                {
                    "interactions": filtered_interactions(start_date=start_date, end_date=end_date, user_name=user_name, limit=20),
                    **report_overview(start_date=start_date, end_date=end_date, user_name=user_name),
                },
            )
            return
        if self.path.startswith("/assets/"):
            self.handle_asset_file()
            return
        if self.path.startswith("/knowledge/"):
            self.handle_knowledge_file()
            return
        json_response(self, {"ok": False, "message": "Ruta no encontrada."}, status=404)

    def handle_report_xlsx(self) -> None:
        if not is_admin_request(self):
            json_response(self, {"ok": False, "message": "Acceso restringido."}, status=403)
            return
        try:
            start_date, end_date, user_name = report_filters_from_path(self.path)
            body = create_report_xlsx(start_date=start_date, end_date=end_date, user_name=user_name)
        except UserFacingError as error:
            json_response(self, {"ok": False, "message": str(error)}, status=500)
            return
        filename = f"reporte_consultas_{local_now().date().isoformat()}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_report_pdf(self) -> None:
        if not is_admin_request(self):
            json_response(self, {"ok": False, "message": "Acceso restringido."}, status=403)
            return
        start_date, end_date, user_name = report_filters_from_path(self.path)
        body = create_report_pdf(start_date=start_date, end_date=end_date, user_name=user_name)
        filename = f"reporte_consultas_{local_now().date().isoformat()}.pdf"
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_knowledge_file(self) -> None:
        relative = safe_relative_path(self.path.removeprefix("/knowledge/"))
        if relative is None:
            json_response(self, {"ok": False, "message": "Archivo no encontrado."}, status=404)
            return
        target = (KNOWLEDGE_DIR / relative).resolve()
        root = KNOWLEDGE_DIR.resolve()
        if root not in target.parents or not target.exists() or not target.is_file():
            json_response(self, {"ok": False, "message": "Archivo no encontrado."}, status=404)
            return
        mime_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        body = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_asset_file(self) -> None:
        relative = safe_relative_path(self.path.removeprefix("/assets/"))
        if relative is None:
            json_response(self, {"ok": False, "message": "Archivo no encontrado."}, status=404)
            return
        target = (ASSETS_DIR / relative).resolve()
        root = ASSETS_DIR.resolve()
        if root not in target.parents or not target.exists() or not target.is_file():
            json_response(self, {"ok": False, "message": "Archivo no encontrado."}, status=404)
            return
        mime_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        body = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self) -> None:
        if self.path == "/api/user/login":
            self.handle_user_login()
            return
        if self.path == "/api/admin/login":
            self.handle_admin_login()
            return
        if self.path == "/api/admin/logout":
            self.handle_admin_logout()
            return
        if self.path == "/api/admin/users":
            self.handle_admin_user_save()
            return
        if self.path == "/api/ingest":
            self.handle_ingest()
            return
        if self.path == "/api/ask":
            self.handle_ask()
            return
        if self.path == "/api/rate":
            self.handle_rate()
            return
        json_response(self, {"ok": False, "message": "Ruta no encontrada."}, status=404)

    def handle_user_login(self) -> None:
        payload = read_json_body(self)
        event_time = client_now(payload)
        client_timezone = str(payload.get("client_timezone", "")).strip()
        username = str(payload.get("username", "")).strip()
        password = str(payload.get("password", ""))
        user = get_app_user(username)
        if user and not verify_password(password, user["password_hash"]) and legacy_admin_password_ok(user["username"], password):
            upsert_app_user(user["username"], password=password, role=user["role"], active=bool(user["active"]), event_time=event_time)
            user = get_app_user(username)
        if not user or not user["active"] or not verify_password(password, user["password_hash"]):
            json_response(self, {"ok": False, "message": "Usuario o contrasena incorrectos."}, status=401)
            return
        detail = ROLE_LABELS.get(user["role"], user["role"])
        if client_timezone:
            detail += f"; zona: {client_timezone}"
        log_audit_event(user["username"], "user_login", user["username"], detail, event_time=event_time)
        json_response(
            self,
            {
                "ok": True,
                "username": user["username"],
                "role": user["role"],
                "role_label": ROLE_LABELS.get(user["role"], "Invitado"),
            },
        )

    def handle_admin_login(self) -> None:
        payload = read_json_body(self)
        event_time = client_now(payload)
        client_timezone = str(payload.get("client_timezone", "")).strip()
        username = str(payload.get("username", "")).strip()
        password = str(payload.get("password", ""))
        user = get_app_user(username)
        if user and not verify_password(password, user["password_hash"]) and legacy_admin_password_ok(user["username"], password):
            upsert_app_user(user["username"], password=password, role=user["role"], active=bool(user["active"]), event_time=event_time)
            user = get_app_user(username)
        if not user or not user["active"] or not verify_password(password, user["password_hash"]):
            json_response(self, {"ok": False, "message": "Usuario o contrasena incorrectos."}, status=401)
            return
        if user["role"] != "super_usuario":
            json_response(self, {"ok": False, "message": "Este usuario no tiene acceso a super usuario."}, status=403)
            return
        token = secrets.token_urlsafe(32)
        ADMIN_SESSIONS[token] = {"username": user["username"], "role": user["role"], "active": bool(user["active"])}
        detail = "Ingreso al panel de super usuario"
        if client_timezone:
            detail += f"; zona: {client_timezone}"
        log_audit_event(user["username"], "admin_login", user["username"], detail, event_time=event_time)
        json_response(self, {"ok": True, "token": token, "role": user["role"], "username": user["username"]})

    def handle_admin_logout(self) -> None:
        payload = read_json_body(self)
        event_time = client_now(payload)
        client_timezone = str(payload.get("client_timezone", "")).strip()
        token = admin_token_from(self)
        session = current_session(self)
        if token:
            ADMIN_SESSIONS.pop(token, None)
        if session.get("username"):
            detail = "Cierre del panel de super usuario"
            if client_timezone:
                detail += f"; zona: {client_timezone}"
            log_audit_event(session["username"], "admin_logout", session["username"], detail, event_time=event_time)
        json_response(self, {"ok": True})

    def handle_admin_user_save(self) -> None:
        if not is_super_admin_request(self):
            json_response(self, {"ok": False, "message": "Acceso restringido."}, status=403)
            return
        payload = read_json_body(self)
        event_time = client_now(payload)
        client_timezone = str(payload.get("client_timezone", "")).strip()
        session = current_session(self)
        try:
            user = upsert_app_user(
                username=str(payload.get("username", "")).strip(),
                password=str(payload.get("password", "")),
                role=str(payload.get("role", "invitado")),
                active=parse_active_flag(payload.get("active", True)),
                event_time=event_time,
            )
        except UserFacingError as error:
            json_response(self, {"ok": False, "message": str(error)}, status=400)
            return
        detail = f"Rol: {ROLE_LABELS.get(user.get('role'), user.get('role'))}; estado: {'activo' if user.get('active') else 'desactivado'}"
        if str(payload.get("password", "")):
            detail += "; contrasena actualizada"
        if client_timezone:
            detail += f"; zona: {client_timezone}"
        log_audit_event(session.get("username", "super_usuario"), "user_save", str(user.get("username", "")), detail, event_time=event_time)
        json_response(
            self,
            {
                "ok": True,
                "user": {
                    "username": user.get("username"),
                    "role": user.get("role"),
                    "active": bool(user.get("active")),
                },
            },
        )

    def handle_ingest(self) -> None:
        try:
            fingerprint = knowledge_fingerprint()
            index = build_index()
            write_index_metadata(index, fingerprint)
        except UserFacingError as error:
            json_response(self, {"ok": False, "message": str(error)}, status=400)
            return
        except Exception as error:
            json_response(self, {"ok": False, "message": f"No pude crear el indice. Revisa la configuracion e intenta otra vez. Detalle: {error}"}, status=500)
            return
        json_response(self, {"ok": True, "message": f"Indice creado con {len(index['chunks'])} fragmentos."})

    def handle_ask(self) -> None:
        payload = read_json_body(self)
        event_time = client_now(payload)
        question = str(payload.get("question", "")).strip()
        user_name = str(payload.get("user_name", "")).strip()
        if not user_name:
            json_response(self, {"ok": False, "message": "Ingresa tu nombre para iniciar."}, status=400)
            return
        if not question:
            json_response(self, {"ok": False, "message": "Escribe una pregunta."}, status=400)
            return
        try:
            history = recent_user_history(user_name)
            result = answer_question_with_sources(question, user_name=user_name, history=history)
        except UserFacingError as error:
            json_response(self, {"ok": False, "message": str(error)}, status=400)
            return
        except Exception as error:
            json_response(self, {"ok": False, "message": f"No pude responder todavia. Revisa la configuracion e intenta otra vez. Detalle: {error}"}, status=500)
            return
        image_candidates = []
        seen = set()
        combined_text = f"{question} {result['answer']}".lower()
        for source in result["sources"]:
            source_path = str(source.get("source", ""))
            extension = Path(source_path).suffix.lower()
            if extension not in IMAGE_FILE_TYPES or source_path in seen:
                continue
            seen.add(source_path)
            normalized_source = source_path.replace("\\", "/")
            relative_source = normalized_source.removeprefix("knowledge_base/")
            image_candidates.append(
                {
                    "name": Path(source_path).name,
                    "path": source_path,
                    "url": "/knowledge/" + quote(relative_source, safe="/"),
                }
            )
        matched_images = [
            image
            for image in image_candidates
            if any(token in combined_text for token in Path(image["name"]).stem.lower().replace("_", " ").split() if len(token) > 2)
        ]
        images = matched_images or image_candidates
        images = images[:4]
        interaction_id = log_interaction(user_name, question, result["answer"], result["sources"], images, event_time=event_time)
        json_response(self, {"ok": True, "answer": result["answer"], "images": images, "interaction_id": interaction_id})

    def handle_rate(self) -> None:
        payload = read_json_body(self)
        event_time = client_now(payload)
        try:
            interaction_id = int(payload.get("interaction_id", 0))
            rating = int(payload.get("rating", 0))
        except (TypeError, ValueError):
            json_response(self, {"ok": False, "message": "Calificacion invalida."}, status=400)
            return
        note = str(payload.get("note", "")).strip()
        if interaction_id <= 0 or rating < 1 or rating > 5:
            json_response(self, {"ok": False, "message": "Calificacion invalida."}, status=400)
            return
        if not save_rating(interaction_id, rating, note, event_time=event_time):
            json_response(self, {"ok": False, "message": "No encontre la respuesta a calificar."}, status=404)
            return
        json_response(self, {"ok": True})

    def log_message(self, format: str, *args) -> None:
        return


def run() -> None:
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    ensure_index_in_background()
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Interfaz lista en http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    run()

